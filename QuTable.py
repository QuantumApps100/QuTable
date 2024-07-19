#!/usr/bin/python3
# coding: utf-8

# from tkinter import Scrollbar as tk_Scrollbar, Toplevel as tk_Toplevel, Button as tk_Button, Menu as tk_Menu, Label as tk_Label, Entry as tk_Entry, Frame as tk_Frame, Text as tk_Text, Tk as tk_Tk, Button as tk_Button, NONE as tk_NONE, HORIZONTAL as tk_HORIZONTAL, PanedWindow as tk_PanedWindow

# pyinstaller --exclude-module tensorflow -n QuTable QuTable-OpenPyxl.py
# pyinstaller --exclude-module tensorflow -n QuTable --onefile QuTable-OpenPyxl.py
# cmd /c "pyinstaller --exclude-module tensorflow -n QuTable --distpath "C:\Program Files\QuTable" QuTable-OpenPyxl.py"

# runas /savecred /user:Administrator cmd /c "pyinstaller --exclude-module tensorflow -n QuTable --distpath "C:\Program Files\QuTable" QuTable-OpenPyxl.py"
# runas /savecred /user:Administrator QuTable-Pyins.bat

from ctypes import windll as ctypes_windll
ctypes_windll.shcore.SetProcessDpiAwareness(1)

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

from PIL import Image, ImageTk
import matplotlib.pyplot as plt

import tkinter as tk
import tkinter.font as tkFont
from tkinter import ttk, filedialog, messagebox, colorchooser

from sympy import srepr
from sympy.parsing import sympy_parser
from sympy.parsing.sympy_parser import (standard_transformations, implicit_multiplication_application)
parse_expr = sympy_parser.parse_expr

import logging
from traceback import format_exc

import threading
import multiprocessing
from concurrent.futures import ThreadPoolExecutor

from decimal import Decimal, InvalidOperation

import numpy as np
import numbers
import functools
import subprocess
import datetime
import re
import os
import sys
import json
import time
import string
import statistics
import pandas as pd
import requests as rq
import zlib
import gzip
import base64

import klembord
klembord.init()

from QuTableExit import QuTableExitContent
from QuTable_icons import imageDict

# Configure logging to both console and file
# logging.basicConfig(filename='app.log', level=logging.DEBUG)

### my broken solution:
# so = se = open("a.log", 'w', 0)
# sys.stdout = os.fdopen(sys.stdout.fileno(), 'w', 0)

# os.dup2(sys.stdout.fileno(), so.fileno())
# os.dup2(sys.stderr.fileno(), se.fileno())

# Redirect stdout and stderr to a file
# sys.stdout = open("stdout.log", "w")
# sys.stderr = open("stderr.log", "w")

class LoadingApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Warming up...")
        self.configure(bg='#FFFFFF')
        self.option_add('*background', '#FFFFFF')
        self.label = tk.Label(self, text="Welcome to QuTable!\nby Quantum Apps\n\nThe Quantum-Inspired Virtual Device is warming up.\nPlease kindly wait and be patient...", font='Arial 12')
        self.label.pack(padx=20, pady=20, anchor="center", expand=True, fill="both")
        self.after(200, self.task)
        self.center_window(width=700, height=250)
        
        self.icon = ImageTk.PhotoImage(image=imageDict['QuTableIcon'] , master=self )
        self.iconphoto(True, self.icon)

    def task(self):
        global QBO, licensePage
        import QBasicOperation as QBO
        self.destroy()
        licensePage = LicensePage()
        
    def center_window(window, width, height):
        # Get the screen width and height
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()

        # Calculate the position of the window to center it on the screen
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

        # Set the geometry of the window to center it on the screen
        window.geometry(f"{width}x{height}+{x}+{y}")

def is_browser_only_link(url):
    try:
        response = rq.head(url)
        content_type = response.headers.get('Content-Type')
        if content_type:
            if orExec('text/html' in content_type, 'image' in content_type):
                return True
            else:
                return False
        else:
            return False
    except rq.exceptions.RequestException as e:
        return False

experimentalOverall = False

keys = string.ascii_letters+string.digits

originalBorderColor = '#ffffff'
selectBorderColor = '#000000'

pixels_per_point = 10.5

vertical_list_standard = [0, 1]
to_end_list, vertical_list = [0, 0, 1, 1], vertical_list_standard*2

Integer = int
Float = lambda num, precision: float(num)

app = None

def evalDefine(cmd):
    return eval(cmd, {self.sheetName:self})

class HScrollBar:
    def __init__(self, parent, target=None, pack=True, *args, **kwargs):
        self.Bar = tk.Scrollbar(parent, orient='horizontal', command=self.xview, *args, **kwargs)
        self.install() if pack else None
        self.uninstall = self.Bar.forget
        self.parent = parent
        self.target = target if isinstance(target, tk.Text) else parent
        # print(f'HScrollBar target = {target}')
        # print(f'HScrollBar bool(target) = {bool(target)}')
        # print(f'HScrollBar self.target = {self.target}')
        self.target['xscrollcommand'] = self.Bar.set
        
    def install(self, *args, **kwargs):
        self.Bar.pack(side='bottom', fill='x', *args, **kwargs)
        
    def xview(self, *args, **kwargs):
        self.target.xview(*args, **kwargs)
        # print('xview =', args)


class VScrollBar:
    def __init__(self, parent, target=None, pack=True, *args, **kwargs):
        self.Bar = tk.Scrollbar(parent, orient='vertical', command=self.yview, *args, **kwargs)
        self.install() if pack else None
        self.uninstall = self.Bar.forget
        self.parent = parent
        self.target = target if isinstance(target, tk.Text) else parent
        # print(f'VScrollBar target = {target}')
        # print(f'VScrollBar bool(target) = {bool(target)}')
        # print(f'VScrollBar self.target = {self.target}')
        self.target['yscrollcommand'] = self.Bar.set
        
    def install(self, *args, **kwargs):
        self.Bar.pack(side='right', fill='y', *args, **kwargs)
        
    def yview(self, *args, **kwargs):
        self.target.yview(*args, **kwargs)
        # print('yview =', args)


class BothScrollBar:
    def __init__(self, h_parent, v_parent, target=None, pack=False, *args, **kwargs):
        self.HBarObj = HScrollBar(parent=h_parent, target=target, pack=pack, *args, **kwargs)
        self.VBarObj = VScrollBar(parent=v_parent, target=target, pack=pack, *args, **kwargs)
        self.HBar = self.HBarObj.Bar
        self.VBar = self.VBarObj.Bar
        
    def install(self, *args, **kwargs):
        self.HBarObj.install(*args, **kwargs)
        self.VBarObj.install(*args, **kwargs)

    def uninstall(self, *args, **kwargs):
        self.HBarObj.uninstall(*args, **kwargs)
        self.VBarObj.uninstall(*args, **kwargs)

def translate_formula(formula, offset):
    # Regular expression to match cell references
    pattern = r'([A-Z]+)([0-9]+)'
    
    def replace(match):
        col, row = match.groups()
        return col + str(int(row) + offset)
    
    # Replace cell references with adjusted references
    translated_formula = re.sub(pattern, replace, formula)
    return translated_formula

def is_tk_widget(obj):
    return isinstance(obj, (tk.Widget, tk.Misc))

def createVarList(expr_str):
    return eval( srepr( parse_expr( expr_str.strip('=') ) ) )

def detectQuantumMode():
    if not app:
        # return True
        return False
    for book in app:
        if book.window.focus_get():
            return book.quantum_mode

def orPairExec(n1, n2):
    if 0: #detectQuantumMode():
        return QBO.or_gate(n1, n2)
    else:
        return n1 or n2
        
def andPairExec(n1, n2):
    if 0: #detectQuantumMode():
        return QBO.and_gate(n1, n2)
    else:
        return n1 and n2

def orExec(*args): return functools.reduce(orPairExec, args)
def andExec(*args): return functools.reduce(andPairExec, args)

def float_int_convert(string, converter=str):
    if not converter:
        converter = lambda x: x
    if is_numeric(string):
        float_string = float(string)
        int_string = int(float_string)
        if float(string) == int_string:
            return converter(int_string).lstrip('0') if converter == str else converter(int_string)
        else:
            return converter(float_string).strip('0') if converter == str else converter(float_string)
    else:
        return converter(string).strip('0') if converter == str else converter(string)

def properize(value, shift=1):
    is_float_cond = '.' in value if andExec(is_numeric(value), type(value) == str) else type(value) == float
    if is_float_cond:
        part1, part2 = str(value).split('.')
        setOf0s = "0"*(len(part2)-shift)
        # print(setOf0s)
        original_num = Decimal(value)
        proper_num1 = original_num.quantize(Decimal(f'0.{setOf0s}1'))
        proper_num2 = float_int_convert(proper_num1)
        return str(proper_num2)
    else:
        return str(value)

def Symbol(data):  return data
def Add(*data):    return data
def Mul(*data):    return data
def Pow(*data):    return data
    
# QuTable code handling
def handleIndex(colCodeWIndex, book=None, sheet=None):
    if colCodeWIndex.startswith('C'):
        if 'R' in colCodeWIndex:
            indices1 = re.split('[CR]', colCodeWIndex)[1:]
            indices2 = tuple( int(index)-1 for index in indices1 )
            if sheet:
                return sheet[indices2]
            for book in app:
                if book.window.focus_get():
                    return book.sheet[indices2]
        else:
            indices1 = colCodeWIndex.split('C')[1]
            indices2 = int(indices1)-1
            if sheet:
                return sheet[indices2]
            for book in app:
                if orExec(book.window.focus_get(), book.columnFormulaWidget.focus_get()):
                    return book.sheet[indices2]
    
# def Add(n1,n2): return float_int_convert( Decimal(str(n1))+Decimal(str(n2)) )
    
# def Mul(n1,n2):
    # if andExec(is_numeric(n1), is_numeric(n2)):
        # return float_int_convert( Decimal(str(n1))*Decimal(str(n2)) )
    # elif is_numeric(n1):
        # return functools.reduce(lambda n2elem: n1*n2elem, n2)
    # elif is_numeric(n2):
        # return functools.reduce(lambda n1elem: n2*n1elem, n1)
    # else:
        # return n1*n2
        # # result = []
        # # for n1elem, n2elem in zip(n1, n2):
            # # result.append( float(n1elem)*float(n2elem) )
        # # return result

# def Pow(n1,n2): return float_int_convert( Decimal(str(n1))**Decimal(str(n2)) )

def remove(iterable, obj):
    List = list(iterable)
    List.remove(obj)
    return iterable.__class__(List)

def hex_to_rgb(hex_color):
    """
    Convert hexadecimal color code to RGB tuple.
    """
    # Remove '#' if present
    hex_color = hex_color.lstrip('#')
    # Convert hex to integer
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    return rgb

def hex_to_rgba(hex_color):
    """
    Convert hexadecimal color code to RGB tuple.
    """
    # Remove '#' if present
    hex_color = hex_color.lstrip('#')
    # Convert hex to integer
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    return rgb + (255,)

def rgb_to_hex(rgb, context):
    """
    Convert RGB tuple to hexadecimal color code.
    """
    if not rgb:
        if context == 'fg':
            return '#000000'
        elif context == 'bg':
            return '#ffffff'
    # Ensure the RGB values are integers within the valid range (0-255)
    r, g, b = [int(x) for x in rgb]
    # Convert RGB values to hexadecimal format and concatenate them
    hex_code = "#{:02x}{:02x}{:02x}".format(r, g, b)
    return hex_code

def mix_colors(hex_codes, context='bg'):
    # Convert hex codes to RGB tuples
    rgb_colors = [hex_to_rgb(hex_code) for hex_code in hex_codes]
    
    # Calculate average RGB values
    avg_color = [sum(color_channel) // len(hex_codes) for color_channel in zip(*rgb_colors)]
    
    # Convert average RGB to hex code
    avg_hex = rgb_to_hex(avg_color, context)
    
    return avg_hex

def convert_to_css(tk_font_str):
    tk_font = tkFont.Font(font=tk_font_str)
    css_font = ""
    
    # Font family
    families = tk_font.actual()['family']
    if isinstance(families, str):
        families = [families]
    css_font += "font-family: {}; ".format(", ".join(families))
    
    # Font size
    size = tk_font.actual()['size']
    css_font += "font-size: {}pt; ".format(size)
    
    # Font weight
    weight = tk_font.actual()['weight']
    if weight == "bold":
        css_font += "font-weight: bold; "
    else:
        css_font += "font-weight: normal; "
    
    # Italic
    slant = tk_font.actual()['slant']
    if slant == "italic":
        css_font += "font-style: italic; "
    else:
        css_font += "font-style: normal; "
    
    # Underline
    underline = tk_font.actual()['underline']
    if underline:
        css_font += "text-decoration: underline; "
    else:
        css_font += "text-decoration: none; "
    
    return css_font

def font_to_dict(openpyxl_font):
    font = openpyxl_font
    return {
        "name": font.name,
        "charset": font.charset,
        "family": font.family,
        "bold": font.b,
        "italic": font.i,
        "strike": font.strike,
        "outline": font.outline,
        "shadow": font.shadow,
        "condense": font.condense,
        "color": {
            "rgb": font.color.rgb,
            "indexed": font.color.indexed,
            "auto": font.color.auto,
            "theme": font.color.theme,
            "tint": font.color.tint,
            "type": font.color.type,
            "extend": font.color.extend
        },
        "sz": font.sz,
        "u": font.u,
        "vertAlign": font.vertAlign,
        "scheme": font.scheme
    }

def dict_to_font(font_dict):
    color_dict = font_dict.pop("color")
    font_dict["color"] = Color(**color_dict)
    openpyxl_font = Font(**font_dict)
    return openpyxl_font
    
def set_real_item_by_index(starting_index, the_list, wanted_type, replacement, delete_cond=False):
    index = starting_index
    item = the_list[index]
    while andExec(not isinstance(item, wanted_type), len(the_list) > 1):
        if delete_cond:     del the_list[index]
        else:               index -= 1
        the_list[index] = replacement
        
def get_real_item_by_index(starting_index, the_list, wanted_type, delete_cond=True):
    index = starting_index
    item = the_list[index]
    while andExec(not isinstance(item, wanted_type), len(the_list) > 1):
        if delete_cond:     del the_list[index]
        else:               index -= 1
        item = the_list[index]
    if isinstance(item, wanted_type):
        return item

def get_text_height(font):
    # Extract font family and font size using regex
    match = re.match(r'(.+)\s(\d+)$', font)
    if match:
        font_family, font_size = match.groups()
    else:
        raise ValueError("Invalid font format. It should be 'Font Family Size', e.g., 'Arial 12'.")

    # Convert font size to an integer
    font_size = int(font_size)

    # Create a Font object with the family and size
    font_object = tkFont.Font(family=font_family, size=font_size)

    # Get font metrics
    font_metrics = font_object.metrics()

    # Extract the font height
    font_height = font_metrics["linespace"]

    return font_height
    
def is_excel_index(text):
    pattern = r'^[A-Z]+[0-9]+$'
    return re.match(pattern, text) is not None

def convert_qutable_to_excel(qutableFormula):
    if not qutableFormula.startswith('='):
        return qutableFormula
    excelFormula = qutableFormula.strip('=')
    qutableVars = createVarList(qutableFormula.strip('='))
    # excelVars = []
    # print(qutableVars)
    for qutableVar in qutableVars:
        qutableIndex = re.split('[CR]', qutableVar)[1:]
        tupleIndex = convert_tuple_to_index(qutableIndex, have_headers=False, col_increment=-1, row_increment=0)
        # excelVars.append( tupleIndex )
        excelFormula = excelFormula.replace(qutableVar, tupleIndex)
    return '='+excelFormula
        
def convert_tuple_to_index(cell_tuple, have_headers, col_increment=0, row_increment=1):
    """
    Convert a tuple representing cell indices (row, column) to MS Excel/Google Sheets cell code.
    """
    col, row = cell_tuple
    col = int(col) + col_increment
    row = int(row) + row_increment
    
    # Convert column index to Excel column code
    col_code = ' '*(col < 0)
    while col >= 0:
        col_code = chr(col % 26 + 65) + col_code
        col //= 26
        col -= 1
    
    # Convert row index to Excel row code (1-based index)
    # row_increment = 1 #int(have_headers)
    row_code = row + 1
    return f"{col_code}{row_code}"

def convert_index_to_tuple(index, have_headers, increment=0):
    book = openpyxl.workbook.workbook.Workbook()
    sheet = openpyxl.worksheet.worksheet.Worksheet(parent=book)
    cell = sheet[index]
    return (cell.column-1+increment, cell.row-2+increment)
    # decrement = not have_headers
    # return (cell.column-1+increment, cell.row-1+increment-decrement)
    
def convert_excel_to_qutable_index(index, have_headers):
    tupleIndex = convert_index_to_tuple(index, have_headers, increment=1)
    return 'C%sR%s' % tupleIndex

def convert_excel_to_qutable_formula(formula, have_headers):
    def repl(match):
        col, row = match.groups()
        col_idx = ord(col.upper()) - ord('A') + 1
        return "C{}R{}".format(col_idx, int(row)-have_headers)
        # return "R{}C{}".format(row, col_idx)
    return re.sub(r"([A-Za-z]+)(\d+)", repl, formula)
    
def donothing():
    filewin = tk.Toplevel()
    button = tk.Button(filewin, text="Do nothing button")
    button.pack()

def removeBinds(specifiedWidget, bindTypeToRemove):
    bindtags = list(specifiedWidget.bindtags())
    bindtags.remove(bindTypeToRemove)
    return specifiedWidget.bindtags(tuple(bindtags))

def sheetDeepCopy(widget, book):
    parent = widget.nametowidget(widget.winfo_parent())
    widgetDeepCopied = type(widget)(parent, book)
    for key in widget.configure():
        try: widgetDeepCopied.configure({key: widget[key]})
        except: pass
    return widgetDeepCopied

def sheetBtnDeepCopy(widget, book, sheetBtnNum):
    parent = widget.nametowidget(widget.winfo_parent())
    widgetDeepCopied = type(widget)(parent, book)
    for key in widget.configure():
        try: widgetDeepCopied.configure({key: widget[key]})
        except: pass
    return widgetDeepCopied

def widgetDeepCopy(widget, widgetDeepCopied):
    for key in widget.configure():
        try: widgetDeepCopied.configure({key: widget[key]})
        except: pass
    return widgetDeepCopied
    
def get_parent(widget):
    return widget.nametowidget(widget.winfo_parent())
    
def resize_permanent(image_path, size):
    Image.open(image_path).resize(size).save( image_path )
    
def find_matching_pixels(image, target_color):
    # Convert the image to a NumPy array
    img_array = np.array(image)

    # Mask for pixels within the tolerance range for each channel
    mask = np.all(img_array == np.array(target_color), axis=-1)

    # Find the coordinates of matching pixels
    matching_pixels = np.argwhere(mask)

    return matching_pixels.tolist()

def is_numeric(string):
    if type(string) == str:
        if string.strip() == '-':
            return True
    try:
        # Attempt to convert the string to a numeric type
        float(string)
        return True
    except (TypeError, ValueError):
        # If conversion fails, return False
        return False

def is_valid_date(date_str):
    # Define common date formats
    date_formats = [
        r"\d{1,2}/\d{1,2}/\d{4}",   # MM/DD/YYYY
        r"\d{4}-\d{1,2}-\d{1,2}"    # YYYY-MM-DD
        # Add more formats if needed
    ]
    
    # Check if the string matches any of the formats
    for date_format in date_formats:
        if re.match(date_format, date_str):
            return True
    return False

def convert_to_datetime(value):
    if is_valid_date(str(value)):
        return pd.to_datetime(value)
    else:
        return value

def convert_dates_to_datetime(df):
    return df.map(convert_to_datetime)

def is_json_serializable(obj):
    try:
        json.dumps(obj)
        return True
    except TypeError:
        return False

def sanitize_sheet_name(name):
    # Replace invalid characters with underscores
    name = re.sub(r'[\/\\?\*:\[\]]', '_', name)
    
    # Trim leading/trailing spaces and apostrophes
    name = name.strip().strip("'")
    
    # Ensure sheet name is not blank and is within 31 characters
    if not name:
        name = "Sheet1"  # Default name if blank
    elif len(name) > 31:
        name = name[:31]  # Truncate if longer than 31 characters
    
    # Check if the sheet name is "History" and modify it
    if name.lower() == "history":
        name += "_1"
    
    return name

def compress_ascii_to_byte(file_addr, output_file):
    with open(file_addr, 'r') as f:
        ascii_data = f.read().encode('ascii')  # Read ASCII file and encode to bytes
    
    compressed_data = zlib.compress(ascii_data)  # Compress the ASCII data
    
    with open(output_file, 'wb') as f:
        f.write(compressed_data)  # Write compressed data to byte file
        
def write_json(json_obj, output_file):
    json_str = json.dumps(json_obj)  # Convert Python object to JSON string
    
    with open(output_file, 'w') as f:
        f.write(json_str)  # Write compressed data to byte file
    

def compress_json_to_byte(json_obj, output_file):
    json_str = json.dumps(json_obj)  # Convert Python object to JSON string
    json_bytes = json_str.encode('utf-8')  # Encode JSON string to bytes (ASCII)
    
    # Convert JSON bytes to base64
    json_base64 = base64.b64encode(json_bytes)
    
    compressed_data = zlib.compress(json_base64)  # Compress the base64 encoded JSON data
    
    with open(output_file, 'wb') as f:
        f.write(compressed_data)  # Write compressed data to byte file

def decompress_byte_to_json(file_addr):
    with open(file_addr, 'rb') as f:
        compressed_data = f.read()  # Read compressed data from byte file
    
    # Decompress the data
    json_base64 = zlib.decompress(compressed_data)
    
    # Decode base64
    json_bytes = base64.b64decode(json_base64)
    
    # Decode bytes to JSON string
    json_str = json_bytes.decode('utf-8')
    
    # Parse JSON string to Python object
    json_obj = json.loads(json_str)
    
    return json_obj



alignmentRule = {(None, 'top'): ( ('nw', 'left'), ('ne', 'right') ) ,
                 (None, 'center'): ( ('w', 'left'), ('e', 'right') ) ,
                 (None, None): ( ('w', 'left'), ('e', 'right') ) ,
                 (None, 'bottom'): ( ('sw', 'left'), ('se', 'right') ) ,
                 ('center', 'top'): ( ('n', 'center'), )*2 ,
                 ('right', 'top'): ( ('ne', 'right'), )*2 ,
                 ('center', None): ( ('center', 'center'), )*2 ,
                 ('center', 'center'): ( ('center', 'center'), )*2 ,
                 ('right', 'center'): ( ('e', 'right'), )*2 ,
                 ('right', None): ( ('e', 'right'), )*2 ,
                 ('right', 'bottom'): ( ('se', 'right'), )*2 ,
                 ('center', 'bottom'): ( ('s', 'center'), )*2 ,
                 ('left', 'bottom'): ( ('sw', 'left'), )*2 ,
                 ('left', None): ( ('w', 'left'), )*2 ,
                 ('left', 'center'): ( ('w', 'left'), )*2 ,
                 ('left', 'top'): ( ('nw', 'left'), )*2 }

alignmentReverse = {'nw': ('left', 'top'), 'n': ('center', 'top'), 'ne': ('right', 'top'), 'w': ('left', 'center'), 'center': ('center', 'center'), 'e': ('right', 'center'), 'se': ('right', 'bottom'), 'sw': ('left', 'bottom'), 's': ('center', 'bottom')}

class LicensePage:
    def __init__(self):
        local_app_data_dir = os.getenv('LOCALAPPDATA')
        qu_table_folder = os.path.join(local_app_data_dir, "QuTable")
        qu_table_agreement_file = os.path.join(qu_table_folder, "QuTable-License.txt")
        self.qu_table_folder = qu_table_folder
        self.qu_table_agreement_file = qu_table_agreement_file
        
        qu_table_agreement_file_cont = open(qu_table_agreement_file).read() if os.path.isfile(qu_table_agreement_file) else None
        
        self.license_text = '''Attribution-NoDerivatives 4.0 International

=======================================================================

Creative Commons Corporation ("Creative Commons") is not a law firm and
does not provide legal services or legal advice. Distribution of
Creative Commons public licenses does not create a lawyer-client or
other relationship. Creative Commons makes its licenses and related
information available on an "as-is" basis. Creative Commons gives no
warranties regarding its licenses, any material licensed under their
terms and conditions, or any related information. Creative Commons
disclaims all liability for damages resulting from their use to the
fullest extent possible.

Using Creative Commons Public Licenses

Creative Commons public licenses provide a standard set of terms and
conditions that creators and other rights holders may use to share
original works of authorship and other material subject to copyright
and certain other rights specified in the public license below. The
following considerations are for informational purposes only, are not
exhaustive, and do not form part of our licenses.

     Considerations for licensors: Our public licenses are
     intended for use by those authorized to give the public
     permission to use material in ways otherwise restricted by
     copyright and certain other rights. Our licenses are
     irrevocable. Licensors should read and understand the terms
     and conditions of the license they choose before applying it.
     Licensors should also secure all rights necessary before
     applying our licenses so that the public can reuse the
     material as expected. Licensors should clearly mark any
     material not subject to the license. This includes other CC-
     licensed material, or material used under an exception or
     limitation to copyright. More considerations for licensors:
    wiki.creativecommons.org/Considerations_for_licensors

     Considerations for the public: By using one of our public
     licenses, a licensor grants the public permission to use the
     licensed material under specified terms and conditions. If
     the licensor's permission is not necessary for any reason--for
     example, because of any applicable exception or limitation to
     copyright--then that use is not regulated by the license. Our
     licenses grant only permissions under copyright and certain
     other rights that a licensor has authority to grant. Use of
     the licensed material may still be restricted for other
     reasons, including because others have copyright or other
     rights in the material. A licensor may make special requests,
     such as asking that all changes be marked or described.
     Although not required by our licenses, you are encouraged to
     respect those requests where reasonable. More considerations
     for the public:
    wiki.creativecommons.org/Considerations_for_licensees


=======================================================================

Creative Commons Attribution-NoDerivatives 4.0 International Public
License

By exercising the Licensed Rights (defined below), You accept and agree
to be bound by the terms and conditions of this Creative Commons
Attribution-NoDerivatives 4.0 International Public License ("Public
License"). To the extent this Public License may be interpreted as a
contract, You are granted the Licensed Rights in consideration of Your
acceptance of these terms and conditions, and the Licensor grants You
such rights in consideration of benefits the Licensor receives from
making the Licensed Material available under these terms and
conditions.


Section 1 -- Definitions.

  a. Adapted Material means material subject to Copyright and Similar
     Rights that is derived from or based upon the Licensed Material
     and in which the Licensed Material is translated, altered,
     arranged, transformed, or otherwise modified in a manner requiring
     permission under the Copyright and Similar Rights held by the
     Licensor. For purposes of this Public License, where the Licensed
     Material is a musical work, performance, or sound recording,
     Adapted Material is always produced where the Licensed Material is
     synched in timed relation with a moving image.

  b. Copyright and Similar Rights means copyright and/or similar rights
     closely related to copyright including, without limitation,
     performance, broadcast, sound recording, and Sui Generis Database
     Rights, without regard to how the rights are labeled or
     categorized. For purposes of this Public License, the rights
     specified in Section 2(b)(1)-(2) are not Copyright and Similar
     Rights.

  c. Effective Technological Measures means those measures that, in the
     absence of proper authority, may not be circumvented under laws
     fulfilling obligations under Article 11 of the WIPO Copyright
     Treaty adopted on December 20, 1996, and/or similar international
     agreements.

  d. Exceptions and Limitations means fair use, fair dealing, and/or
     any other exception or limitation to Copyright and Similar Rights
     that applies to Your use of the Licensed Material.

  e. Licensed Material means the artistic or literary work, database,
     or other material to which the Licensor applied this Public
     License.

  f. Licensed Rights means the rights granted to You subject to the
     terms and conditions of this Public License, which are limited to
     all Copyright and Similar Rights that apply to Your use of the
     Licensed Material and that the Licensor has authority to license.

  g. Licensor means the individual(s) or entity(ies) granting rights
     under this Public License.

  h. Share means to provide material to the public by any means or
     process that requires permission under the Licensed Rights, such
     as reproduction, public display, public performance, distribution,
     dissemination, communication, or importation, and to make material
     available to the public including in ways that members of the
     public may access the material from a place and at a time
     individually chosen by them.

  i. Sui Generis Database Rights means rights other than copyright
     resulting from Directive 96/9/EC of the European Parliament and of
     the Council of 11 March 1996 on the legal protection of databases,
     as amended and/or succeeded, as well as other essentially
     equivalent rights anywhere in the world.

  j. You means the individual or entity exercising the Licensed Rights
     under this Public License. Your has a corresponding meaning.


Section 2 -- Scope.

  a. License grant.

       1. Subject to the terms and conditions of this Public License,
          the Licensor hereby grants You a worldwide, royalty-free,
          non-sublicensable, non-exclusive, irrevocable license to
          exercise the Licensed Rights in the Licensed Material to:

            a. reproduce and Share the Licensed Material, in whole or
               in part; and

            b. produce and reproduce, but not Share, Adapted Material.

       2. Exceptions and Limitations. For the avoidance of doubt, where
          Exceptions and Limitations apply to Your use, this Public
          License does not apply, and You do not need to comply with
          its terms and conditions.

       3. Term. The term of this Public License is specified in Section
          6(a).

       4. Media and formats; technical modifications allowed. The
          Licensor authorizes You to exercise the Licensed Rights in
          all media and formats whether now known or hereafter created,
          and to make technical modifications necessary to do so. The
          Licensor waives and/or agrees not to assert any right or
          authority to forbid You from making technical modifications
          necessary to exercise the Licensed Rights, including
          technical modifications necessary to circumvent Effective
          Technological Measures. For purposes of this Public License,
          simply making modifications authorized by this Section 2(a)
          (4) never produces Adapted Material.

       5. Downstream recipients.

            a. Offer from the Licensor -- Licensed Material. Every
               recipient of the Licensed Material automatically
               receives an offer from the Licensor to exercise the
               Licensed Rights under the terms and conditions of this
               Public License.

            b. No downstream restrictions. You may not offer or impose
               any additional or different terms or conditions on, or
               apply any Effective Technological Measures to, the
               Licensed Material if doing so restricts exercise of the
               Licensed Rights by any recipient of the Licensed
               Material.

       6. No endorsement. Nothing in this Public License constitutes or
          may be construed as permission to assert or imply that You
          are, or that Your use of the Licensed Material is, connected
          with, or sponsored, endorsed, or granted official status by,
          the Licensor or others designated to receive attribution as
          provided in Section 3(a)(1)(A)(i).

  b. Other rights.

       1. Moral rights, such as the right of integrity, are not
          licensed under this Public License, nor are publicity,
          privacy, and/or other similar personality rights; however, to
          the extent possible, the Licensor waives and/or agrees not to
          assert any such rights held by the Licensor to the limited
          extent necessary to allow You to exercise the Licensed
          Rights, but not otherwise.

       2. Patent and trademark rights are not licensed under this
          Public License.

       3. To the extent possible, the Licensor waives any right to
          collect royalties from You for the exercise of the Licensed
          Rights, whether directly or through a collecting society
          under any voluntary or waivable statutory or compulsory
          licensing scheme. In all other cases the Licensor expressly
          reserves any right to collect such royalties.


Section 3 -- License Conditions.

Your exercise of the Licensed Rights is expressly made subject to the
following conditions.

  a. Attribution.

       1. If You Share the Licensed Material, You must:

            a. retain the following if it is supplied by the Licensor
               with the Licensed Material:

                 i. identification of the creator(s) of the Licensed
                    Material and any others designated to receive
                    attribution, in any reasonable manner requested by
                    the Licensor (including by pseudonym if
                    designated);

                ii. a copyright notice;

               iii. a notice that refers to this Public License;

                iv. a notice that refers to the disclaimer of
                    warranties;

                 v. a URI or hyperlink to the Licensed Material to the
                    extent reasonably practicable;

            b. indicate if You modified the Licensed Material and
               retain an indication of any previous modifications; and

            c. indicate the Licensed Material is licensed under this
               Public License, and include the text of, or the URI or
               hyperlink to, this Public License.

          For the avoidance of doubt, You do not have permission under
          this Public License to Share Adapted Material.

       2. You may satisfy the conditions in Section 3(a)(1) in any
          reasonable manner based on the medium, means, and context in
          which You Share the Licensed Material. For example, it may be
          reasonable to satisfy the conditions by providing a URI or
          hyperlink to a resource that includes the required
          information.

       3. If requested by the Licensor, You must remove any of the
          information required by Section 3(a)(1)(A) to the extent
          reasonably practicable.


Section 4 -- Sui Generis Database Rights.

Where the Licensed Rights include Sui Generis Database Rights that
apply to Your use of the Licensed Material:

  a. for the avoidance of doubt, Section 2(a)(1) grants You the right
     to extract, reuse, reproduce, and Share all or a substantial
     portion of the contents of the database, provided You do not Share
     Adapted Material;

  b. if You include all or a substantial portion of the database
     contents in a database in which You have Sui Generis Database
     Rights, then the database in which You have Sui Generis Database
     Rights (but not its individual contents) is Adapted Material; and

  c. You must comply with the conditions in Section 3(a) if You Share
     all or a substantial portion of the contents of the database.

For the avoidance of doubt, this Section 4 supplements and does not
replace Your obligations under this Public License where the Licensed
Rights include other Copyright and Similar Rights.


Section 5 -- Disclaimer of Warranties and Limitation of Liability.

  a. UNLESS OTHERWISE SEPARATELY UNDERTAKEN BY THE LICENSOR, TO THE
     EXTENT POSSIBLE, THE LICENSOR OFFERS THE LICENSED MATERIAL AS-IS
     AND AS-AVAILABLE, AND MAKES NO REPRESENTATIONS OR WARRANTIES OF
     ANY KIND CONCERNING THE LICENSED MATERIAL, WHETHER EXPRESS,
     IMPLIED, STATUTORY, OR OTHER. THIS INCLUDES, WITHOUT LIMITATION,
     WARRANTIES OF TITLE, MERCHANTABILITY, FITNESS FOR A PARTICULAR
     PURPOSE, NON-INFRINGEMENT, ABSENCE OF LATENT OR OTHER DEFECTS,
     ACCURACY, OR THE PRESENCE OR ABSENCE OF ERRORS, WHETHER OR NOT
     KNOWN OR DISCOVERABLE. WHERE DISCLAIMERS OF WARRANTIES ARE NOT
     ALLOWED IN FULL OR IN PART, THIS DISCLAIMER MAY NOT APPLY TO YOU.

  b. TO THE EXTENT POSSIBLE, IN NO EVENT WILL THE LICENSOR BE LIABLE
     TO YOU ON ANY LEGAL THEORY (INCLUDING, WITHOUT LIMITATION,
     NEGLIGENCE) OR OTHERWISE FOR ANY DIRECT, SPECIAL, INDIRECT,
     INCIDENTAL, CONSEQUENTIAL, PUNITIVE, EXEMPLARY, OR OTHER LOSSES,
     COSTS, EXPENSES, OR DAMAGES ARISING OUT OF THIS PUBLIC LICENSE OR
     USE OF THE LICENSED MATERIAL, EVEN IF THE LICENSOR HAS BEEN
     ADVISED OF THE POSSIBILITY OF SUCH LOSSES, COSTS, EXPENSES, OR
     DAMAGES. WHERE A LIMITATION OF LIABILITY IS NOT ALLOWED IN FULL OR
     IN PART, THIS LIMITATION MAY NOT APPLY TO YOU.

  c. The disclaimer of warranties and limitation of liability provided
     above shall be interpreted in a manner that, to the extent
     possible, most closely approximates an absolute disclaimer and
     waiver of all liability.


Section 6 -- Term and Termination.

  a. This Public License applies for the term of the Copyright and
     Similar Rights licensed here. However, if You fail to comply with
     this Public License, then Your rights under this Public License
     terminate automatically.

  b. Where Your right to use the Licensed Material has terminated under
     Section 6(a), it reinstates:

       1. automatically as of the date the violation is cured, provided
          it is cured within 30 days of Your discovery of the
          violation; or

       2. upon express reinstatement by the Licensor.

     For the avoidance of doubt, this Section 6(b) does not affect any
     right the Licensor may have to seek remedies for Your violations
     of this Public License.

  c. For the avoidance of doubt, the Licensor may also offer the
     Licensed Material under separate terms or conditions or stop
     distributing the Licensed Material at any time; however, doing so
     will not terminate this Public License.

  d. Sections 1, 5, 6, 7, and 8 survive termination of this Public
     License.


Section 7 -- Other Terms and Conditions.

  a. The Licensor shall not be bound by any additional or different
     terms or conditions communicated by You unless expressly agreed.

  b. Any arrangements, understandings, or agreements regarding the
     Licensed Material not stated herein are separate from and
     independent of the terms and conditions of this Public License.


Section 8 -- Interpretation.

  a. For the avoidance of doubt, this Public License does not, and
     shall not be interpreted to, reduce, limit, restrict, or impose
     conditions on any use of the Licensed Material that could lawfully
     be made without permission under this Public License.

  b. To the extent possible, if any provision of this Public License is
     deemed unenforceable, it shall be automatically reformed to the
     minimum extent necessary to make it enforceable. If the provision
     cannot be reformed, it shall be severed from this Public License
     without affecting the enforceability of the remaining terms and
     conditions.

  c. No term or condition of this Public License will be waived and no
     failure to comply consented to unless expressly agreed to by the
     Licensor.

  d. Nothing in this Public License constitutes or may be interpreted
     as a limitation upon, or waiver of, any privileges and immunities
     that apply to the Licensor or You, including from the legal
     processes of any jurisdiction or authority.

=======================================================================

Creative Commons is not a party to its public
licenses. Notwithstanding, Creative Commons may elect to apply one of
its public licenses to material it publishes and in those instances
will be considered the “Licensor.” The text of the Creative Commons
public licenses is dedicated to the public domain under the CC0 Public
Domain Dedication. Except for the limited purpose of indicating that
material is shared under a Creative Commons public license or as
otherwise permitted by the Creative Commons policies published at
creativecommons.org/policies, Creative Commons does not authorize the
use of the trademark "Creative Commons" or any other trademark or logo
of Creative Commons without its prior written consent including,
without limitation, in connection with any unauthorized modifications
to any of its public licenses or any other arrangements,
understandings, or agreements concerning use of licensed material. For
the avoidance of doubt, this paragraph does not form part of the
public licenses.

Creative Commons may be contacted at creativecommons.org.

By proceeding with the installation of this software, you agree to abide by the terms and conditions of the license agreement.
        '''
        
        self.license_file_content = f'When you (user or users) first opened QuTable, you have agreed to this following license below by clicking the Checkmark with the label "I agree with the software license" and then clicking the "Proceed" button. Here is the license below:\n\n{self.license_text}'
        
        # Rename if that same path is not a file
        if os.path.exists(qu_table_agreement_file):
            if not os.path.isfile(qu_table_agreement_file):
                file_addr, extension = os.path.splitext(qu_table_agreement_file)
                extension = extension.lower()
                i = 1
                while 1:
                    newFileAddr = qu_table_agreement_file.replace(file_addr, f'{file_addr}-{i}')
                    if not os.path.exists(newFileAddr):
                        break
                    i += 1
                os.rename(qu_table_agreement_file, newFileAddr)
        
        if andExec(os.path.isfile(qu_table_agreement_file), qu_table_agreement_file_cont == self.license_file_content):
            self.after_agreeing()
        
        else:
            self.window = tk.Tk()
            self.window.title('License Agreement')
            self.window.configure(bg='#FFFFFF')
            self.window.option_add('*background', '#FFFFFF')
            
            self.icon = ImageTk.PhotoImage(image=imageDict['QuTableIcon'] , master=self.window )
            self.window.iconphoto(True, self.icon)

            self.center_window()
            
            # Big heading
            tk.Label(self.window, text='Software License Agreement', font=('Arial', 18, 'bold'), bg='#FFFFFF').grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky=tk.W)

            tk.Label(self.window, text='Please read carefully before using QuTable.\n(Note that this Warning would NOT appear again as long as a file related to your affirmation to this warning is still written in a file having a location of "%LOCALAPPDATA%/QuTable/QuTable-License.txt")\n\nThis software is provided under the terms and conditions of the license agreement included with the software package, shown as per following:', bg='#FFFFFF', wraplength=800).grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky='w')
            self.license_textbox = tk.Text(self.window, width=77, height=17)
            self.license_textbox.insert(tk.END, self.license_text)
            self.license_textbox.config(state=tk.DISABLED)
            self.license_textbox.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky='nsew')
            
            self.agree_var = False
            self.agree_checkbox = tk.Checkbutton(self.window, text='I agree with the software license', font=("Arial", 10))
            self.agree_checkbox.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky='w')
            self.agree_checkbox.bind('<Button-1>', lambda event: self.agree_var_change())
            self.agree_checkbox.bind('<Button-2>', lambda event: self.agree_var_change())
            self.agree_checkbox.bind('<Button-3>', lambda event: self.agree_var_change())

            self.proceed_button_license = ttk.Button(self.window, text='Proceed', command=self.proceed_by_agreeing, state='disabled')
            self.proceed_button_license.grid(row=4, column=1, padx=10, pady=5, sticky='e')
            
            if getattr(sys, 'frozen', False):
                self.window.mainloop()
        
    def agree_var_change(self):
        self.agree_var = not self.agree_var
        self.proceed_button_license.config(state="enabled" if self.agree_var else "disabled")
        
    def proceed_by_agreeing(self):
        if self.agree_var:
            self.window.destroy()
            os.makedirs(self.qu_table_folder, exist_ok=True)
            licenseIO = open(self.qu_table_agreement_file, 'w+')
            licenseIO.write(self.license_file_content)
            licenseIO.close()
            self.after_agreeing()
        else:
            messagebox.showwarning('Warning', 'You must agree to the software license agreement.', parent=self.window)
        
    def after_agreeing(self_object):
        global app, self
        app = QuTable()
        self = app[0,0]
        
    def center_window(self, width=805, height=630):
        window = self.window
        
        if not width:  width =  window.winfo_width()
        if not height: height = window.winfo_height()
        
        # Get the screen width and height
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()

        # Calculate the position of the window to center it on the screen
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

        # Set the geometry of the window to center it on the screen
        window.geometry(f"{width}x{height}+{x}+{y}")

class RightClickMenu(tk.Menu):
    def __init__(self, widget, book, parent, purpose=None, funcBefore=None, funcAfter=None, *args, **kwargs):
        super().__init__(parent, tearoff=0, bg='#FFFFFF', *args, **kwargs)
        self.purpose = purpose
        self.book = book
        self.widget = widget
        self.funcBefore = funcBefore
        self.funcAfter = funcAfter
        self.activate()
        
    def when_right_clicked(self, event=None):
        if self.funcBefore:
            self.funcBefore()
        if type(self.purpose) == SheetToggleButton:
            if len(self.book.sheets) <= 1:
                self.entryconfig('Delete', state="disabled")
        try: 
            self.tk_popup(event.x_root, event.y_root) 
        finally: 
            self.grab_release()
        if self.funcAfter:
            self.funcAfter()
        
    def activate(self, widget=None):
        if widget is None:
            widget = self.widget
        widget.bind("<Button-3>", self.when_right_clicked)
        
class ButtonTooltip:
    def __init__(self, widget, text, y_translation):
        self.widget = widget
        self.text = text
        self.y_translation = y_translation
        self.tooltipWindow = tooltipWindow = tk.Toplevel(self.widget)
        self.tooltipMsg = None
        tooltipWindow.wm_overrideredirect(1)
        tooltipWindow.withdraw()

    def showTooltip(self, event):
        tooltipWindow = self.tooltipWindow
        self.tooltipMsg = tk.Label(tooltipWindow, text=self.text, background='#FFFFFF', relief='solid', borderwidth=1)
        self.tooltipMsg.pack(ipadx=1)
        # xPos, yPos, = event.x_root, event.y_root
        xPos, yPos, _, _ = self.widget.bbox('insert')
        xPos += event.x_root
        yPos += event.y_root
        tooltipWindow.wm_geometry(f'+{xPos}+{yPos+self.y_translation}')
        tooltipWindow.deiconify()

    def hideTooltip(self, event):
        # print(event.widget)
        if self.tooltipMsg:
            self.tooltipMsg.destroy()
        if self.tooltipWindow:
            self.tooltipWindow.withdraw()
        self.tooltipMsg = None

class SuperLabel(tk.Label):
    def __init__(self, book, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        sheet = book.sheet
        self.configure(relief='flat')
        self.configure(background='#FFFFFF')

class SuperEntry(tk.Entry):
    def __init__(self, book, parent, tooltipText, focusOutFuncExt, image_path=None, to_end=None, vertical=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.tooltip = ButtonTooltip(self, tooltipText, y_translation=10)
        self.book = book
        self.focusOutFuncExt = focusOutFuncExt
        self.bind('<Enter>', self.hoverButton)
        self.bind('<Leave>', self.leaveButton)
        self.bind("<FocusIn>", self.focusInFunc )
        
    def focusInFunc(self, event=None):
        self.bind("<Return>", self.focusOutFunc )
        self.bind("<FocusOut>", self.focusOutFunc )
        # Moving across cells using keyboard arrows
        self.book.window.unbind('<Key>')
        self.book.window.unbind('<KeyRelease>')
        
    def focusOutFunc(self, event=None):
        # if event.keysym == 'Return':
            # self.book.window.unbind('<FocusOut>')
        # # elif event.type == 10:
        # else:
            # self.book.window.unbind('<Return>')
        self.focusOutFuncExt(event=event)
        # print(1270, event.keysym, event.num, event.state, event.type)
        # print(1271, dir(event))
        # Moving across cells using keyboard arrows
        self.book.window.bind('<Key>', lambda event: self.book.sheet.keyHandle(event=event) )
        self.book.window.bind('<KeyRelease>', lambda event: self.book.sheet.keyReleaseHandle(event=event) )
        
    def hoverButton(self, event=None):
        if self.tooltip:
            if not self.tooltip.tooltipMsg:
                self.tooltip.showTooltip(event=event)
        
    def leaveButton(self, event=None):
        x, y = event.x_root, event.y_root
        widgetHovered = event.widget.winfo_containing(x, y)
        if self.tooltip:
            if widgetHovered != self.tooltip.tooltipMsg:
                self.tooltip.hideTooltip(event=event)
            else:
                self.tooltip.hideTooltip(event=event)
                self.tooltip.showTooltip(event=event)
                if self.tooltip.tooltipMsg:
                    self.tooltip.tooltipMsg.bind('<Leave>', self.leaveTooltip)
        
    def leaveTooltip(self, event=None):
        x, y = event.x_root, event.y_root
        widgetHovered = event.widget.winfo_containing(x, y)
        if widgetHovered != self:
            self.configure(background='#FFFFFF')
            self.tooltip.hideTooltip(event=event)
    
class SuperMenuButton(tk.Menubutton):
    switchedOn = False
    
    def __init__(self, book, parent, icon, image_path=None, to_end=None, vertical=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        sheet = book.sheet
        self.sheet = sheet
        self.image_path = image_path
        self.hoveredIconName = image_path.replace('.png', '-Hovered.png') if image_path else ''
        self.tooltip = ButtonTooltip(self, kwargs['text'], y_translation=30) if 'text' in kwargs else None
        self['text'] = icon
        self.bind('<Enter>', self.hoverButton)
        self.bind('<Leave>', self.leaveButton)
        self.configure(relief='flat')
        self.configure(background='#FFFFFF')
        # self.bind('<1>', lambda event: self.sheet.focus_set() )
        
        self.to_end = to_end
        self.vertical = vertical
        
        self.menu = tk.Menu(self, tearoff=False)
        self.configure(menu=self.menu)
        
        # print('book =', repr(book) )
        
    def add_command(self, *args, **kwargs):
        self.menu.add_command(*args, **kwargs)
        
    def delete(self, *args, **kwargs):
        self.menu.delete(*args, **kwargs)
        
    def switchOn(self, event=None):
        self.switchedOn = True
        self.configure(background='#b3b3b3')
        
    def switchOff(self, event=None):
        self.switchedOn = False
        self.configure(background='#FFFFFF')
        
    def hoverButton(self, event=None):
        if self['state'] != 'disabled':
            if self.switchedOn:
                self.configure(background='#a3a3a3')
            else:
                self.configure(background='#d3d3d3')
        if self.tooltip:
            if not self.tooltip.tooltipMsg:
                self.tooltip.showTooltip(event=event)
        
    def leaveButton(self, event=None):
        x, y = event.x_root, event.y_root
        widgetHovered = event.widget.winfo_containing(x, y)
        if self.tooltip:
            if widgetHovered != self.tooltip.tooltipMsg:
                if self.switchedOn:
                    self.configure(background='#b3b3b3')
                else:
                    self.configure(background='#FFFFFF')
                self.tooltip.hideTooltip(event=event)
            else:
                self.tooltip.hideTooltip(event=event)
                self.tooltip.showTooltip(event=event)
                if self.tooltip.tooltipMsg:
                    self.tooltip.tooltipMsg.bind('<Leave>', self.leaveTooltip)
        
    def leaveTooltip(self, event=None):
        x, y = event.x_root, event.y_root
        widgetHovered = event.widget.winfo_containing(x, y)
        if widgetHovered != self:
            self.configure(background='#FFFFFF')
            self.tooltip.hideTooltip(event=event)
            
        
class SuperButton(tk.Button):
    switchedOn = False
    
    def __init__(self, book, parent, image_path=None, to_end=None, vertical=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        sheet = book.sheet
        self.sheet = sheet
        self.image_path = image_path
        self.hoveredIconName = image_path.replace('.png', '-Hovered.png') if image_path else ''
        self.tooltip = ButtonTooltip(self, kwargs['text'], y_translation=30) if 'text' in kwargs else None
        self.bind('<Enter>', self.hoverButton)
        self.bind('<Leave>', self.leaveButton)
        self.configure(relief='flat')
        self.configure(background='#FFFFFF')
        # self.bind('<1>', lambda event: self.sheet.focus_set() )
        
        self.to_end = to_end
        self.vertical = vertical
        
        # print('book =', repr(book) )
        
    def switchOn(self, event=None):
        self.switchedOn = True
        self.configure(background='#b3b3b3')
        
    def switchOff(self, event=None):
        self.switchedOn = False
        self.configure(background='#FFFFFF')
        
    def hoverButton(self, event=None):
        if self['state'] != 'disabled':
            if self.switchedOn:
                self.configure(background='#a3a3a3')
            else:
                self.configure(background='#d3d3d3')
        if self.tooltip:
            if not self.tooltip.tooltipMsg:
                self.tooltip.showTooltip(event=event)
        
    def leaveButton(self, event=None):
        x, y = event.x_root, event.y_root
        widgetHovered = event.widget.winfo_containing(x, y)
        if self.tooltip:
            if widgetHovered != self.tooltip.tooltipMsg:
                if self.switchedOn:
                    self.configure(background='#b3b3b3')
                else:
                    self.configure(background='#FFFFFF')
                self.tooltip.hideTooltip(event=event)
            else:
                self.tooltip.hideTooltip(event=event)
                self.tooltip.showTooltip(event=event)
                if self.tooltip.tooltipMsg:
                    self.tooltip.tooltipMsg.bind('<Leave>', self.leaveTooltip)
        
    def leaveTooltip(self, event=None):
        x, y = event.x_root, event.y_root
        widgetHovered = event.widget.winfo_containing(x, y)
        if widgetHovered != self:
            self.configure(background='#FFFFFF')
            self.tooltip.hideTooltip(event=event)
            
        
class SheetToggleButton(tk.Button):
    def __init__(self, sheet, book, parent, sheetBtnNum, *args, **kwargs):
        if parent is None: parent = widget.nametowidget(widget.winfo_parent())
        super().__init__(parent, *args, **kwargs)
        self.parent = parent
        self.sheet = sheet
        self.book = book
        self.sheetBtnNum = sheetBtnNum
        self.activate()
        self.pack(side='left')
        self.configure(relief='flat')
        self.bind('<1>', lambda event: self.sheet.focus_set() )
        
    def toggleTheSheet(self, sheetBtnNum=None):
        if sheetBtnNum is None:
            sheetBtnNum=self.sheetBtnNum
        self.book.toggleSheet(sheetBtnNum)
        
    def hover(self, event=None):
        self.configure(background='#d3d3d3')
        self.configure(font=('Arial', 14, 'bold') )
        
    def leave(self, event=None):
        self.configure(background='#FFFFFF')
        self.configure(font=('Arial', 14) )
        
    def activate(self):
        book = self.book
        parent = self.parent
        sheetBtnNum = self.sheetBtnNum
        self.configure(command=self.toggleTheSheet)
        self.bind_unclicked()
        rightClickMenu = RightClickMenu(widget=self, book=book, purpose=self, parent=parent)
        rightClickMenu.add_command(label="Rename", command=lambda sheetBtnNum=sheetBtnNum: book.renameSheet(sheetBtnNum) )
        rightClickMenu.add_command(label="Duplicate", command=lambda sheetBtnNum=sheetBtnNum: book.duplicateSheet(sheetBtnNum) )
        rightClickMenu.add_command(label="Delete", command=lambda sheetBtnNum=sheetBtnNum: book.deleteSheet(sheetBtnNum) )
        self.rightClickMenu = rightClickMenu
        
    def bind_unclicked(self):
        self.bind('<Enter>', self.hover)
        self.bind('<Leave>', self.leave)
        self['background'] = '#FFFFFF'
        self['font'] = ('Arial', 14)
        
    def unbind_clicked(self):
        self.unbind('<Enter>')
        self.unbind('<Leave>')
        self['background'] = 'lightblue'
        self['font'] = ('Arial', 14, 'bold')
        
dummy_root = tk.Tk()
dummy_root.withdraw()


class SheetSeparator(tk.Frame):
    def __repr__(self):
        return f'{self.__class__.__name__}(num={self.num}, parent={self.parent}, sepType={self.sepType}, sheet=app[0,0])'
        
    def set_colIndex(self, _colIndex):
        self.col_grid = 2*self._colIndex+self.colIndex_inc
        if self.winfo_manager() == 'grid':
            self.takeoff()
            self.install()
        
    def set_rowIndex(self, _rowIndex):
        self.row_grid = 2*self._rowIndex+self.rowIndex_inc
        if self.winfo_manager() == 'grid':
            self.takeoff()
            self.install()
        
    @property
    def length(self):
        return self._length
    @length.setter
    def length(self, new_length):
        self._length = new_length
        if orExec(type(self) == ColSeparator, type(self) == RowSeparator):
            self[self.length_indicator] = new_length
        
    @property
    def thickness(self):
        return self._thickness
    @thickness.setter
    def thickness(self, new_thickness):
        self._thickness = new_thickness
        if orExec(type(self) == ColSeparator, type(self) == RowSeparator):
            self[self.thickness_indicator] = new_thickness
        
    @property
    def colIndex(self):
        return self._colIndex
    @colIndex.setter
    def colIndex(self, _colIndex):
        self._colIndex = _colIndex
        # self.set_colIndex(_colIndex)
        self.col_grid = 2*self._colIndex+self.colIndex_inc
        # if self.winfo_manager() == 'grid':
            # self.takeoff()
            # self.install()
        
    @property
    def rowIndex(self):
        return self._rowIndex
    @rowIndex.setter
    def rowIndex(self, _rowIndex):
        self._rowIndex = _rowIndex
        # self.set_rowIndex(_rowIndex)
        self.row_grid = 2*self._rowIndex+self.rowIndex_inc
        # if self.winfo_manager() == 'grid':
            # self.takeoff()
            # self.install()
    
    @property
    def col_grid(self):
        return self._col_grid
    @col_grid.setter
    def col_grid(self, new_col_grid):
        self._col_grid = new_col_grid
        if self.winfo_manager():
            self.grid_configure(column=new_col_grid)
        
    @property
    def row_grid(self):
        return self._row_grid
    @row_grid.setter
    def row_grid(self, new_row_grid):
        self._row_grid = new_row_grid
        if self.winfo_manager():
            self.grid_configure(row=new_row_grid)
    
    def changeColor(self, color):
        self.bg = color
    
    def manage(self, select, *args, **kwargs):
        if select:
            self.install(*args, **kwargs)
        else:
            self.takeoff(*args, **kwargs)
    
    def install(self):
        self.grid(column=self.col_grid, row=self.row_grid, sticky='nsew')
        
    def takeoff(self):
        self.grid_forget()
        
class ColSeparator(SheetSeparator):
    length_indicator    = 'height'
    thickness_indicator = 'width'
    colIndex_inc, rowIndex_inc = 3, 3
        
    def __init__(self, parent, num, book, sheet, thickness, length=28, sepType='Selector', x=None, y=None, *args, **kwargs):
        tk.Frame.__init__(self, sheet.sheetInner, width=thickness, *args, **kwargs)
        self.num = num
        self.book = book
        self.sheet = sheet
        self.parent = parent
        self.sepType = sepType
        
        self.thickness = thickness
        self.length = length if length else parent.height
        
        self.bg = self['bg'] = kwargs['bg'] if 'bg' in kwargs else '#000000'
            
        if sepType == 'ShadowBorder':
            self.sheet.shadowBorders.add(self)
        self.sheet.rightClickMenu.activate(widget=self)
        
        self.colIndex, self.rowIndex = parent.index_tuple
        
    def install(self, pink_purple=False, self_induced=True):
        self.grid(column=self.col_grid, row=self.row_grid, sticky='nsew')
        if pink_purple:
            # self.changeColor('#FF6BFF')
            self.changeColor('#FFA3A3')
        if self.sepType == 'Selector':
            self.lift()
        elif self.sepType == 'Border':
            if self.parent.colSelector.winfo_manager():
                self.lift(aboveThis=self.parent.colSelector)
            else:
                self.lift()
            if self_induced:
                self.parent.moveRight().left_border = True
                self.parent.right_border = True
        elif self.sepType == 'ShadowBorder':
            if self.parent.colBorder.winfo_manager():
                self.lift(aboveThis=self.parent.colBorder)
            elif self.parent.colSelector.winfo_manager():
                self.lift(aboveThis=self.parent.colSelector)
            else:
                self.lift()
        
    def takeoff(self, self_induced=True):
        self.grid_forget()
        if isinstance(self.parent, Column):
            return
        if self.sepType == 'Border':
            self.lift()
            if self_induced:
                self.parent.moveRight().left_border = False
                self.parent.right_border = False
        elif self.sepType == 'ShadowBorder':
            self.lift(aboveThis=self.parent.colBorder)
        
class RowSeparator(SheetSeparator):
    length_indicator    = 'width'
    thickness_indicator = 'height'
    colIndex_inc, rowIndex_inc = 2, 4
    
    def __init__(self, parent, num, book, sheet, thickness, length=None, sepType='Selector', x=None, y=None, *args, **kwargs):
        tk.Frame.__init__(self, sheet.sheetInner, height=thickness, *args, **kwargs)
        self.num = num
        self.book = book
        self.sheet = sheet
        self.parent = parent
        self.sepType = sepType
        
        self.thickness = thickness
        self.length = self['width'] = length if length else parent.width
        
        self.bg = self['bg'] = kwargs['bg'] if 'bg' in kwargs else '#000000'
        
        if isinstance(parent, IndexCell):
            self.changeColor('#D5D5D5')
            
        if sepType == 'ShadowBorder':
            self.sheet.shadowBorders.add(self)
        self.sheet.rightClickMenu.activate(widget=self)
        
        self.colIndex, self.rowIndex = parent.index_tuple
        
    def install(self, pink_purple=False, self_induced=True):
        self.grid(column=self.col_grid, row=self.row_grid, sticky='nsew')
        if pink_purple:
            # self.changeColor('#FF6BFF')
            self.changeColor('#FFA3A3')
        if self.sepType == 'Selector':
            self.lift()
        elif self.sepType == 'Border':
            if self.parent.rowSelector.winfo_manager():
                self.lift(aboveThis=self.parent.rowSelector)
            else:
                self.lift()
            if self_induced:
                self.parent.moveDown().top_border = True
                self.parent.bottom_border = True
        elif self.sepType == 'ShadowBorder':
            if self.parent.rowBorder.winfo_manager():
                self.lift(aboveThis=self.parent.rowBorder)
            elif self.parent.rowSelector.winfo_manager():
                self.lift(aboveThis=self.parent.rowSelector)
            else:
                self.lift()
        
    def takeoff(self, self_induced=True):
        self.grid_forget()
        if isinstance(self.parent, Column):
            return
        if self.sepType == 'Border':
            self.lift()
            if self_induced:
                self.parent.moveDown().top_border = False
                self.parent.bottom_border = False
        elif self.sepType == 'ShadowBorder':
            self.lift(aboveThis=self.parent.colBorder)
        
        
class CellFrame(SheetSeparator):
    colIndex_inc, rowIndex_inc = 2, 3
    
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, *args, **kwargs)
        self.parent = parent
        self.sepType = 'CellFrame'
        self.colIndex, self.rowIndex = parent.index_tuple
        
    def __repr__(self):
        return f'CellFrame for {self.parent}'
        
    def __str__(self):
        return self.__repr__()
        
        
class Cell:
    deleted = False
    default = False
    formula_context = False
    indicator = None
    event_state, event_num, event_widget = None, None, None
    widgetEnd = None
        
    def __init__(self, parent, book, index, column, sheet, col_width, row_height, indirect_inheritance=True, inserted=False, *args, **kwargs):
        # self.is_placeholder = self.column.is_placeholder
        self._index = index
        self._counting_index = (index[0]+1, index[1]+1)
        self.column = column
        
        self.book = book
        self.sheet = sheet
        self.cell = self
        
        self.parent = parent
        
        self.inserted = inserted
        
        self.widgetEnd_prev = self
        self.widgetEnd_prev_debug = self
        
        if indirect_inheritance:
            self.color_reinforce()
            self.alignment_reinforce()
            self._font = self['font']
        
        # self.width  = self['width']  = kwargs['width']  if 'width' in kwargs  else self.winfo_width()
        # self.height = self['height'] = kwargs['height'] if 'height' in kwargs else self.winfo_height()
        
        # print(f'kwargs = {kwargs}')
        self._width  = col_width  = kwargs['width']  if 'width' in kwargs  else col_width
        self._height = row_height = kwargs['height'] if 'height' in kwargs else row_height
        
        self.cellFrame = CellFrame(parent=self, master=sheet.sheetInner, width=self.width, height=self.height, bg='#FFFFFF')
        # if sheet.sheetName == 'Winter 2024':
            # print(f'{self} -- cellFrame-width = {self.width}')
        
        if indirect_inheritance:
            removeBinds(self, 'all')
        
        self.colSelector = ColSeparator(self, index[0], book, sheet, thickness=2, length=self.height, bg='#FF0000', sepType='Selector')
        self.rowSelector = RowSeparator(self, index[1], book, sheet, thickness=2, length=self.width, bg='#FF0000', sepType='Selector')
        self.colBorder = ColSeparator(self, index[0], book, sheet, thickness=2, length=self.height, sepType='Border')
        self.rowBorder = RowSeparator(self, index[1], book, sheet, thickness=2, length=self.width, sepType='Border')
        self.colShadowBorder = ColSeparator(self, index[0], book, sheet, thickness=2, length=self.height, bg='#E1E3E1', sepType='ShadowBorder')
        self.rowShadowBorder = RowSeparator(self, index[1], book, sheet, thickness=2, length=self.width, bg='#E1E3E1', sepType='ShadowBorder')
        
        self.cell_parts = [self, self.cellFrame, self.colSelector, self.rowSelector, self.colBorder, self.rowBorder, self.colShadowBorder, self.rowShadowBorder]
        self.cell_separators = [self.colSelector, self.rowSelector, self.colBorder, self.rowBorder, self.colShadowBorder, self.rowShadowBorder]
        self.col_separators = [self.colSelector, self.colBorder, self.colShadowBorder]
        self.row_separators = [self.rowSelector, self.rowBorder, self.rowShadowBorder]
        
        [self._top_border, self._bottom_border, self._left_border, self._right_border] = [False]*4
        
        self.index = index
        self.counting_index = (index[0]+1, index[1]+1)
        self.cellOpenPyxlIndex = convert_tuple_to_index(index, sheet.have_headers)
        self.cellIndexCode = 'C%sR%s' % self.counting_index
    
    def changeTextWidget(self, _content):
        try:
            self.config(text=_content)
        except Exception as e:
            print("%s\t%s\t%s\n%s\n\n" % (1808, self, e, format_exc()) , file=sys.stderr)
    
    def contentExt(self, _content):
        try:
            if isinstance(self, CellLabel):
                if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                    self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['content'] = _content
            rowIndex = self.index[1]
            if isinstance(self, ColumnTitle):
                if self.sheet.edit_status:
                    self.sheet.df.rename(columns={self.content: _content})
            elif isinstance(self, CellLabel):
                if not self.is_placeholder:
                    if self.sheet.edit_status:
                        self.sheet.df.loc[rowIndex, self.column.colTitle] = _content
                    if not self.inserted:
                        self.column.cellValues[rowIndex] = _content
            # if isinstance(self, CellLabel):
                # # self.formula = _content
                # if not self.column.is_placeholder:
                    # if not self.cellOpenPyxl.value:
                        # self.cellOpenPyxl.value = _content
        except Exception as e:
            print("%s\t%s\t%s\n%s\n\n" % (1835, self, e, format_exc()) , file=sys.stderr)
    
    @property
    def content(self):
        try:
            return self._content
        except AttributeError:
            return ''
    @content.setter
    def content(self, _content):
        # _content = float_int_convert(_content)
        _content = properize(_content)
        self._content = _content
        # t = threading.Thread(target=Cell.changeTextWidget, args=(self, _content,) )
        # t.start()
        self.changeTextWidget(_content)
        t = threading.Thread(target=Cell.contentExt, args=(self, _content,) )
        t.start()
        if isinstance(self, ColumnTitle):
            self.column.colTitle = _content
        # self.contentExt(_content)
        
    @property
    def colIndex(self):
        return self._colIndex
    @colIndex.setter
    def colIndex(self, new_colIndex):
        self._colIndex = new_colIndex
        self._index = (new_colIndex, self._index[1])
        self._counting_index = (new_colIndex+1, self._index[1]+1)
        self.cellIndexCode = 'C%sR%s' % self._counting_index
        self.cellFrame.colIndex = new_colIndex
        for cell_separator in self.cell_separators:
            cell_separator.colIndex = new_colIndex
        
    @property
    def rowIndex(self):
        return self._rowIndex
    @rowIndex.setter
    def rowIndex(self, new_rowIndex):
        self._rowIndex = new_rowIndex
        self._index = (self._index[0], new_rowIndex)
        self._counting_index = (self._index[0]+1, new_rowIndex+1)
        self.cellIndexCode = 'C%sR%s' % self._counting_index
        self.column.ylocCell = self.column.RowColTitleDist + (self.column.RowDist+2)*new_rowIndex
        self.cellFrame.rowIndex = new_rowIndex
        for cell_separator in self.cell_separators:
            cell_separator.rowIndex = new_rowIndex
        
    @property
    def index(self):
        return self._index
    @index.setter
    def index(self, new_index_tuple):
        self._index = self._index_tuple = self._colIndex, self._rowIndex = new_index_tuple
        self._counting_index = (self._colIndex+1, self._rowIndex+1)
        self.cellIndexCode = 'C%sR%s' % self._counting_index
        if isinstance(self, CellLabel):
            self.column.ylocCell = self.column.RowColTitleDist + (self.column.RowDist+2)*self._rowIndex
        self.cellFrame.colIndex = self._colIndex
        self.cellFrame.rowIndex = self._rowIndex
        for cell_separator in self.cell_separators:
            cell_separator.colIndex = self._colIndex
            cell_separator.rowIndex = self._rowIndex
            
    @property
    def index_tuple(self):
        return self._index
    @index_tuple.setter
    def index_tuple(self, new_index_tuple):
        self.index = new_index_tuple
        
    @property
    def counting_index(self):
        return self._counting_index
    @counting_index.setter
    def counting_index(self, new_counting_index):
        self.index = (new_counting_index[0]-1, new_counting_index[1]-1)
    
    @property
    def bg(self):
        return self._bg
    @bg.setter
    def bg(self, color):
        self._bg = self['bg'] = self['highlightbackground'] = color
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                # print(1903, self.indicator, self.cellFlatIndex, self.sheet.structData[self.indicator] )
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['bg'] = color
        
    @property
    def fg(self):
        return self._fg
    @fg.setter
    def fg(self, color):
        self._fg = self['fg'] = color
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['fg'] = color
        
    @property
    def anchorVar(self):
        return self._anchor
    @anchorVar.setter
    def anchorVar(self, anchor_value):
        self._anchor = self['anchor'] = anchor_value
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['anchor'] = anchor_value
        
    @property
    def justify(self):
        return self._justify
    @justify.setter
    def justify(self, justify_value):
        self._justify = self['justify'] = justify_value
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['justify'] = justify_value
        
    @property
    def top_border(self):
        return self._top_border
    @top_border.setter
    def top_border(self, new_top_border):
        self._top_border = self.moveUp()._bottom_border = installCond = new_top_border
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['top_border'] = int(new_top_border)
        self.switchBorderTop(self_induced=False, install=installCond)
        
    @property
    def bottom_border(self):
        return self._bottom_border
    @bottom_border.setter
    def bottom_border(self, new_bottom_border):
        self._bottom_border = self.moveDown()._top_border = installCond = new_bottom_border
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['bottom_border'] = int(new_bottom_border)
        self.switchBorderBottom(self_induced=False, install=installCond)
        
    @property
    def left_border(self):
        return self._left_border
    @left_border.setter
    def left_border(self, new_left_border):
        self._left_border = self.moveLeft()._right_border = installCond = new_left_border
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['left_border'] = int(new_left_border)
        self.switchBorderLeft(self_induced=False, install=installCond)
        
    @property
    def right_border(self):
        return self._right_border
    @right_border.setter
    def right_border(self, new_right_border):
        self._right_border = self.moveRight()._left_border = installCond = new_right_border
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['right_border'] = int(new_right_border)
        self.switchBorderRight(self_induced=False, install=installCond)
        
    @property
    def width(self):
        return self._width
    @width.setter
    def width(self, new_width):
        self._width = self.cellFrame['width'] = new_width
        for row_separator in self.row_separators:
            row_separator.length = new_width
        
    @property
    def height(self):
        return self._height
    @height.setter
    def height(self, new_height):
        self._height = self.cellFrame['height'] = new_height
        for col_separator in self.col_separators:
            col_separator.length = new_height
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['height'] = new_height
        
    @property
    def font(self):
        return self._font
    @font.setter
    def font(self, new_font):
        self._font = self['font'] = new_font
        if isinstance(self, CellLabel):
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['font'] = new_font
        
    def toggleFormatting(self, attr):
        formatting1 = attr
        formatting2 = f' {formatting1}'
        if type(self.font) == str:
            if formatting2 in self.font:
                self.font = self.font.replace(formatting2, '')
            else:
                self.font += formatting2
        elif hasattr(self.font, '__iter__'):
            if formatting1 in self.font:
                self.font = tuple(fontType for fontType in self.font if fontType != formatting1)
            else:
                self.font += (formatting1,)
        
    def bold(self):
        self.toggleFormatting('bold')
        
    def italic(self):
        self.toggleFormatting('italic')
        
    def underline(self):
        self.toggleFormatting('underline')
        
    def color_reinforce(self):
        # if self.cget('bg') == 'SystemButtonFace':
        self.bg = originalBorderColor
        self.fg = self.cget('fg')
        
    def alignment_reinforce(self):
        self.anchorVar = self.cget('anchor')
        self.justify = self.cget('justify')
        
    def __repr__(self):
        return f'{self.cellIndexCode} ==> {repr(self.content)}'

    def __str__(self):
        return self.__repr__() if not self.default else str(self)

    def __int__(self):
        if is_numeric(self._content):
            return int(self._content)
        elif self._content == '':
            return 0
        else:
            return 1

    def __float__(self):
        if is_numeric(self._content):
            return float(self._content)
        elif self._content == '':
            return 0.0
        else:
            return 1.0
        # return float(self._content)

    def __lt__(self, other):
        return self.index < other.index

    def __gt__(self, other):
        return self.index > other.index

    def __le__(self, other):
        return self.index <= other.index

    def __ge__(self, other):
        return self.index >= other.index

    def __mul__(self, other):
        return float_int_convert( Decimal(str(self._content)) * Decimal(str(other._content)) )
        
    def __div__(self, other):
        return float_int_convert( Decimal(str(self._content)) / Decimal(str(other._content)) )
        
    def __add__(self, other):
        if self.formula_context:
            return float_int_convert( Decimal(str(self._content)) + Decimal(str(other._content)) )
        else: return self.index[0]+other.index[0], self.index[1]+other.index[1]

    def __sub__(self, other):
        if self.formula_context:
            return float_int_convert( Decimal(str(self._content)) - Decimal(str(other._content)) )
        else: return self.index[0]-other.index[0], self.index[1]-other.index[1]

    def __hash__(self):
        return hash(self.index)
        
    def cell_bind(self, debug=None):
        self.sheet.rightClickMenu.activate(widget=self)

    def changeTextColor(self, color):
        self.fg = color
        
    def changeFillColor(self, color):
        self.bg = color
        
    def setLink(self, link, debug=0):
        self.link = link
        
    def moveLeft(self, units=1):
        return self.moveRight(-units)
        
    def moveRight(self, units=1):
        colIndex, rowIndex = self.index
        colIndexTranslated = colIndex+units
        if isinstance(self, ColumnTitle):
            if andExec(colIndexTranslated >= 0, colIndexTranslated < self.sheet.nCols-1):
                return self.sheet.colTitleCells[colIndexTranslated]
            elif colIndexTranslated <= -1:
                return self.sheet.indexColumn.colTitleCell
            elif colIndexTranslated >= self.sheet.nCols-1:
                return self.sheet.colTitleCells[-1]
        elif orExec(isinstance(self, CellLabel), isinstance(self, IndexCell)):
            if andExec(colIndexTranslated >= 0, colIndexTranslated < self.sheet.nCols-1):
                return self.sheet[colIndexTranslated, rowIndex]
            elif colIndexTranslated <= -1:
                return self.sheet.indexColumn[rowIndex]
            elif colIndexTranslated >= self.sheet.nCols-1:
                return self.sheet[self.sheet.nCols-1, rowIndex]
        elif andExec(colIndexTranslated >= 0, colIndexTranslated < self.sheet.nCols-1):
            return self.sheet[colIndexTranslated, rowIndex]
        else:
            return self
        
    def moveDown(self, units=1):
        colIndex, rowIndex = self.index
        rowIndexTranslated = rowIndex+units
        if orExec(isinstance(self, CellLabel), isinstance(self, ColumnTitle)):
            if andExec(rowIndexTranslated >= 0, rowIndexTranslated < self.sheet.nRows-1):
                # print(self.sheet.cells, rowIndexTranslated)
                return self.sheet[colIndex, rowIndexTranslated]
            elif rowIndexTranslated <= -1:
                if colIndex == self.sheet.nCols:
                    return self.sheet.placeholderColumn.colTitleCell
                else:
                    return self.sheet.colTitleCells[colIndex]
            elif rowIndexTranslated >= self.sheet.nRows-1:
                return self.sheet[colIndex, self.sheet.nRows-1]
        elif isinstance(self, IndexCell):
            if andExec(rowIndexTranslated >= 0, rowIndexTranslated < self.sheet.nRows-1):
                return self.sheet.indexColumn[rowIndexTranslated]
            elif rowIndexTranslated <= -1:
                return self.sheet.indexColumn.colTitleCell
            elif rowIndexTranslated >= self.sheet.nRows-1:
                return self.sheet.indexColumn[-1]
        elif andExec(rowIndexTranslated >= 0, rowIndexTranslated < self.sheet.nRows-1):
            return self.sheet[colIndex, rowIndexTranslated]
        else:
            return self
        
    def moveUp(self, units=1):
        return self.moveDown(-units)
        
    def switchBorderLeft(self, self_induced=True, install=True):
        otherCell = self.moveLeft()
        if otherCell == self:
            otherCell = self.sheet.indexColumn[self.index[1]]
        if install:
            otherCell.colBorder.install(self_induced=self_induced)
        else:
            otherCell.colBorder.takeoff(self_induced=self_induced)
        
    def switchBorderRight(self, self_induced=True, install=True):
        if install:
            self.colBorder.install(self_induced=self_induced)
        else:
            self.colBorder.takeoff(self_induced=self_induced)
        
    def switchBorderTop(self, self_induced=True, install=True):
        if isinstance(self, ColumnTitle):
            self.column.rowBorder.install(self_induced=self_induced)
        else:
            otherCell = self.moveUp()
            if install:
                otherCell.rowBorder.install(self_induced=self_induced)
            else:
                otherCell.rowBorder.takeoff(self_induced=self_induced)
        
    def switchBorderBottom(self, self_induced=True, install=True):
        if install:
            self.rowBorder.install(self_induced=self_induced)
        else:
            self.rowBorder.takeoff(self_induced=self_induced)
    
    def toggleCellColor(self, eventInherited=None, select=True, selectMultipleInOne=True, auto=True, debug=False, firstCellSelect=None):
        if debug:
            print(f'too_lonely_to_unmark =', too_lonely_to_unmark)
            if selectMultipleInOne:
                print('self =', self)
                print('firstCellSelect =', firstCellSelect)
        
        selectedCellsSet = self.sheet.selectedCellsSet
        if selectMultipleInOne:
            if select:
                self['bg'] = self['highlightbackground'] = mix_colors([self.bg, '#4F85D5'])
            else:
                self['bg'] = self['highlightbackground'] = self.bg
            
        # else:
            # if select:
                # # Turn on last range
                # if selectedCellsSet and not auto:
                    # selectedCellsSet[-1].switchOnSelector()
            # elif not too_lonely_to_unmark:
                # # Turn off last range
                # if selectedCellsSet and not auto:
                    # selectedCellsSet[-1].switchOffSelector()
        
    def generateIndexList(self):
        """ Used for Insert/Delete Rows/Columns """
        return [self.index[0]], [self.index[1]]
        
    # def switchOffSelector(self):
        # self.rowSelector.takeoff()
        # self.colSelector.takeoff()
        
    def generateRange(self, single_cell_clicked=True):
        start, end = self.index
        return CellRange( ColRange(start, start) , RowRange(end, end) , sheet=self.sheet , originCell=self , single_cell_clicked=single_cell_clicked )
        
    def within(self, cell_range):
        col_range, row_range = cell_range.col_range, cell_range.row_range
        colIndex, rowIndex = self.index
        return andExec(col_range.start <= colIndex, colIndex <= col_range.stop, row_range.start <= rowIndex, rowIndex <= row_range.stop)
        
    @property
    def ranger(self):
        return self.generateRange()
        
    @property
    def switchOnOuterBorder(self):
        return self.ranger.switchOnOuterBorder
        
    @property
    def switchOffOuterBorder(self):
        return self.ranger.switchOffOuterBorder
        
    @property
    def toggleOuterBorder(self):
        return self.ranger.toggleOuterBorder
        
    @property
    def switchOnSelector(self):
        return self.ranger.switchOnSelector
        
    @property
    def switchOffSelector(self):
        return self.ranger.switchOffSelector
        
    def __getitem__(self, index):
        if type(index) == int:
            return self.index[index]
        
    
class CellLabel(tk.Label, Cell):
    formula = ''
    formula_context = False
    
    def __init__(self, parent, book, index, column, sheet, col_width, row_height, colTitle=None, is_placeholder=False, indirect_inheritance=False, inserted=False, *args, **kwargs):
        Cell.__init__(self, parent, book, index, column, sheet, col_width, row_height, is_placeholder=is_placeholder, indirect_inheritance=indirect_inheritance, inserted=inserted)
        
        tk.Label.__init__(self, self.cellFrame, bd=0, fg="#000000", borderwidth=0, *args, **kwargs)
        
        self.configure(cursor='plus')
        self.cell_bind()
        
        self.is_placeholder = is_placeholder
        
        self._link = None
        self._formula = ''
        
        # 0       : fill (bg)
        # 1       : font (fg)
        # 2,3     : alignment (anchor, justify)
        # 4       : hyperlink
        # 5,6,7,8 : border (top, bottom, left, right)
        # 9       : row_height
        # 10      : content (plain/result)
        # 11      : content (formula)
        # 12      : font
        colIndex, rowIndex = self.index
        self.row_height = row_height
        self.indicator = indicator = 'colTitleCells' if type(self) == ColumnTitle else 'cells'
        self.cellFlatIndex = cellFlatIndex = colIndex if type(self) == ColumnTitle else colIndex*self.sheet.nRows+rowIndex
        
        if andExec(not self.is_placeholder, not self.column.is_placeholder):
            if isinstance(self, ColumnTitle):
                print(2295, self, (not self.is_placeholder, not self.column.is_placeholder), andExec(not self.is_placeholder, not self.column.is_placeholder) )
            self.autoInit()
                
        if orExec(self.book.extension != '.qutable', self.sheet.ready):
            # if andExec(not self.is_placeholder, not self.column.is_placeholder):
                # self.autoInit()
            if isinstance(self, ColumnTitle):
                self.formula = colTitle
            elif isinstance(self, CellLabel):
                self.formula = ''
        
        elif not column.is_placeholder:
            if isinstance(self, ColumnTitle):
                self.formula = colTitle
            elif isinstance(self, CellLabel):
                self._content = ''
        
        self.color_reinforce()
        self.alignment_reinforce()
        self._font = self['font']
        self['wraplength'] = col_width
        removeBinds(self, 'all')
        
    def autoInit(self):
        # [None]*12
        # try:  # if self.cellFlatIndex < len(self.sheet.structData[self.indicator]):
            # self.bg, self.fg, self.anchorVar, self.justify, self.link, self.row_height, self.content, self.formula, self.font = self['bg'], self['fg'], self['anchor'], self['justify'], '', self['height'], self['text'], '', self['font']
            # self.structData = {'bg':self.bg, 'fg':self.fg, 'anchor':self.anchorVar, 'justify':self.justify, 'link':self.link, 'top_border':int(self.top_border), 'bottom_border':int(self.bottom_border), 'left_border':int(self.left_border), 'right_border':int(self.right_border), 'height':self.row_height, 'content':self.content, 'formula':self.formula, 'font':self.font}
            # # self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)] = self.structData
            # # return self.structData
        # except KeyError:
            # self._bg, self._fg, self._anchorVar, self._justify, self._link, self._row_height, self._content, self._formula, self._font = self['bg'], self['fg'], self['anchor'], self['justify'], '', self['height'], self['text'], '', self['font']
            # self.structData = {'bg':self._bg, 'fg':self._fg, 'anchor':self._anchorVar, 'justify':self._justify, 'link':self._link, 'top_border':int(self._top_border), 'bottom_border':int(self._bottom_border), 'left_border':int(self._left_border), 'right_border':int(self._right_border), 'height':self._row_height, 'content':self._content, 'formula':self._formula, 'font':self._font}
            # if andExec(not self.is_placeholder, self.sheet.ready):
                # self.sheet.structData[self.indicator].append( self.structData )
                # print(2327, self, self.sheet.structData[self.indicator] )
                # print(2327, self, self.structData)
                # if isinstance(self, ColumnTitle):
                    # self.sheet.structData[self.indicator].extend( [{}] )
                # elif type(self) == CellLabel:
                    # self.sheet.structData[self.indicator].extend( [{} for i2 in range(self.sheet.nRows)] )
                    
        try:
            self._bg, self._fg, self._anchorVar, self._justify, self._link, self._row_height, self._content, self._formula, self._font = self['bg'], self['fg'], self['anchor'], self['justify'], '', self['height'], self['text'], '', self['font']
            self.structData = {'bg':self._bg, 'fg':self._fg, 'anchor':self._anchorVar, 'justify':self._justify, 'link':self._link, 'top_border':int(self._top_border), 'bottom_border':int(self._bottom_border), 'left_border':int(self._left_border), 'right_border':int(self._right_border), 'height':self._row_height, 'content':self._content, 'formula':self._formula, 'font':self._font}
            # print(2341, self.sheet.structData[self.indicator], str(self.index[0])) if isinstance(self, ColumnTitle) else None
            self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)] = self.structData
            return self.structData
        except Exception as e:
            print(format_exc(), file=sys.stderr)
        
    def autoFormat(self, structData):
        # print(2337, structData)
        self.bg, self.fg, self.anchorVar, self.justify, self.link, self.top_border, self.bottom_border, self.left_border, self.right_border, self.row_height, self.content, self.formula, self.font = structData['bg'], structData['fg'], structData['anchor'], structData['justify'], structData['link'], structData['top_border'], structData['bottom_border'], structData['left_border'], structData['right_border'], structData['height'], structData['content'], structData['formula'], structData['font']
        self.structData = self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)] = structData
        if isinstance(self, ColumnTitle):
            if not structData['top_border']:
                self.column.rowBorder.takeoff()
            
            # print(structData)
            # print(self.structData)
        
    @property
    def link(self):
        return self._link
    @link.setter
    def link(self, link):
        if type(self) == CellLabel:
            if not self.column.is_placeholder:
                # self.cellOpenPyxl.hyperlink = self._link = self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['link'] = link
                if link:
                    self.bind('<ButtonRelease>', lambda event: self.sheet.clickAction(event=event, link=link) )
        else:
            self._link = link
            if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
                self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['link'] = link
        if link:
            self.configure(cursor='hand2')
        
    def is_numeric(self):
        return is_numeric(self.content)
        
    def parse_qutable_expr(self):
        qutableVars = createVarList( self.formula_expr )
        try:
            result = eval(self.formula_expr, {var:self.sheet[var] for var in qutableVars})
        except InvalidOperation:
            result = 'Error!'
        return result
        
    def convert_excel_to_qutable_formula(self):
        self._formula = '='+convert_excel_to_qutable_formula(self.formula_expr, self.sheet.have_headers)
        if andExec(self.indicator, isinstance(self.cellFlatIndex, numbers.Number), not self.column.is_placeholder):
            self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['formula'] = self._formula
        
    def parse_excel_expr(self):
        def Symbol(index):
            # print(index, self.sheet.sheetOpenPyxl[index].value)
            return Decimal(str(self.sheet.sheetOpenPyxl[index].value))
        def Add(*data):
            return sum(data)
        def Mul(*data):
            return functools.reduce(lambda n1,n2: n1*n2, data)
        def Pow(*data):
            return functools.reduce(lambda n1,n2: n1**n2, data)
        exprStruct = self.sheet.parse_expr_with_sheet(self.formula_expr)
        result = float_int_convert(eval(exprStruct))
        self.convert_excel_to_qutable_formula()
        return result
        
    @property
    def formula(self):
        return self._formula
    @formula.setter
    def formula(self, _formula):
        if isinstance(_formula, tuple):
            _formula, formulaEntryNOTReturnStatus = _formula
            if andExec(not _formula, formulaEntryNOTReturnStatus):
                # self.book.formulaEntryReturnStatus = False
                return
        else:
            _formula = str(_formula)
            
        self._formula = _formula
        # if self.book.ready:
            # self.book.formulaEntry.delete(0, 'end')
            # self.book.formulaEntry.insert(0, _formula)
        self.formula_expr = self.formula.strip('=')
        if andExec(self.indicator, not self.is_placeholder, not self.column.is_placeholder):
            self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)]['formula'] = str(_formula)
        # if isinstance(self, ColumnTitle):
            # pass
        # print(2247, _formula)
        
        self.parse_formula_determination()
        
    def parse_formula_determination(self):
        if self._formula.startswith('='):
            if andExec(self.book.extension.startswith('.xls'), self.sheet.init):
                self.content = self.parse_excel_expr()
            else:
                self.content = self.parse_qutable_expr()
        else:
            self.content = self._formula
        
    def __repr__(self):
        return Cell.__repr__(self)
        
    def __str__(self):
        return Cell.__str__(self)
        
    def deepcopy(self):
        cellConfig = self.configure()
        cellPackConfig = self.pack_info()
        cellFrameGridConfig = self.cellFrame.grid_info()
        del cellPackConfig['in']
        del cellFrameGridConfig['in']
        newCell = self.__class__(self.parent, self.book, self.index, self.column, self.sheet)
        newCell.formula = self.formula
        # newCell.content = self.content
        for config_property in cellConfig:
            newCell[config_property] = self[config_property]
        for cellPackConfig_property in cellPackConfig:
            newCell.pack_configure({cellPackConfig_property:cellPackConfig[cellPackConfig_property]})
        for cellFrameGridConfig_property in cellFrameGridConfig:
            newCell.grid_configure({cellFrameGridConfig_property:cellFrameGridConfig[cellFrameGridConfig_property]})
        return newCell
        
    def regenerateCell(self):
        newCell = self.deepcopy()
        colIndex, rowIndex = self.index
        self.column.cells[rowIndex] = newCell
        self.sheet.cells[colIndex*self.sheet.nRows+rowIndex] = newCell
        
    def indexDiff(self, other):
        return abs(self.index[0]-other.index[0]), abs(self.index[1]-other.index[1])
        
    def destroy(self, event=None):
        super().destroy()
        self.sheet.indexColumn.cells.remove(self)
        
    def cell_bind(self, debug=None):
        Cell.cell_bind(self, debug=debug)
        
        if debug: print('self.sheet.currentCell =', self.sheet.currentCell)
        
        self.selected = False
        
        self.selectionMethod = '<Button-1>'
        # self.entry
        # removeBinds(self.entry, 'Entry')
        
        self.dragRight = None
        self.dragDown = None
        
        for cell in self.cell_parts:
            cell.bind('<Control-1>', self.toggleCell )
            cell.bind('<Control-3>', self.toggleCell )
            cell.bind('<Button-1>', lambda event: ( self.focusOnCell(event=event) , None ) )
            cell.bind('<MouseWheel>', self.sheet.mouseWheel )
            cell.bind('<Double-1>', lambda event: ( self.edit_start(event=event) , None ) )
            cell.bind('<B1-Motion>', lambda event: ( self.selectbyDragging(event=event) , None ) )
            # Cursor Affairs
            cell.bind('<Enter>', self.hover )
            cell.bind('<Leave>', self.leave )
            
    def edit_start(self, event=None, text=None):
        self.sheet.edit_start_type = type(self)
        
        # if self.column.is_automated:
            # messagebox.showerror("Column Automated", "Column has an automated formula, so it cannot be edited.", parent=self.book.window)
            # return
        
        if isinstance(self, ColumnTitle):
            self.sheet.currentCell = self
        
        # Moving across cells using keyboard arrows
        self.book.window.unbind('<Key>')
        self.book.window.unbind('<KeyRelease>')
        
        # for keyboardCMD in self.book.cellKeyBinds:
            # self.book.window.unbind(keyboardCMD)
        
        self.sheet.entry = entry = self.entry = tk.Entry(self.cellFrame, borderwidth=0, font=self['font'])
        
        # entry['highlightbackground'] = 
        entry['bg'] = self.cget('bg')
        entry['fg'] = self.cget('fg')
        
        entry.bind('<MouseWheel>', self.sheet.mouseWheel )
        
        entry.bind('<Tab>', lambda event: ( self.edit_stop(event=event), self.sheet.MoveMarkByCoord(+1, 0) ) )
        entry.bind('<Shift-Tab>', lambda event: ( self.edit_stop(event=event), self.sheet.MoveMarkByCoord(-1, 0) ) )
        entry.bind('<Shift-Return>', lambda event: ( self.edit_stop(event=event), self.sheet.MoveMarkByCoord(0, -1) ) )
        entry.bind('<Return>', lambda event: ( self.edit_stop(event=event), self.sheet.MoveMarkByCoord(0, +1) ) )
        entry.bind('<Escape>', self.edit_stop)
        entry.bind('<FocusOut>', lambda event: ( self.edit_stop(event=event, focusOut=True), self.sheet.resetQStats() ) )
        
        # self.cell_pack_info = self.pack_info()
        self.pack_forget()
        
        entry.place(x=0, y=0, relwidth=1.0, relheight=1.0)
        entry.focus_set()
        
        # Detect whether the keyboard cursor is in the left end or right end of the Entry Box
        entry.bind('<Right>', self.MoveMarkByCoordEntryRight)
        entry.bind('<Left>',  self.MoveMarkByCoordEntryLeft)
        
        # Move to other cells on top or bottom of current cell
        entry.bind('<Up>', lambda event: self.sheet.MoveMarkByCoord(0, +1) )
        entry.bind('<Down>',  lambda event: self.sheet.MoveMarkByCoord(0, -1) )
        
        if text is None:
            text = self.formula
            
        entry.insert('0', text)

    def edit_stop(self, event=None, focusOut=False):
        # print('edit_stop')
        # print(type(self))
        entry = self.entry
        entryVal = entry.get()
        self.column.previous_cont = self.getContent()
        # notYetBind = True
            
        entry.destroy()
        self.pack(side='top', fill='both', expand=True)
        
        self.sheet.entry = self.entry = None
        
        if andExec( self.is_placeholder , self.column.is_placeholder):
            # rightMostCol = self.sheet.insertColRightMost(nColsSelected=1)
            # self.sheet.insertRowBottomMost(nRowsSelected=1)
            
            rightMostCol, bottomMostRow = self.sheet.reformSheetInner(self.sheet.nCols+1, self.sheet.nRows+1)
            if isinstance(self, ColumnTitle):
                rightMostCol.colTitle = entryVal
            elif isinstance(self, CellLabel):
                rightMostCol.colTitle = 'NewCol'
            
            colIndex, rowIndex = self.index
            # self.sheet[colIndex, rowIndex].formula = entryVal
            # rightMostCol[rowIndex].switchOffSelector()
            # rightMostCol[rowIndex-1].focusOnCell()
            # rightMostCol.current_cont = entryVal
            rightMostCol[rowIndex-1].formula = entryVal
            self.sheet.resetQStats()
            self.sheet.reExecFormula()
            
        elif self.is_placeholder:
            self.sheet.insertRowBottomMost(nRowsSelected=1)
            colIndex, rowIndex = self.index
            self.sheet[colIndex, rowIndex-1].focusOnCell()
            self.sheet[colIndex, rowIndex].switchOffSelector()
            self.sheet[colIndex, rowIndex-1].formula = entryVal
            self.column.current_cont = entryVal
            self.sheet.resetQStats()
            self.sheet.reExecFormula()
            
        elif self.column.is_placeholder:
            rightMostCol = self.sheet.insertColRightMost(nColsSelected=1)
            rowIndex = self.index[1]
            if isinstance(self, ColumnTitle):
                rightMostCol.colTitle = entryVal
            elif isinstance(self, CellLabel):
                rightMostCol[rowIndex].formula = rightMostCol.current_cont = entryVal
                rightMostCol.colTitle = 'NewCol'
            self.sheet.resetQStats()
            self.sheet.reExecFormula()
            
        elif event:
            # print(event.keysym)
            if orExec(event.keysym != 'Escape', focusOut):
                self.replaceFormula( entryVal )
                if isinstance(self, ColumnTitle):
                    self.column.colTitle = entryVal
                elif isinstance(self, CellLabel):
                    self.column.current_cont = entryVal
                    # print(self.column.current_cont, self.column.previous_cont)
                    self.sheet.resetQStats()
                    self.sheet.reExecFormula()
        
        elif focusOut:
            self.replaceFormula( entryVal )
            if isinstance(self, CellLabel):
                self.column.current_cont = entryVal
                # print(self.column.current_cont, self.column.previous_cont)
                self.sheet.resetQStats()
                self.sheet.reExecFormula()
            
        else:
            self.replaceFormula( entryVal )
        
        # if notYetBind:
            # notYetBind = False
            
        # Moving across cells using keyboard arrows
        self.book.window.bind('<Key>', lambda event: self.sheet.keyHandle(event=event) )
        self.book.window.bind('<KeyRelease>', lambda event: self.sheet.keyReleaseHandle(event=event) )
        
        # print(entry.winfo_width())
        # print(entry.index('insert'))
        
        self.sheet.actionCollect(action=f'Edit Cell - {self.cellIndexCode}')
        
        # self.pack(self.cell_pack_info)
        
        # for keyboardCMD, cellLambdaFunction in self.book.cellFunctions:
            # self.book.window.bind(keyboardCMD, cellLambdaFunction)
        
    def hover(self, event=None):
        if self.link:
            event.widget.configure(cursor='hand2')
        
    def leave(self, event=None):
        if self.link:
            event.widget.configure(cursor='plus')

    def generateColRange(self):
        return ColRange(self.index[0], self.index[0], sheet=self.sheet, book=self.book)

    def generateRowRange(self):
        return RowRange(self.index[1], self.index[1], sheet=self.sheet, book=self.book)

    def getEntireColumn(self):
        return self.generateColRange().generateRange()

    def getEntireRow(self):
        return self.generateRowRange().generateRange()
        
    def toggleAlignH(self, HPos):
        anchor = self.anchorVar
        new_anchor = (HPos, alignmentReverse[anchor][1])
        self.anchorVar, self.justify = alignmentRule[new_anchor][self.is_numeric()]
        self.sheet.switchAlignmentBtn()
        
    def toggleAlignLeft(self):
        self.toggleAlignH('left')
        
    def toggleAlignCenterH(self):
        self.toggleAlignH('center')
        
    def toggleAlignRight(self):
        self.toggleAlignH('right')
        
    def toggleAlignV(self, VPos):
        anchor = self.anchorVar
        new_anchor = (alignmentReverse[anchor][0], VPos)
        self.anchorVar, self.justify = alignmentRule[new_anchor][self.is_numeric()]
        self.sheet.switchAlignmentBtn()
        
    def toggleAlignTop(self):
        self.toggleAlignV('top')
        
    def toggleAlignMiddleV(self):
        self.toggleAlignV('center')
        
    def toggleAlignBottom(self):
        self.toggleAlignV('bottom')
        
    def changeColorBtn(self):
        # Change color in TextColor_btn and FillColor_btn
        menu = self.book.SuperMenuWidget
        
        TextColor_btn_img_array = np.array(menu.TextColor_btn_imgPIL)
        TextColor_btn_img_array[20:27, 2:27] = hex_to_rgba( self['fg'] )
        
        menu.TextColor_btn_imgPIL = Image.fromarray( TextColor_btn_img_array )
        menu.TextColor_btn_img = ImageTk.PhotoImage( menu.TextColor_btn_imgPIL , master=self.book.window )
        menu.TextColor_btn['image'] = menu.TextColor_btn_img
        
        
        FillColor_btn_img_array = np.array(menu.FillColor_btn_imgPIL)
        FillColor_btn_img_array[23:27, 2:27] = hex_to_rgba( self['bg'] )
        
        menu.FillColor_btn_imgPIL = Image.fromarray( FillColor_btn_img_array )
        menu.FillColor_btn_img = ImageTk.PhotoImage( menu.FillColor_btn_imgPIL , master=self.book.window )
        menu.FillColor_btn['image'] = menu.FillColor_btn_img
    
    # def toggleCell(self, event=None):
        # self.toggleCell(event=event, auto=False)
        
    def toggleCell(self, event=None, eventInherited=None, select=True, selectMultipleInOne=False, justChangeColor=False, debug=False, firstCellSelect=None):
        # self.toggleCellColor(eventInherited=eventInherited, select=select, selectMultipleInOne=selectMultipleInOne, debug=debug, firstCellSelect=firstCellSelect)
        
        self_range = self.generateRange()
        too_lonely_to_unmark = andExec(self_range.is_one_cell(), len(self.sheet.selectedCellsSet) <= 1, self.selected)
        
        if debug: print('selectMultipleInOne =', selectMultipleInOne)
        
        # if not justChangeColor:
        if debug:
            print(f'{self.cellIndexCode}.selected =', self.selected)
        
        if event:
            if event.widget == self:
                self.event_state, self.event_num, self.event_widget = event.state, event.num, event.widget
            
        if andExec(self_range in self.sheet.selectedCellsSet, not self.sheet.control_is_pressed):
            self_range.switchOffSelector()
            for cell_range in self.sheet.selectedCellsSet:
                cell_range.switchOnSelector()
        else:
            self_range.switchOnSelector()
        self.sheet.switchBtnsCombo()
        
        if not self.selected:
            self.selected = True
            if not selectMultipleInOne:
                self.sheet.selectedCellsSet.add( self_range )
        elif not too_lonely_to_unmark:
            self.selected = False
            if not selectMultipleInOne:
                if debug: print(f'{self.cellIndexCode}_range =', self_range)
                self.sheet.selectedCellsSet.remove( self_range )
        
        if debug:
            print()
        
        return selectMultipleInOne
        
    def focusOnCell(self, event=None, debug=None):
        if debug: print(event.state, event.num)
        
        # Provide Selection Mark on selected cell
        colIndex, rowIndex = self.index
        self.sheet.SelectionMarkByCoord(colIndex, rowIndex, event=event)
        
        if debug: print(self, self.index)
        
        if event is None:
            # print('focusOnCell executed', event.state, '\t', event.num)
            self.sheet.selectedCellsSet.start( self.sheet.currentCell )
        
        elif orExec(event.state != 264, event.state != 268, event.state != 12):
            # print('focusOnCell executed', event.state, '\t', event.num)
            self.sheet.selectedCellsSet.start( self.sheet.currentCell )
        
        if event:
            self.event_state, self.event_num, self.event_widget = event.state, event.num, event.widget
                
        if debug:
            print('self.sheet.selectedCellsSet =', self.sheet.selectedCellsSet)
            print()
            
        self.changeColorBtn()
        self.sheet.switchBtnsCombo()
        
    def selectbyDragging(self, event=None, col_range=None, row_range=None, col_start_end=None, widgetEnd=None, debug=0):
        if event:
            x, y = event.x_root, event.y_root
        
        widgetStart = self
        if not widgetEnd:
            widgetEnd = self.winfo_containing(x, y)
            
        if andExec(widgetEnd != widgetStart, widgetEnd != self.widgetEnd_prev_debug, event, debug):
            print(f'widgetEnd, x, y = {repr(widgetEnd)}, {x}, {y}')
            self.widgetEnd_prev_debug = widgetEnd
        
        if isinstance(widgetEnd, LineRange):
            widgetEnd = LineRange.parent
            
        if isinstance(widgetEnd, CellLabel):
            if andExec(widgetStart == widgetEnd, not self.sheet.control_is_pressed):
                widgetStart.focusOnCell(event=event)
                
            elif orExec(widgetEnd != self.widgetEnd_prev, andExec(widgetStart == widgetEnd, self.sheet.control_is_pressed)):
                widgetStart.switchOffSelector()
                
                self.dragRight = widgetEnd.index[0] > widgetStart.index[0]
                self.dragDown = widgetEnd.index[1] > widgetStart.index[1]
                
                self.widgetEnd = self.sheet.lastCell = widgetEnd
                
                # if not col_range:
                startCol, endCol = widgetStart.index_tuple[0], widgetEnd.index_tuple[0]
                if endCol < startCol:
                    startCol, endColMod = endCol, startCol
                else:
                    startCol, endColMod = startCol, endCol
                col_range = ColRange(startCol, endColMod)
                
                if debug:
                    print(f'startCol, endCol, endColMod = {startCol}, {endCol}, {endColMod}')
                    print('col_range =', col_range)
                    print()
                
                # if not row_range:
                startRow, endRow = widgetStart.index_tuple[1], widgetEnd.index_tuple[1]
                if endRow < startRow:
                    startRow, endRowMod = endRow, startRow
                else:
                    startRow, endRowMod = startRow, endRow
                row_range = RowRange(startRow, endRowMod)
                if debug:
                    print(f'startRow, endRow, endRowMod = {startRow}, {endRow}, {endRowMod}')
                    print('row_range =', row_range)
                    print()
                
                current_range = CellRange(col_range, row_range, self.book, self.sheet, originCell=widgetStart)
                if debug: print('current_range =', current_range)
                
                widgetEnd_prev = self.widgetEnd_prev
                
                if debug:
                    # print('dir(self) =', dir(self))
                    print('widgetStart =', widgetStart)
                    print('widgetEnd_prev =', widgetEnd_prev)
                    print('widgetEnd =', widgetEnd)
                    print('widgetEnd < widgetEnd_prev =', widgetEnd < widgetEnd_prev)
                    print('widgetStart.index_tuple =', widgetStart.index_tuple)
                    print('widgetEnd.index_tuple =', widgetEnd.index_tuple)
                
                if len(self.sheet.selectedCellsSet) >= 1:
                    last_range = self.sheet.selectedCellsSet[-1]
                    last_range_cond = not last_range.within(current_range)
                else:
                    last_range_cond = False
                    
                if andExec(last_range_cond, self.sheet.control_is_pressed, self.sheet.dragFirstTouch):
                    self.sheet.selectedCellsSet.add ( current_range )
                    self.sheet.dragFirstTouch = False
                else:
                    if debug: print('Executed: self.sheet.selectedCellsSet[-1] = current_range')
                    self.sheet.selectedCellsSet[-1] = current_range
                    
                if debug:
                    print('self.sheet.selectedCellsSet =', self.sheet.selectedCellsSet)
                    print()
                
                # Each has different firstCellSelect
                firstCellSelect = widgetStart
                current_range.toggleCellColor(eventInherited=event, select=True, previous_range=self.sheet.previous_range)
                self.sheet.currentCell.toggleCellColor(eventInherited=event, select=False)
                
                col_diff_w_start, row_diff_w_start = widgetEnd - widgetStart
                col_diff_w_prev, row_diff_w_prev = widgetEnd - widgetEnd_prev
                # print('col_diff_w_start, row_diff_w_start =', col_diff_w_start, row_diff_w_start)
                # print('col_diff_w_prev, row_diff_w_prev =', col_diff_w_prev, row_diff_w_prev)
                
                cond_col_opposite_dir = orExec( andExec(col_diff_w_start < 0, col_diff_w_prev > 0), (col_diff_w_start > 0, col_diff_w_prev < 0) )
                cond_row_opposite_dir = orExec( andExec(row_diff_w_start < 0, row_diff_w_prev > 0), (row_diff_w_start > 0, row_diff_w_prev < 0) )
                
                if debug:
                    print('cond_col_opposite_dir =', cond_col_opposite_dir)
                    print('cond_row_opposite_dir =', cond_row_opposite_dir)
                    print()
                
                prevSelectedCells = self.sheet.cells_to_CellRange(widgetStart, widgetEnd_prev,
                                                        cond_col_opposite_dir=cond_col_opposite_dir,
                                                        cond_row_opposite_dir=cond_row_opposite_dir)
                                                        
                currentSelectedCells = self.sheet.cells_to_CellRange(widgetStart, widgetEnd,
                                                        cond_col_opposite_dir=cond_col_opposite_dir,
                                                        cond_row_opposite_dir=cond_row_opposite_dir)
                
                if debug:
                    print('prevSelectedCells =', prevSelectedCells)
                    print('currentSelectedCells =', currentSelectedCells)
                
                # try:
                unSelectedCellsRanges = prevSelectedCells - currentSelectedCells
                # except:
                    # print('prevSelectedCells =', prevSelectedCells)
                    # print('currentSelectedCells =', currentSelectedCells)
                    
                if debug: print('unSelectedCellsRanges =', unSelectedCellsRanges)
                
                
                unSelectedCellsRanges.toggleCellColor(eventInherited=event, select=False, firstCellSelect=firstCellSelect)
                
                # if isinstance(unSelectedCellsRanges, EmptyCellRange):
                    # if isinstance(self.sheet.previous_range, CellRange):
                        # # print('self.sheet.previous_range =', self.sheet.previous_range)
                        # if orExec(col_diff_w_prev == 0, row_diff_w_prev == 0):
                            # if col_diff_w_prev != 0:
                                # self.sheet.previous_range.switchOffSelector(to_end_list=[col_diff_w_prev > 0], vertical_list=[1], debug=0)
                            # elif row_diff_w_prev != 0:
                                # self.sheet.previous_range.switchOffSelector(to_end_list=[row_diff_w_prev > 0], vertical_list=[0], debug=0)
                        # else:
                            # to_end_list = [col_diff_w_prev > 0, row_diff_w_prev > 0] if col_diff_w_prev == row_diff_w_prev else [col_diff_w_prev < 0, row_diff_w_prev < 0]
                            # self.sheet.previous_range.switchOffSelector(to_end_list=to_end_list, vertical_list=vertical_list_standard, debug=0)
                # elif isinstance(unSelectedCellsRanges, CellSet):
                    # # print('unSelectedCellsRanges =', unSelectedCellsRanges)
                    # # print('type(unSelectedCellsRanges) =', type(unSelectedCellsRanges))
                    # # print()
                    # if orExec(col_diff_w_prev == 0, row_diff_w_prev == 0):
                        # if col_diff_w_prev != 0:
                            # function = lambda cell_range: cell_range.switchOffSelector(to_end_list=[col_diff_w_prev < 0], vertical_list=[1], debug=0) if isinstance(cell_range, CellRange) else None
                        # elif row_diff_w_prev != 0:
                            # function = lambda cell_range: cell_range.switchOffSelector(to_end_list=[row_diff_w_prev < 0], vertical_list=[0], debug=0) if isinstance(cell_range, CellRange) else None
                    # else:
                        # to_end_list = [col_diff_w_prev < 0, row_diff_w_prev < 0] if col_diff_w_prev == row_diff_w_prev else [col_diff_w_prev > 0, row_diff_w_prev > 0]
                        # function = lambda cell_range: cell_range.switchOffSelector(to_end_list=to_end_list, vertical_list=vertical_list_standard, debug=0) if isinstance(cell_range, CellRange) else None
                    # self.sheet.functionRemoval = function
                    # unSelectedCellsRanges.map(function, cell_absorbant=False)
                    
                if debug:
                    print()
                    print()
                
                self.sheet.previous_range = current_range
                
            self.widgetEnd_prev = widgetEnd
                
        if event:
            self.event_state, self.event_num, self.event_widget = event.state, event.num, event.widget
            
        self.sheet.switchBtnsCombo()
        
    def releaseAfterDragging(self, event=None):
        self.sheet.selectedCellsSet.add(None)
        if event:
            self.event_state, self.event_num, self.event_widget = event.state, event.num, event.widget
        self.sheet.dragFirstTouch = True
        
    def MoveMarkByCoordEntryRight(self, event=None):
        entryCursor = self.entry.index('insert')
        if entryCursor == len(self.entry.get()):
            self.sheet.MoveMarkByCoord(+1, 0)
        
    def MoveMarkByCoordEntryLeft(self, event=None):
        entryCursor = self.entry.index('insert')
        if entryCursor == 0:
            self.sheet.MoveMarkByCoord(-1, 0)
        
    def setFormulaContext(self, formula_context):
        self.formula_context = formula_context
        
    def getFormula(self):
        return self.formula
        
    def getContent(self):
        return self.content
        
    def replaceFormula(self, formula):
        self.formula = formula
        
    def recalibrateIndex(self, col_translation=None, row_translation=None, new_col_index=None, new_row_index=None):
        index = self.index
        
        if isinstance(col_translation, numbers.Number):
            self.colIndex += col_translation
        elif not isinstance(new_col_index, numbers.Number):
            self.colIndex = new_col_index
            
        if isinstance(row_translation, numbers.Number):
            self.rowIndex += row_translation
        elif isinstance(new_row_index, numbers.Number):
            self.rowIndex = new_row_index
            
        # print('new_col_index =', new_col_index)
        # print('new_row_index =', new_row_index)
        
        # cell_part = self.cellFrame
        # grid_info = cell_part.grid_info()
        # row    = grid_info['row']
        # column = grid_info['column'] + col_translation*2 if isinstance(col_translation, numbers.Number) else grid_info['column']
        # cell_part.grid_configure(row=row, column=column)
        
        # if isinstance(col_translation, numbers.Number):
            # for cell_separator in self.cell_separators:
                # cell_separator.colIndex += col_translation
                
        
class IndexCell(tk.Label, Cell):
    def __init__(self, parent, book, index, column, sheet, col_width, row_height, is_placeholder=False, indirect_inheritance=False, inserted=False, *args, **kwargs):
        Cell.__init__(self, parent, book, index, column, sheet, col_width, row_height, is_placeholder=is_placeholder, indirect_inheritance=indirect_inheritance, inserted=inserted)
        tk.Label.__init__(self, self.cellFrame, bd=0, fg="#000000", background='#FFFFFF', borderwidth=0, *args, **kwargs)
        # tk.Label.__init__(self, sheet.sheetInner, bd=0, fg="#000000", background='#FFFFFF', borderwidth=0, *args, **kwargs)
        # tk.Label.__init__(self, parent, bd=0, fg="#000000", background='#FFFFFF', borderwidth=0, *args, **kwargs)
        
        self.color_reinforce()
        self.alignment_reinforce()
        self._font = self['font']
        self['wraplength'] = col_width
        removeBinds(self, 'all')
        
        # self.index = self.index_tuple = index
        # self.counting_index = (index[0]+1, index[1]+1)
        self.rowIndex = self.index[1]
        
        self.book = book
        self.sheet = sheet
        self.column = column
        self.cell = self
        
        self.content = ''
        
        self.selected = False
        
        self.cell_bind()
        
        # self.bg = self['bg'] = originalBorderColor
        # self.bg = self['bg'] = 'WhiteSmoke'
        
        self.configure(cursor='plus')
        
    def __repr__(self):
        return f'Index #{self.rowIndex+1}'
        
    def __str__(self):
        return Cell.__str__(self)
        
    def cell_bind(self, debug=None):
        Cell.cell_bind(self, debug=debug)
        for cell in self.cell_parts:
            cell.bind('<Control-1>', self.toggleRowCells)
            cell.bind('<Control-3>', self.toggleRowCells)
            cell.bind('<Button-1>', self.focusOnRowCells)
            cell.bind('<B1-Motion>', self.selectbyDragging)
        
    def replaceFormula(self, content):
        self['text'] = self.content = content
        
    def toggleRowCells(self, event=None, auto=False):
        row = self.sheet[:, self.rowIndex]
        self.selected = not self.selected
        for cell in row:
            cell.toggleCell(event=event, select=self.selected, selectMultipleInOne=True, auto=True)
        if not auto:
            self.sheet.switchBtnsCombo()
        
    def focusOnRowCells(self, event=None):
        self.sheet.removeSelectionMark(event=event)
        col_range = ColRange(0, self.sheet.nCols-1)
        row_range = RowRange(self.rowIndex, self.rowIndex)
        current_range = CellRange(col_range, row_range, self.book, self.sheet)
        self.sheet.selectedCellsSet.start( current_range )
        current_range.toggleCellColor(eventInherited=event, select=True)
        self.sheet.previous_range = current_range
        self.sheet.currentCell = current_range.getStartingCell()
        self.sheet.lastCell    = current_range.getLastCell()
        self.sheet.currentCell.toggleCellColor(eventInherited=event, select=False)
        
        if self.book.ready:
            self.book.formulaEntry.delete(0, 'end')
            self.book.formulaEntry.insert(0, self.sheet.currentCell.formula)
        
    def selectbyDragging(self, event=None, col_range=None, row_range=None, col_start_end=None, widgetEnd=None, debug=None):
        x, y = event.x_root, event.y_root
        
        widgetStart = self
        if not widgetEnd:
            widgetEnd = self.winfo_containing(x, y)
        # print(f'widgetEnd = {widgetEnd}')
        
        if orExec(isinstance(widgetEnd, CellLabel), isinstance(widgetEnd, IndexCell)):
            if debug: print('Executed: IndexCell.selectbyDragging')
        
            if widgetStart == widgetEnd:
                if self.sheet.control_is_pressed:
                    widgetStart.toggleRowCells(event=event, select=True, selectMultipleInOne=False)
                else:
                    widgetStart.focusOnRowCells(event=event)
                
            elif widgetEnd != self.widgetEnd_prev:
                self.dragRight = widgetEnd.index[0] > widgetStart.index[0]
                self.dragDown = widgetEnd.index[1] > widgetStart.index[1]
                
                self.widgetEnd = self.sheet.lastCell = widgetEnd
                
                # if not col_range:
                startCol, endCol, endColMod = 0, self.sheet.nCols, self.sheet.nCols-1
                col_range = ColRange(startCol, endColMod)
                if debug:
                    print(f'startCol, endCol, endColMod = {startCol}, {endCol}, {endColMod}')
                    print('col_range =', col_range)
                    print()
                
                # if not row_range:
                startRow, endRow = widgetStart.index_tuple[1], widgetEnd.index_tuple[1]
                if endRow < startRow:
                    startRow, endRowMod = endRow, startRow
                else:
                    startRow, endRowMod = startRow, endRow
                row_range = RowRange(startRow, endRowMod)
                if debug:
                    print(f'startRow, endRow, endRowMod = {startRow}, {endRow}, {endRowMod}')
                    print('row_range =', row_range)
                    print()
                
                current_range = CellRange(col_range, row_range, self.book, self.sheet, originCell=widgetStart.moveRight())
                if debug: print('current_range =', current_range)
                
                widgetEnd_prev = self.widgetEnd_prev
                if debug:
                    print('dir(self) =', dir(self))
                    print('widgetStart =', widgetStart)
                    print('widgetEnd_prev =', widgetEnd_prev)
                    print('widgetEnd =', widgetEnd)
                    print('widgetEnd < widgetEnd_prev =', widgetEnd < widgetEnd_prev)
                    print('widgetStart.index_tuple =', widgetStart.index_tuple)
                    print('widgetEnd.index_tuple =', widgetEnd.index_tuple)
                
                if len(self.sheet.selectedCellsSet) >= 1:
                    last_range = self.sheet.selectedCellsSet[-1]
                    last_range_cond = not last_range.within(current_range)
                else:
                    last_range_cond = False
                    
                if andExec(last_range_cond, self.sheet.control_is_pressed, self.sheet.dragFirstTouch):
                    self.sheet.selectedCellsSet.add ( current_range )
                    self.sheet.dragFirstTouch = False
                else:
                    self.sheet.selectedCellsSet[-1] = current_range
                    
                if debug:
                    print('self.sheet.selectedCellsSet =', self.sheet.selectedCellsSet)
                    print()
                
                # Each has different firstCellSelect
                firstCellSelect = self.sheet[0, widgetStart.index[1]]
                current_range.toggleCellColor(eventInherited=event, select=True, previous_range=self.sheet.previous_range)
                self.sheet.currentCell.toggleCellColor(eventInherited=event, select=False)
                
                col_diff_w_start, row_diff_w_start = widgetEnd - widgetStart
                col_diff_w_prev, row_diff_w_prev = widgetEnd - widgetEnd_prev
                
                cond_col_opposite_dir = orExec( andExec(col_diff_w_start < 0, col_diff_w_prev > 0), (col_diff_w_start > 0, col_diff_w_prev < 0) )
                cond_row_opposite_dir = orExec( andExec(row_diff_w_start < 0, row_diff_w_prev > 0), (row_diff_w_start > 0, row_diff_w_prev < 0) )
                
                if debug:
                    print('cond_col_opposite_dir =', cond_col_opposite_dir)
                    print('cond_row_opposite_dir =', cond_row_opposite_dir)
                    print()
                
                prevSelectedCells = self.sheet.previous_range
                currentSelectedCells = current_range
                
                if debug:
                    print('prevSelectedCells =', prevSelectedCells)
                    print('currentSelectedCells =', currentSelectedCells)
                
                unSelectedCellsRanges = prevSelectedCells - currentSelectedCells
                if debug: print('unSelectedCellsRanges =', unSelectedCellsRanges)
                
                unSelectedCellsRanges.toggleCellColor(eventInherited=event, select=False, firstCellSelect=firstCellSelect)
                
                if debug:
                    print()
                    print()
                
                self.sheet.previous_range = current_range
                
            self.widgetEnd_prev = widgetEnd
                
        if event:
            self.event_state, self.event_num, self.event_widget = event.state, event.num, event.widget
            
        self.sheet.switchBtnsCombo()
        
    def destroy(self, event=None):
        super().destroy()
        self.sheet.indexColumn.cells.remove(self)
        
    # def cell_bind(self):
        # super().cell_bind()
        # self.bind('<Control-1>', self.toggleRowCells)
        # self.bind('<Control-3>', self.toggleRowCells)
        # self.bind('<Button-1>', self.focusOnRowCells)
        
    # def toggleRowCells(self, event=None):
        # for cell in range(len(sheet)):
            # cell.toggleRowCells(event=event)
        
    # def focusOnRowCells(self, event=None):
        # for Row in self.sheet:
            # Row.removeSelectionMark()
            
        # for cell in self.Row:
            # cell.toggleRowCells(event=event)

class SelectAll(IndexCell):
    def __repr__(self):
        return 'Select All'
        
    def cell_bind(self, debug=None):
        Cell.cell_bind(self, debug=debug)
        for cell in self.cell_parts:
            cell.bind('<Button-1>', self.focusOnAllCells)
        
    def focusOnAllCells(self, event=None):
        self.sheet.removeSelectionMark(event=event)
        col_range = ColRange(0, self.sheet.nCols-1)
        row_range = RowRange(0, self.sheet.nRows-1)
        current_range = CellRange(col_range, row_range, self.book, self.sheet)
        self.sheet.selectedCellsSet.start( current_range )
        current_range.toggleCellColor(eventInherited=event, select=True)
        self.sheet.previous_range = current_range
        self.sheet.currentCell = current_range.getStartingCell()
        self.sheet.lastCell    = current_range.getLastCell()
        self.sheet.currentCell.toggleCellColor(eventInherited=event, select=False)
    
    
class ColumnTitle(CellLabel):
    def cell_bind(self, debug=None):
        CellLabel.cell_bind(self, debug=debug)
        for cell in self.cell_parts:
            if self.column.is_placeholder:
                cell.unbind('<Control-1>')
                cell.unbind('<Control-3>')
                cell.unbind('<Double-1>')
                cell.unbind('<Button-1>')
            else:
                cell.unbind('<Control-1>')
                cell.unbind('<Control-3>')
                cell.unbind('<Button-1>')
                cell.bind('<Control-1>', self.toggleColumnCells)
                cell.bind('<Control-3>', self.toggleColumnCells)
                cell.bind('<Button-1>', self.focusOnColumnCells)
            cell.unbind('<B1-Motion>')
            cell.bind('<B1-Motion>', self.selectbyDragging)
        
    def toggleColumnCells(self, event=None, auto=False):
        if self.book.colTitles_select_mode:
            super().toggleCell(event=event)
            self.sheet[self.colIndex].rowSelector.install()
            return
            
        column = self.column
        cell = column[0]
        cell.toggleCell(event=event)
        self.selected = not self.selected
        for cell in column[1:]:
            cell.toggleCell(event=event, select=self.selected, selectMultipleInOne=True, auto=True)
            
        if not auto:
            self.sheet.switchBtnsCombo()
        
    def focusOnColumnCells(self, event=None):
        if self.book.colTitles_select_mode:
            super().focusOnCell(event=event)
            self.sheet[self.colIndex].rowSelector.install()
            return
            
        # if self.column.is_placeholder:
            # lastColCellIndex = (self.sheet.nCols, self.index[1])
            # self.sheet.insertColRightMost(nColsSelected=1)
            # return
        
        # Otherwise:
        # print(f'event.widget = {event.widget}')
        
        self.sheet.removeSelectionMark(event=event)
        col_range = ColRange(self.column.index, self.column.index)
        row_range = RowRange(0, self.sheet.nRows-1)
        current_range = CellRange(col_range, row_range, self.book, self.sheet)
        self.sheet.selectedCellsSet.start( current_range )
        current_range.toggleCellColor(eventInherited=event, select=True)
        self.sheet.previous_range = current_range
        self.sheet.currentCell = current_range.getStartingCell()
        self.sheet.lastCell    = current_range.getLastCell()
        self.sheet.currentCell.toggleCellColor(eventInherited=event, select=False)
        
        if self.book.ready:
            self.book.formulaEntry.delete(0, 'end')
            self.book.formulaEntry.insert(0, self.sheet.currentCell.formula)
        
        # determiner = self.sheet.columnFormulaEdit
        # if isinstance(determiner, numbers.Number):
            # expr = self.book.columnFormulaInput.get('0.0', 'end')
            # for oper_sym in '+-*/=':
                # expr = expr.replace(oper_sym, f';{oper_sym};')
            # expr1 = expr.split(';')
            # expr1[-1] = self.column.colCodeWIndex
            # expr2 = ''.join(expr1)
            # self.book.columnFormulaInput.delete('0.0', 'end')
            # self.book.columnFormulaInput.insert('end', expr2)
            # self.book.columnFormulaWidget.deiconify()
        
    def selectbyDragging(self, event=None, col_range=None, row_range=None, col_start_end=None, widgetEnd=None, debug=None):
        if self.book.colTitles_select_mode:
            super().selectbyDragging(event=event)
            for cell_range in self.sheet.selectedCellsSet:
                if cell_range.row_range.start == -1:
                    for column, colTitleCell in zip(self.sheet[cell_range.col_range], self.sheet.colTitleCells[cell_range.col_range]):
                        column.rowSelector.install()
                        colTitleCell.toggleCellColor(eventInherited=event, select=True, selectMultipleInOne=True)
                    cell_range.originCell.toggleCellColor(eventInherited=event, select=False, selectMultipleInOne=True)
            return
            
        x, y = event.x_root, event.y_root
        
        widgetStart = self
        if not widgetEnd:
            widgetEnd = self.winfo_containing(x, y)
        # print(f'widgetEnd = {widgetEnd}')
        
        if isinstance(widgetEnd, CellLabel):
            if debug: print('Executed: ColumnTitle.selectbyDragging')
        
            if widgetStart == widgetEnd:
                if self.sheet.control_is_pressed:
                    widgetStart.toggleColumnCells(event=event, select=True, selectMultipleInOne=False)
                else:
                    widgetStart.focusOnColumnCells(event=event)
                
            elif widgetEnd != self.widgetEnd_prev:
                self.dragRight = widgetEnd.index[0] > widgetStart.index[0]
                self.dragDown = widgetEnd.index[1] > widgetStart.index[1]
                
                self.widgetEnd = self.sheet.lastCell = widgetEnd
                
                # if not col_range:
                startCol, endCol = widgetStart.index_tuple[0], widgetEnd.index_tuple[0]
                if endCol < startCol:
                    startCol, endColMod = endCol, startCol
                else:
                    startCol, endColMod = startCol, endCol
                col_range = ColRange(startCol, endColMod)
                if debug:
                    print(f'startCol, endCol, endColMod = {startCol}, {endCol}, {endColMod}')
                    print('col_range =', col_range)
                    print()
                
                # if not row_range:
                startRow, endRow, endRowMod = 0, self.sheet.nRows, self.sheet.nRows-1
                row_range = RowRange(startRow, endRowMod)
                if debug:
                    print(f'startRow, endRow, endRowMod = {startRow}, {endRow}, {endRowMod}')
                    print('row_range =', row_range)
                    print()
                
                current_range = CellRange(col_range, row_range, self.book, self.sheet, originCell=widgetStart.moveDown())
                if debug: print('current_range =', current_range)
                
                widgetEnd_prev = self.widgetEnd_prev
                if debug:
                    print('dir(self) =', dir(self))
                    print('widgetStart =', widgetStart)
                    print('widgetEnd_prev =', widgetEnd_prev)
                    print('widgetEnd =', widgetEnd)
                    print('widgetEnd < widgetEnd_prev =', widgetEnd < widgetEnd_prev)
                    print('widgetStart.index_tuple =', widgetStart.index_tuple)
                    print('widgetEnd.index_tuple =', widgetEnd.index_tuple)
                
                if len(self.sheet.selectedCellsSet) >= 1:
                    last_range = self.sheet.selectedCellsSet[-1]
                    last_range_cond = not last_range.within(current_range)
                else:
                    last_range_cond = False
                    
                if andExec(last_range_cond, self.sheet.control_is_pressed, self.sheet.dragFirstTouch):
                    self.sheet.selectedCellsSet.add ( current_range )
                    self.sheet.dragFirstTouch = False
                else:
                    self.sheet.selectedCellsSet[-1] = current_range
                    
                if debug:
                    print('self.sheet.selectedCellsSet =', self.sheet.selectedCellsSet)
                    print()
                
                # Each has different firstCellSelect
                firstCellSelect = self.sheet[widgetStart.index[0], 0]
                current_range.toggleCellColor(eventInherited=event, select=True, previous_range=self.sheet.previous_range)
                self.sheet.currentCell.toggleCellColor(eventInherited=event, select=False)
                
                col_diff_w_start, row_diff_w_start = widgetEnd - widgetStart
                col_diff_w_prev, row_diff_w_prev = widgetEnd - widgetEnd_prev
                
                cond_col_opposite_dir = orExec( andExec(col_diff_w_start < 0, col_diff_w_prev > 0), (col_diff_w_start > 0, col_diff_w_prev < 0) )
                cond_row_opposite_dir = orExec( andExec(row_diff_w_start < 0, row_diff_w_prev > 0), (row_diff_w_start > 0, row_diff_w_prev < 0) )
                
                if debug:
                    print('cond_col_opposite_dir =', cond_col_opposite_dir)
                    print('cond_row_opposite_dir =', cond_row_opposite_dir)
                    print()
                
                prevSelectedCells = self.sheet.previous_range
                currentSelectedCells = current_range
                
                if debug:
                    print('prevSelectedCells =', prevSelectedCells)
                    print('currentSelectedCells =', currentSelectedCells)
                
                unSelectedCellsRanges = prevSelectedCells - currentSelectedCells
                if debug: print('unSelectedCellsRanges =', unSelectedCellsRanges)
                
                unSelectedCellsRanges.toggleCellColor(eventInherited=event, select=False, firstCellSelect=firstCellSelect)
                
                if debug:
                    print()
                    print()
                
                self.sheet.previous_range = current_range
                
            self.widgetEnd_prev = widgetEnd
                
        if event:
            self.event_state, self.event_num, self.event_widget = event.state, event.num, event.widget
            
        self.sheet.switchBtnsCombo()
        
class Column:
    init = True
    ready = False
    _sumStatus = False
    _sumNum = _previous_cont = _current_cont = None
    
    def __init__(self, parent, book, colIndex, nRows, rowDimRawList, font, sheet, fonts=None, RowDist=25, cellType=CellLabel, colTitle='', QStatsType=None, is_placeholder=False, inserted=False, width=None, height=None):
        self.book = book
        self.sheet = sheet
        self.column = self
        
        self.is_placeholder = is_placeholder
        self.inserted = inserted
        
        self.indicator = 'columns'
        self._colIndex = self._index = colIndex
        self.index_tuple = (colIndex, -1)
        
        if not is_placeholder:
            # print(3442, self, inserted)
            if self.book.extension == '.qutable' and not inserted:
                if orExec(cellType == ColumnTitle, cellType == CellLabel):
                    try:
                        self.structData = sheet.structData['columns'][colIndex]
                    except KeyError:
                        self.structData = sheet.structData['columns'][str(colIndex)]
                    self.sumStatus = self.structData['sumStatus'] if 'sumStatus' in self.structData else False
                    self.sumNum = self.structData['sumNum'] if 'sumNum' in self.structData else None
                    self.previous_cont = self.structData['previous_cont'] if 'previous_cont' in self.structData else None
                    self.current_cont = self.structData['current_cont'] if 'current_cont' in self.structData else None
                    self.is_automated = self.structData['is_automated'] if 'is_automated' in self.structData else None
                    # print('Line 3445:', cellType)
                    # print('Line 3446:', self.structData)
                else:
                    self.structData = {}
                structData_N_A = False
            else:
                structData_N_A = True
                self.structData = {}
                if orExec(cellType == ColumnTitle, cellType == CellLabel):
                    sheet.structData['columns'][self.index] = self.structData
                self.sumStatus = False
                self.sumNum = None
                self.previous_cont = '0'
                self.current_cont = '0'
                self.is_automated = False
                
        self.width  = width     if width     else 0
        self.height = height    if height    else RowDist*(nRows+5)
        
        self.cells = CellSet(self.book, self.sheet)
        self.cellValues = CellSet(self.book, self.sheet, ['']*nRows)
        
        self.index_tuple = (colIndex, -2)
        self.rowSelector = RowSeparator(self, -1, book, sheet, thickness=2, length=self.width, bg='#FF0000', sepType='Selector')
        self.rowBorder = RowSeparator(self, -1, book, sheet, thickness=2, length=self.width, sepType='Border')
        self.rowShadowBorder = RowSeparator(self, -1, book, sheet, thickness=2, length=self.width, bg='#E1E3E1', sepType='ShadowBorder')
        
        self.rowShadowBorder.install()
        
        self.rowTopSeparators = CellSet(self.book, self.sheet, [self.rowSelector, self.rowBorder, self.rowShadowBorder])
        
        self.colSelectors = CellSet(self.book, self.sheet)
        self.rowSelectors = CellSet(self.book, self.sheet, [self.rowSelector])
        self.colShadowBorders = CellSet(self.book, self.sheet)
        self.rowShadowBorders = CellSet(self.book, self.sheet, [self.rowShadowBorder])
        self.colBorders = CellSet(self.book, self.sheet)
        self.rowBorders = CellSet(self.book, self.sheet, [self.rowBorder])
        
        self.top_border = False
        self.bottom_border = False
        
        self.cellType = cellType
        self.font = font
        # self.fonts = fonts = fonts if type(fonts) == list else []
        
        self.colTitleCell = None
        self.colTitle = self.title = colTitle
        self.colCodeWIndex = f'C{colIndex+1}'
        
        # if not is_placeholder and structData_N_A:
            # self.structData = {'title': colTitle, 'sumStatus': self.sumStatus, 'sumNum': self.sumNum, 'previous_cont': self.previous_cont, 'current_cont': self.current_cont, 'width': self.width, 'height': self.height}
            # sheet.structData['columns'].append(self.structData)
        
        self.RowDist = RowDist
        self.RowColTitleDist = RowDist+3
        
        # self.sheet.nRows = nRows = nRows if QStatsType is None else len(QStatsType)
        
        xloc = self.sheet.columnsTotalWidth + colIndex/4 if cellType == CellLabel else colIndex/4
        
        self.rowDimRawList = rowDimRawList
        
        pixels_per_point = 1
        row_height_top_raw = rowDimRawList[0].height if rowDimRawList else 0
        row_height_top = row_height_top_raw*pixels_per_point if andExec(self.sheet.have_headers, rowDimRawList, row_height_top_raw) else 0
        row_height_top = int(row_height_top)
        self.sheet.rowDimList.append(row_height_top)
        
        # Column Titles/Headers
        if cellType == IndexCell:
            ylocCol = self.RowColTitleDist
            
            # widget_title = '•••'
            widget_title = '...'
            
            self.colTitleCell = self.sheet.selectAll = SelectAll(self, book=book, index=(colIndex, -1), text=widget_title, font=font, sheet=self.sheet, column=self, col_width=int(self.width), row_height=row_height_top)
            
            # self.indexCellSeparator = tk.Frame(self, height=3, bg='#F5F5F5')
            # self.indexCellSeparator.place(x=0, y=ylocCol+RowDist, anchor="nw", relwidth=1.0)
            
        elif cellType == CellLabel:
            ylocCol = 0
            # if andExec(fonts, self.sheet.have_headers):
                # font = fonts[0]
            colTitleCellFont = font+('bold',) if andExec(not self.book.extension.startswith('.xls'), self.book.extension != '.qutable') else font
            self.colTitleCell = ColumnTitle(self, book=book, index=(colIndex, -1), font=colTitleCellFont, sheet=self.sheet, column=self, colTitle=colTitle, col_width=int(self.width), row_height=row_height_top)
            self.replaceTitle(colTitle)
            # self.colTitleCell.colBorder.install()
            
        self.colTitleCell.cellFrame.grid(column=2*colIndex+2, row=1, sticky='nsew')
        self.colTitleCell.pack(side='top', fill='both', expand=True)
        

        # if not is_placeholder:
        col_grid = 2*colIndex+2
        self.colTitleCell.colShadowBorder.install()
        self.colTitleCell.rowShadowBorder.install()
        
        # for colIndex, colTitle in zip(range(self.nCols), self.colTitles):
        for rowIndex in range(nRows):
            # Conversion factor using the pixel-to-point conversion rate
            try:
                row_height_points = rowDimRawList[rowIndex-1 if self.sheet.have_headers else rowIndex].height
                row_height_pixels = row_height_points * pixels_per_point if row_height_points else 0
                row_height_pixels = int(row_height_pixels)
            except (IndexError, UnboundLocalError):
                row_height_pixels = 0
                
            self.sheet.rowDimList.append(row_height_pixels)
            
            # if len(fonts) >= rowIndex+1:
                # if self.sheet.have_headers:
                    # font = fonts[rowIndex+1]
                # else:
                    # font = fonts[rowIndex]
                    
            cell = cellType(self, book=book, index=(colIndex, rowIndex), font=font, cursor='arrow', column=self, sheet=self.sheet, col_width=int(self.width), row_height=row_height_pixels)
                
            self.ylocCell = self.RowColTitleDist + (RowDist+2)*rowIndex
            
            # if colIndex == 0:
            row_grid = 2*rowIndex+3
            cell.cellFrame.grid(column=col_grid, row=row_grid, sticky='nsew')
            cell.pack(side='top', fill='both', expand=True)
            
            colBorder, rowBorder, colShadowBorder, rowShadowBorder, colSelector, rowSelector = self.addBordersSelectors(cell.colBorder, cell.rowBorder, cell.colShadowBorder, cell.rowShadowBorder, cell.colSelector, cell.rowSelector)
            
            colShadowBorder.install()
            rowShadowBorder.install()
            
            if cellType == IndexCell:
                # Determine QStatsType
                if QStatsType is None:
                    cell.replaceFormula(rowIndex+1)
                else:
                    cell.formula = QStatsType[rowIndex]
                # rowBorder.install()
            
            self.cells.add(cell, init=True, cell_raw=True)
            # self.sheet.cells.add(cell, init=True, cell_raw=True)
        
        # Placeholder Cell
        rowIndex += 1
        row_grid = 2*rowIndex+3
        self.ylocCell = self.RowColTitleDist + (RowDist+2)*rowIndex
        self.placeholderCell = cell = cellType(self, book=book, index=(colIndex, rowIndex), font=font, cursor='arrow', column=self, sheet=self.sheet, is_placeholder=True, col_width=int(self.width), row_height=row_height_pixels)
        if cellType == IndexCell:
            cell.content = '+'
        cell.cellFrame.grid(column=col_grid, row=row_grid, sticky='nsew')
        cell.pack(side='top', fill='both', expand=True)
        
        colBorder, rowBorder, colShadowBorder, rowShadowBorder, colSelector, rowSelector = self.addBordersSelectors(cell.colBorder, cell.rowBorder, cell.colShadowBorder, cell.rowShadowBorder, cell.colSelector, cell.rowSelector)
        
        colShadowBorder.install()
        rowShadowBorder.install()
            
        # Overall Column Font Insertion
        if hasattr(font, '__iter__'): colTitleCellFont = font+('bold',)
        elif type(font) == str: colTitleCellFont = font+' bold'
        
        if cellType == IndexCell:
            self.fonts = fonts if fonts else [colTitleCellFont]*(self.sheet.nRows+1)
            return
        
        self.fonts = fonts if fonts else [colTitleCellFont]+[font]*(self.sheet.nRows)
        
        self.ready = True
        self.init = False
        
    def autoFormat(self, structData):
        self.sumStatus, self.sumNum, self.previous_cont, self.current_cont, self.is_automated, self.height, self.title, self.width = structData['sumStatus'], structData['sumNum'], structData['previous_cont'], structData['current_cont'], structData['is_automated'], structData['height'], structData['title'], structData['width']
        self.colTitleCell['wraplength'] = self.structData['width']
        for cell in self:
            cell['wraplength'] = self.structData['width']
        self.structData = self.sheet.structData[self.indicator][str(self.index[0]) if isinstance(self, ColumnTitle) else str(self.index)] = structData
        
    @property
    def cellValues(self):
        return self._cellValues
    @cellValues.setter
    def cellValues(self, new_cellValues):
        self._cellValues = new_cellValues
        for cell, cellValue in zip(self, new_cellValues):
            cell.formula = cellValue
        
    @property
    def colIndex(self):
        return self._colIndex
    @colIndex.setter
    def colIndex(self, new_colIndex):
        self._colIndex = self._index = new_colIndex
        self.colCodeWIndex = f'C{self.colIndex+1}'
        for cell_separator in self.rowTopSeparators:
            cell_separator.colIndex = new_colIndex
        if self.colTitleCell:
            self.colTitleCell.colIndex = new_colIndex
        for cell in self:
            cell.colIndex = new_colIndex
        self.placeholderCell.colIndex = new_colIndex
        
    @property
    def index(self):
        return self._index
    @index.setter
    def index(self, new_colIndex):
        self._index = self._colIndex = new_colIndex
        self.colCodeWIndex = f'C{self.colIndex+1}'
        for cell_separator in self.rowTopSeparators:
            cell_separator.colIndex = new_colIndex
        if self.colTitleCell:
            self.colTitleCell.colIndex = new_colIndex
        for cell in self:
            cell.colIndex = new_colIndex
        
    @property
    def colTitle(self):
        try:
            return self._colTitle
        except AttributeError:
            return f'Column #{self.index}'
    @colTitle.setter
    def colTitle(self, new_title):
        self._title = self._colTitle = new_title
        if self.colTitleCell:
            try:
                self.colTitleCell._formula = self.colTitleCell._content = self.colTitleCell.structData['formula'] = self.colTitleCell.structData['content'] = self.colTitleCell['text'] = new_title
            except AttributeError:
                self.colTitleCell._formula = self.colTitleCell._content = self.colTitleCell['text'] = new_title
        if not self.is_placeholder:
            self.structData['title'] = new_title
        
    @property
    def title(self):
        try:
            return self._title
        except AttributeError:
            return f'Column #{self.index}'
    @title.setter
    def title(self, new_title):
        self._title = self._colTitle = new_title
        if self.colTitleCell:
            try:
                self.colTitleCell._formula = self.colTitleCell._content = self.colTitleCell.structData['formula'] = self.colTitleCell.structData['content'] = self.colTitleCell['text'] = new_title
            except AttributeError:
                self.colTitleCell._formula = self.colTitleCell._content = self.colTitleCell['text'] = new_title
        if not self.is_placeholder:
            self.structData['title'] = new_title
        
    @property
    def sumNum(self):
        return self._sumNum
    @sumNum.setter
    def sumNum(self, new_sumNum):
        if not self.is_placeholder:
            self._sumNum = self.structData['sumNum'] = new_sumNum
        
    @property
    def is_automated(self):
        return self._is_automated
    @is_automated.setter
    def is_automated(self, new_is_automated):
        if not self.is_placeholder:
            self._is_automated = self.structData['is_automated'] = new_is_automated
        
    @property
    def sumStatus(self):
        return self._sumStatus
    @sumStatus.setter
    def sumStatus(self, new_sumStatus):
        if not self.is_placeholder:
            self._sumStatus = self.structData['sumStatus'] = new_sumStatus
        
    @property
    def width(self):
        return self._width
    @width.setter
    def width(self, new_width):
        self._width = new_width
        if self.ready:
            for cell in self:
                cell.width = new_width
        if not self.is_placeholder:
            self.structData['width'] = new_width
        
    @property
    def height(self):
        return self._height
    @height.setter
    def height(self, new_height):
        self._height = new_height
        if not self.is_placeholder:
            self.structData['height'] = new_height
    
    # For Stats purposes
    @property
    def previous_cont(self):
        return self._previous_cont
    @previous_cont.setter
    def previous_cont(self, new_previous_cont):
        if not self.is_placeholder:
            self._previous_cont = self.structData['previous_cont'] = new_previous_cont
        
    # For Stats purposes
    @property
    def current_cont(self):
        return self._current_cont
    @current_cont.setter
    def current_cont(self, new_current_cont):
        if not self.is_placeholder:
            self._current_cont = self.structData['current_cont'] = new_current_cont
        
    @property
    def fonts(self):
        return [self.colTitleCell.font]+[cell.font for cell in self]
    @fonts.setter
    def fonts(self, new_fonts):
        if andExec(new_fonts, self.sheet.have_headers):
            self.colTitleCell.font = new_fonts[0]
        else:
            self.colTitleCell.font = 'Arial 12 bold'
        for cell, new_font in zip(self, new_fonts[self.sheet.have_headers:]):
            cell.font = new_font
        
    def addBordersSelectors(self, colBorder, rowBorder, colShadowBorder, rowShadowBorder, colSelector, rowSelector):
        self.colBorders += [colBorder]
        self.rowBorders += [rowBorder]
        self.colShadowBorders += [colShadowBorder]
        self.rowShadowBorders += [rowShadowBorder]
        self.colSelectors += [colSelector]
        self.rowSelectors += [rowSelector]
        
        self.sheet.colBorders += [colBorder]
        self.sheet.rowBorders += [rowBorder]
        self.sheet.colShadowBorders += [colShadowBorder]
        self.sheet.rowShadowBorders += [rowShadowBorder]
        self.sheet.colSelectors += [colSelector]
        self.sheet.rowSelectors += [rowSelector]
        
        return colBorder, rowBorder, colShadowBorder, rowShadowBorder, colSelector, rowSelector
        
    def column_bind(self):
        # Resize mode
        self.NONE = 0

        self.sizeAdjustStatus = self.NONE
        self.isDragBandHeld = True

        self.cursor = ''
        self.dragBandWidth = 10
        
        borderThickness = 5
        
        self.bind("<ButtonPress-1>", self.SizeAdjustInitialize)
        self.bind("<ButtonRelease-1>", self.SizeAdjustHalt)
        self.bind("<B1-Motion>", self.CursorPosition)
        self.bind('<MouseWheel>', self.sheet.mouseWheel )
        
    def __getitem__(self, index):
        if type(index) == int:
            if index == -1:
                return self.colTitleCell
            elif index == self.sheet.nRows:
                return self.placeholderCell
            else:
                return self.cells[index]
            
        elif isinstance(index, LineRange):
            colTitleCellSet = CellSet(self.book, self.sheet, [self.colTitleCell] if index.start == -1 else [] )
            index_start = index.start
            if andExec(index.start < 0, index.stop > index.start):
                index.start = 0
            ordinaryCellSet = CellSet(self.book, self.sheet, self.cells[index] )
            placeholderCellSet = CellSet(self.book, self.sheet, [self.placeholderCell] if index.stop == self.sheet.nRows else [] )
            index.start = index_start
            selectedCellsSet = colTitleCellSet + ordinaryCellSet + placeholderCellSet
            return CellSet(self.book, self.sheet, selectedCellsSet)
            
        elif type(index) == range:
            slice_range = slice(index.start, index.stop, index.step)
            return self.cells[slice_range]
            
        elif type(index) == str:
            return super().__getitem__(index)
            
        elif hasattr(index, '__iter__'):
            if len(index) == 1:
                return self.cells[index[0]]
            elif len(index) == 2:
                return self.cells[index[0]][index[1]]
            
        elif type(index) == slice:
            return self.cells[index]
            
        else:
            return super().__getitem__(index)
        
    def __setitem__(self, index, value):
        if type(index) == int:
            self.cells[index] = value
            
        elif type(index) == range:
            slice_range = slice(index.start, index.stop, index.step)
            self.cells[slice_range] = value
            
        elif type(index) == str:
            super().__setitem__(index, value)
            
        elif hasattr(index, '__iter__'):
            if len(index) == 1:
                self.cells[index[0]] = value
            elif len(index) == 2:
                self.cells[index[0]][index[1]] = value
            
        elif type(index) == slice:
            self.cells[index] = value
            
        else:
            super().__setitem__(index, value)
        
    def __iter__(self):
        for cell in self.cells:
            yield cell
        
    def __len__(self):
        return len(self.cells)

    def __lt__(self, other):
        return self.cells < other.cells

    def __gt__(self, other):
        return self.cells > other.cells

    def __le__(self, other):
        return self.cells <= other.cells

    def __ge__(self, other):
        return self.cells >= other.cells

    def __repr__(self):
        try:
            if self.colTitle:
                colTitle = self.colTitle.replace("\n", " ").replace("\r", " ").replace("\t", " ")
                return f'{colTitle} = {self.colCodeWIndex}'
            elif self.colCodeWIndex:
                return self.colCodeWIndex
            else:
                colTitle = self.colTitle.replace("\n", " ").replace("\r", " ").replace("\t", " ")
                return colTitle
        except AttributeError:
            return f'Column #{self.index}'
            
    def __str__(self):
        return self.__repr__()
            
    def __add__(self, other):
        multipliedColContents = []
        if type(other) == type(self):
            for cell_in_self, cell_in_other in zip(self, other):
                try: cell_in_self = float(cell_in_self)
                except (ValueError, TypeError): continue
                try: cell_in_other = float(cell_in_other)
                except (ValueError, TypeError): continue
                sum_cell = cell_in_self + cell_in_other
                multipliedColContents += [sum_cell]
        elif orExec(type(other) == int, type(other) == float):
            for cell_in_self in self:
                try: cell_in_self = float(cell_in_self)
                except (ValueError, TypeError): continue
                sum_cell = cell_in_self + other
                multipliedColContents += [sum_cell]
        return multipliedColContents
            
    def __mul__(self, other):
        multipliedColContents = []
        if type(other) == type(self):
            for cell_in_self, cell_in_other in zip(self, other):
                # try: cell_in_self = cell_in_self.content
                # except (ValueError, TypeError): continue
                # try: cell_in_other = cell_in_other.content
                # except (ValueError, TypeError): continue
                product_cell = cell_in_self * cell_in_other
                multipliedColContents += [product_cell]
        elif orExec(type(other) == int, type(other) == float):
            for cell_in_self in self:
                # try: cell_in_self = cell_in_self.content
                # except (ValueError, TypeError): continue
                # try: other = other.content
                # except (ValueError, TypeError): continue
                sum_cell = cell_in_self * other
                multipliedColContents += [sum_cell]
        return multipliedColContents
            
    def __pow__(self, other):
        multipliedColContents = []
        if type(other) == type(self):
            for cell_in_self, cell_in_other in zip(self, other):
                try: cell_in_self = float(cell_in_self)
                except (ValueError, TypeError): continue
                try: cell_in_other = float(cell_in_other)
                except (ValueError, TypeError): continue
                exp_cell = cell_in_self ** cell_in_other
                multipliedColContents += [exp_cell]
        elif orExec(type(other) == int, type(other) == float):
            for cell_in_self in self:
                try: cell_in_self = float(cell_in_self)
                except (ValueError, TypeError): continue
                sum_cell = cell_in_self ** other
                multipliedColContents += [sum_cell]
        return multipliedColContents

        
    def getCellValues(self, form=None, emptyExclude=False):
        if andExec(not form, not emptyExclude):
            return self.cellValues
        
        if orExec(form == float, form == int):
            values = []
            for cell in self.cells:
                content = cell.getContent()
                try:
                    content = form(content)
                except (TypeError, ValueError):
                    content = 0
                if not emptyExclude: values += [content]
                elif andExec(emptyExclude, content): values += [content]
            
        else:
            values = [cell.getContent() for cell in self.cells]
            
        return values
            
    def is_empty(self):
        return set( self.getFormulae() ) == {''}
            
    def sumAll(self):
        result = 0
        for cell in self:
            content = cell.getContent()
            try:
                if self.book.quantum_mode:
                    result = QBO.adder(result, float(content))
                else:
                    result += float(content)
            except (ValueError, TypeError): pass
        # if str(result).endswith('.0'): result = int(result)
        self.sumNum = result
        self.sumStr = float_int_convert(result)
        self.sumStatus = True
        return self.sumStr
            
    def sum(self):
        # if self.sumStatus and isinstance(self.current_cont, numbers.Number):
        if self.sumStatus:
            # print(f'1 current_cont = {self.current_cont}')
            if not is_numeric(self.current_cont):
                self.current_cont = '0'
            if not is_numeric(self.previous_cont):
                self.previous_cont = '0'
            if self.book.quantum_mode:
                diff_cont = QBO.subtractor(self.current_cont, self.previous_cont)
                self.sumStr = QBO.adder(self.sumNum, diff_cont)
                self.sumNum = float(self.sumStr)
                # print(f'sumNum = {self.sumNum}')
                # print()
            else:
                diff_cont = float(self.current_cont) - float(self.previous_cont)
                # print(f'self.sumNum = {self.sumNum}')
                # print(f'current_cont = {self.current_cont}')
                # print(f'previous_cont = {self.previous_cont}')
                # print(f'diff_cont = {diff_cont}')
                # print(f'='*20)
                self.sumNum += diff_cont
                self.sumStr = float_int_convert(self.sumNum)
                # if self.sumNum == int(self.sumNum):
                    # self.sumNum = int(self.sumNum)
            return self.sumStr
        else:
            return self.sumAll()
        
    def count(self, is_num=False):
        try:
            count_num = 0
            for cellValue in self.cellValues:
                cellValueContentDetection = cellValue if is_numeric(cellValue.strip()) else ( 0 if is_num else cellValue.strip() )
                if cellValueContentDetection:
                    # if self.book.quantum_mode:
                        # count_num = QBO.adder(count_num, 1)
                    # else:
                    count_num += 1
            return count_num
        except (ValueError, TypeError, statistics.StatisticsError): return 0
        
    def mean(self):
        try:
            count_num = self.count(is_num=True)
            return self.sumNum/count_num if count_num else 0
            # if self.book.quantum_mode:
                # return self.sumNum/self.count()
                # # return self.sum()/self.count()
            # else:
                # return statistics.mean(self.getCellValues(form=float))
            # # self.avgNum = self.meanNum = result
        except (ValueError, TypeError, statistics.StatisticsError):
            return 0
        
    def average(self):
        return self.mean()
            
    def max(self):
        try: return max(self.getCellValues(form=float))
        except (ValueError, TypeError, statistics.StatisticsError): return 0
            
    def min(self):
        try: return min(self.getCellValues(form=float))
        except (ValueError, TypeError, statistics.StatisticsError): return 0
            
    # def stdev(self):
        # try: return statistics.stdev(self.getCellValues(form=float))
        # except (ValueError, TypeError, statistics.StatisticsError): return 0
            
    # def variance(self, n):
        # try:
            # if self.book.quantum_mode:
                # mu, self.Sigma = self.mean(), 0
                # for cell in self:
                    # x_i = cell.getContent()
                    # diff = QBO.subtractor(x_i, mu)
                    # self.Sigma += QBO.QMul(diff, diff)
                # return self.Sigma/n
            # else:
                # return statistics.variance(self.getCellValues(form=float))
        # except (ValueError, TypeError, statistics.StatisticsError): return 0
    
    def varianceQI(self, cell, mu):
        ''' Variance Q-Intermediate '''
        x_i = cell.getContent()
        diff = QBO.subtractor(x_i, mu)
        self.Sigma = QBO.adder(self.Sigma, QBO.QMul(diff, diff))
        printed_output = f'Sigma = {self.Sigma}'
        
    def variance(self, n):
        try:
            mu, self.Sigma = self.mean(), 0
            # if 0:
                # with ThreadPoolExecutor(max_workers=self.sheet.nRows) as executor:
                    # # for cell in self:
                        # # executor.submit(self.varianceQI, cell, mu).result()
                    # # Submit each text to the executor for parallel processing
                    # future_results = [executor.submit(self.varianceQI, cell, mu) for cell in self]
                    # # Wait for all tasks to complete and get the results
                    # results = [future.result() for future in future_results]
            if self.book.quantum_mode:
                with multiprocessing.Pool(processes=self.sheet.nRows) as pool:
                    results = pool.starmap(self.varianceQI, [(cell, mu) for cell in self])
                # Extract processed texts and printed output
                processed_texts, printed_outputs = zip(*processed_texts_with_output)
            else:
                for cell in self:
                    x_i = float(cell.getContent())
                    diff = x_i - mu
                    self.Sigma += diff**2
            return self.Sigma/n
        except (ValueError, TypeError): raise; return 0
        
    def svariance(self):
        """ Sample Variance """
        try: return self.variance(self.count()-1)
        except (ValueError, TypeError, statistics.StatisticsError): return 0
            
    def pvariance(self):
        """ Population Variance """
        try: return self.variance(self.count())
        except (ValueError, TypeError, statistics.StatisticsError): return 0
        
    def s_stdev(self):
        """ Sample Standard Deviation """
        try: return self.svariance()**0.5
        except (ValueError, TypeError, statistics.StatisticsError): return 0
            
    def p_stdev(self):
        """ Population Standard Deviation """
        try: return self.pvariance()**0.5
        except (ValueError, TypeError, statistics.StatisticsError): return 0
        
    def getStatsReport(self, statsReportVisible=True):
        StatsAttrs = ['Sum', 'Count', 'Average', 'Maximum', 'Minimum', 'Sample Standard Deviation', 'Sample Variance', 'Population Standard Deviation', 'Population Variance']
        StatsValues = [self.sum(), self.count(), self.mean(), self.max(), self.min(), self.s_stdev(), self.svariance(), self.p_stdev(), self.pvariance()]
        
        StatsList = [] #if statsReportVisible else {}
        for StatsAttr, StatsValue in zip(StatsAttrs, StatsValues):
            try:
                if 1: #statsReportVisible:
                    StatsList += [StatsValue]
                else:
                    StatsList[StatsAttr] = StatsValue
            except (ValueError, TypeError):
                if 1: #statsReportVisible:
                    StatsList += ['N/A']
                else:
                    StatsList[StatsAttr] = 'N/A'
        
        StatsDict = {self.colCodeWIndex if statsReportVisible else self.index: StatsList}
        
        self.current_cont = '0'
        self.previous_cont = '0'
        
        return StatsDict
        
    def copyCellContents(self, cell_from, cell_to):
        if type(cell_from) == int:
            cell_from = self[cell_from]
        if type(cell_to) == int:
            cell_to = self[cell_to]
        
        cell_to.formula = cell_from.getContent()

    def generateRowCellsIndex(self, readable=False):
        return [cell.index[1]+1 if readable else cell.index[1] for cell in self]
        
    def insertRowBottomMost(self, nRowsSelected=None):
        colIndex = self.colIndex
        
        if not isinstance(nRowsSelected, numbers.Number):
            nRowsSelected = self.sheet.selectedCellsSet.getRowSize()
        
        last_cell = self[self.sheet.nRows-1]
        # if not self.is_placeholder:
        
        for i in range(nRowsSelected):
            rowIndex = self.sheet.nRows+i
            if self.cellType == IndexCell:
                if type(self.font) == str: self.font += ' bold'
                elif hasattr(self.font, '__iter__'): self.font += ('bold',)
            
            row_height_pixels = self.sheet.rowDimList[-1]
            # print('Line 3699:', self.cellType, rowIndex, rowIndex+i)
            newRow = self.cellType(self, book=self.book, inserted=True, index=(self.colIndex, rowIndex), font=last_cell.font, cursor='arrow', column=self, sheet=self.sheet, col_width=int(self.width), row_height=row_height_pixels, anchor=last_cell.anchorVar, justify=last_cell.justify)
            # print(4179, self.sheet.nRows, i, rowIndex, newRow)
            
            col_grid, row_grid = 2*colIndex+2, 2*rowIndex+3
            newRow.cellFrame.grid(column=col_grid, row=row_grid, sticky='nsew')
            newRow.pack(side='top', fill='both', expand=True)
            
            colBorder, rowBorder, colShadowBorder, rowShadowBorder, colSelector, rowSelector = self.addBordersSelectors(newRow.colBorder, newRow.rowBorder, newRow.colShadowBorder, newRow.rowShadowBorder, newRow.colSelector, newRow.rowSelector)
            
            colShadowBorder.install()
            rowShadowBorder.install()
            
            # self.ylocCell = self.RowColTitleDist + (self.RowDist+2)*rowIndex
            # rowBorder = tk.Frame(self, width=2, bg='#F5F5F5')
            # self.rowBorders += [rowBorder]
            # rowBorder.place(x=0, y=self.ylocCell+self.RowDist, anchor="nw", relwidth=1.0)
            
            if self.cellType == IndexCell:
                newRow.content = rowIndex+1
                
            self.cells.add(newRow, init=True, cell_raw=True)
        
        bottomMostRow = newRow
        return bottomMostRow
            
    def deleteRowCells(self, selectedRowsIndex=None):
        allRowIndexSet = set( range(self.sheet.nRows) ) if self.cellType == IndexCell else set( self.generateRowCellsIndex() )
        if not hasattr(selectedRowsIndex, '__iter__'):
            selectedRowsIndex = self.sheet.selectedRowsIndex
        selectedRowsIndex = set( selectedRowsIndex )
        
        unselectedRowsIndex = allRowIndexSet.difference( selectedRowsIndex )
        self.sheet.nRows = len(unselectedRowsIndex)
        
        unselectedRowsIndexList = list(unselectedRowsIndex) + [None]*len( selectedRowsIndex )
        allRowIndexList = list(allRowIndexSet)
        
        if self.cellType == IndexCell:
            start, end = self.sheet.nRows, self.sheet.nRows+1
            for allRowIndex in range(start, end):
                cell = self[allRowIndex]
                cell.cellFrame.grid_forget()
                cell.pack_forget()
                self.cells[allRowIndex] = ''
                
                sheet_cells = self.sheet.cells
                if cell in sheet_cells: sheet_cells.remove( cell )
        else:
            for unselectedRowIndex, allRowIndex in zip( unselectedRowsIndexList , allRowIndexList ):
                if unselectedRowIndex is None:
                    cell = self[allRowIndex]
                    cell.cellFrame.grid_forget()
                    cell.pack_forget()
                    self.cells[allRowIndex] = ''
                
                    sheet_cells = self.sheet.cells
                    if cell in sheet_cells: sheet_cells.remove( cell )
            
                else:
                    # print('unselectedRowIndex =', unselectedRowIndex)
                    rowContent = self[unselectedRowIndex].getContent()
                    self[allRowIndex].replaceFormula( rowContent )
        
            # print('unselectedRowsIndexList =', unselectedRowsIndexList)
            self[ min(selectedRowsIndex) ].focusOnCell()
        
        self.cells.remove('')
        
    def SizeAdjustInitialize(self, event):
        """Set the resize mode if the left button of the mouse is clicked close to the right or bottom edge of the frame."""
        # Extract the size of the frame
        width = self.winfo_width()
        height = self.winfo_height()

        # If the mouse left button is clicked close to the right edge allow horizontal resizing
        if event.x > width-self.dragBandWidth: 
            self.sizeAdjustStatus = tk.HORIZONTAL
        else:
            self.sizeAdjustStatus = self.NONE

    def CursorPosition(self, event):    
        """Check whether the cursor is close to the right or bottom edge of the frame."""
        # Extract the size of the frame
        width = self.winfo_width()
        height = self.winfo_height()

        # If the cursor is close to the right edge change cursor icon
        if event.x > width-self.dragBandWidth: 
            self.config(cursor='sb_h_double_arrow')
        else:
            self.config(cursor='')

        # If horizontal resizing is allowed then resize frame with cursor
        if self.sizeAdjustStatus == tk.HORIZONTAL:
            initColWidth = self.width
            self.width = event.x
            self.config(width=self.width)
            
            increase = self.width - initColWidth
            self.sheet.columnsTotalWidth += increase
            self.sheet.columnsTotalWidthActual += increase

    def SizeAdjustHalt(self, event):
        """Disable any resize mode and set the standard arrow as cursor."""
        self.sizeAdjustStatus = self.NONE
        self.config(cursor='')

    def ColToDict(self):
        """Convert columns to dicts for dataframes"""
        return {self.colTitle:[cell.getContent() for cell in self.cells]}
        
    def getFormulae(self):
        cellFormulae = []
        for cell in self:
            cellFormulae += [cell.formula]
        return cellFormulae
        
    def replaceFormulae(self, formula_list):
        for cell, formula in zip(self, formula_list):
            cell.formula = formula
        
    def replaceTitle(self, new_title):
        self.colTitle = new_title
        
    def recalibrateIndex(self, col_translation):
        self.colIndex += col_translation
        
    def moveDown(self, units=1):
        return self.colTitleCell.moveDown(units-1)


class Range:
    def __init__(self, start=None, stop=None, step=1, cond_opposite_dir=False, book=None, sheet=None):
        if not isinstance(stop, numbers.Number):
            if isinstance(start, numbers.Number):
                stop = start
                self.slicer = slice(start, stop+1, step)
                self.ranger = range(start, stop+1, step)
            else:
                self.slicer = slice(None)
                self.ranger = None
        else:
            if stop < start:
                start, stop = stop, start
            self.slicer = slice(start, stop+1, step)
            self.ranger = range(start, stop+1, step)
            
        self._start, self._stop, self._step = start, stop, step
        
        self.book = book
        self.sheet = sheet
        
        self.checkUpperLimit()
        
    @property
    def start(self):
        return self._start
    @start.setter
    def start(self, new_start):
        self._start = new_start
        self.slicer = slice(new_start, self.slicer.stop, self.slicer.step)
        self.ranger = range(new_start, self.ranger.stop, self.ranger.step)
        
    @property
    def stop(self):
        return self._stop
    @stop.setter
    def stop(self, new_stop):
        self._stop = new_stop
        self.slicer = slice(self.slicer.start, new_stop, self.slicer.step)
        self.ranger = range(self.ranger.start, new_stop, self.ranger.step)
        
    @property
    def step(self):
        return self._step
    @step.setter
    def step(self, new_step):
        self._step = new_step
        self.slicer = slice(self.slicer.start, self.slicer.stop, new_step)
        self.ranger = range(self.ranger.start, self.ranger.stop, new_step)
        
    def checkUpperLimit(self):
        pass
        
    def __repr__(self):
        return f'{self.__class__.__name__}NaturalNum{self.start+1, self.stop+1}'
        
    def repr_actual(self):
        return f'{self.__class__.__name__}{self.start, self.stop}'
        
    def __str__(self):
        return self.__repr__()
        
    def getSize(self):
        return self.stop+1 - self.start
        
    def within(self, other):
        return andExec(self.start >= other.start, self.stop <= other.stop)
        
    def __iter__(self):
        for num in self.ranger:
            yield num
        
    def same_size_with(self, other):
        return self.getSize() == other.getSize()
        
    def __lt__(self, other):
        return self.getSize() < other.getSize()
        
    def __gt__(self, other):
        return self.getSize() > other.getSize()
        
    def __le__(self, other):
        return self.getSize() <= other.getSize()
        
    def __ge__(self, other):
        return self.getSize() >= other.getSize()
        
    def __mul__(self, other):
        return self.getSize() * other.getSize()
        
    def __add__(self, other):
        if type(other) == type(self):
            start_sum, stop_sum = self.start+other.start, self.stop+other.stop
        elif isinstance(other, numbers.Number):
            start_sum, stop_sum = self.start+other, self.stop+other
        return self.__class__(start_sum, stop_sum, book=self.book, sheet=self.sheet)
        
    def __sub__(self, other):
        if type(other) == type(self):
            if self.within(other):
                self, other = other, self
            start1, end1 = self.start, other.start-1
            start2, end2 = other.stop+1, self.stop
            sub1 = self.__class__(start1, end1, book=self.book, sheet=self.sheet) if end1 >= start1 else EmptyRange()
            sub2 = self.__class__(start2, end2, book=self.book, sheet=self.sheet) if end2 >= start2 else EmptyRange()
            if type(sub1) == EmptyRange: diff = sub2
            elif type(sub2) == EmptyRange: diff = sub1
            else: diff = [sub1, sub2]
            return diff
        elif isinstance(other, numbers.Number):
            start_diff, stop_diff = self.start-other, self.stop-other
            return self.__class__(start_diff, stop_diff, book=self.book, sheet=self.sheet)
        
    def reverse(self):
        start, stop = self.stop-1, self.start-1
        step = -self.step if isinstance(self.step, numbers.Number) else self.step
        self.slicer = slice(start, stop, step)
        self.start, self.stop, self.step = start, stop, step
    
class LineRange(Range):
    def toggleCell(self, event=None, *args, **kwargs):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        if not isinstance(col_range, ColRange):
            col_range = ColRange(0, sheet.nCols)
        if not isinstance(row_range, RowRange):
            row_range = RowRange(0, sheet.nRows)
        cellSet = sheet[col_range, row_range]
        for subset in cellSet:
            for cell in subset:
                cell.toggleCell(event=event, *args, **kwargs)
        
    def shift_n_units(self, translation):
        start = self.start + translation
        stop  = self.stop  + translation
        return self.__class__(start, stop, book=self.book, sheet=self.sheet)
        
    def extend_n_units(self, stretch):
        lowerLimitCond = self.start >= 0
        if isinstance(self.sheet, Sheet):
            if type(self) == ColRange:
                upperLimitCond = self.stop <= self.sheet.nRows
            elif type(self) == RowRange:
                upperLimitCond = self.stop <= self.sheet.nCols
        if andExec(stretch < 0, lowerLimitCond):
            self.start = self.start - stretch
        elif andExec(stretch > 0, upperLimitCond):
            self.stop = self.stop + stretch
        # return self.__class__(start, stop, book=self.book, sheet=self.sheet)
        
    def getLineNum(self):
        return list(self.ranger)
        
    def is_one_line(self):
        return self.start == self.stop
        
    def __eq__(self, other):
        if type(other) == type(self):
            return andExec(self.start == other.start, self.stop == other.stop)
        else:
            return False
        
    def __ne__(self, other):
        if type(other) == type(self):
            return andExec(self.start != other.start, self.stop != other.stop)
        else:
            return False
        
    def generateRange(self, col_range=None, row_range=None, book=None, sheet=None, super_inclusive=False, debug=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        if not isinstance(book, Book):
            book = self.book
        if not isinstance(row_range, RowRange):
            if debug: print('sheet.nRows =', sheet.nRows)
            row_range = RowRange(-1, sheet.nRows) if super_inclusive else RowRange(0, sheet.nRows-1)
        if not isinstance(col_range, ColRange):
            if debug: print('sheet.nCols =', sheet.nCols)
            col_range = ColRange(0, sheet.nCols) if super_inclusive else ColRange(0, sheet.nCols-1)
        return CellRange( col_range , row_range , book=book , sheet=sheet )
        
    def replaceFormulae(self, content, col_range=None, row_range=None, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        self.generateRange(col_range=col_range, row_range=row_range, sheet=sheet).replaceFormulae(content)
        
    def map(self, function, col_range=None, row_range=None, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        self.generateRange(col_range=col_range, row_range=row_range, sheet=sheet).map(function)
        
    def concat(self, other):
        return tuple(self) + tuple(other)
        
    def isHighEnough(self):
        return self.start >= -1 and self.stop >= -1
    
class ColRange(LineRange):
    """ Create a Col slice and range objects with start and stop, inclusive. """
    def checkUpperLimit(self):
        if isinstance(self.sheet, Sheet):
            if self.stop >= self.sheet.nCols:
                self.stop = self.sheet.nCols
        
    def toggleCell(self, event=None, *args, **kwargs):
        super().toggleCell(event=event, *args, **kwargs)
        
    def colRangeToTheRight(self, inclusive=True, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        start = self.start if inclusive else self.start+self.getSize()
        stop = sheet.nCols+1
        return self.__class__(start, stop, sheet=sheet, book=self.book)
        
    def __call__(self, inclusive=True, sheet=None):
        return self.colRangeToTheRight(inclusive=inclusive, sheet=sheet)
        
    def generateColIndexList(self):
        return set(self.ranger)
        
    def generateRange(self, col_range=None, row_range=None, super_inclusive=False, *args, **kwargs):
        return super().generateRange(col_range=self, row_range=row_range, super_inclusive=super_inclusive, *args, **kwargs)
        
    def replaceFormulae(self, content, row_range=None, sheet=None):
        super().replaceFormulae(content, col_range=self, row_range=row_range, sheet=sheet)
        
    def includes(self, cell):
        col_index = cell.index[0]
        return andExec(col_index >= self.start, col_index <= self.stop)
        
    def generate(self, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        return sheet[self, :]
        
class RowRange(LineRange):
    """ Create a Row slice and range objects with start and stop, inclusive. """
    def checkUpperLimit(self):
        if isinstance(self.sheet, Sheet):
            if self.stop >= self.sheet.nRows:
                self.stop = self.sheet.nRows
            
    def toggleCell(self, event=None, *args, **kwargs):
        super().toggleCell(event=event, *args, **kwargs)
        
    def rowRangeToTheBottom(self, inclusive=True, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        start = self.start if inclusive else self.start+self.getSize()
        stop = sheet.nRows+1
        return self.__class__(start, stop, sheet=sheet, book=self.book)
        
    def __call__(self, inclusive=True, sheet=None):
        return self.rowRangeToTheBottom(inclusive=inclusive, sheet=sheet)
        
    def generateRowIndexList(self):
        return set(self.ranger)
        
    def generateRange(self, col_range=None, row_range=None, super_inclusive=False, *args, **kwargs):
        return super().generateRange(col_range=col_range, row_range=self, super_inclusive=super_inclusive, *args, **kwargs)
        
    def replaceFormulae(self, content, col_range=None, sheet=None):
        super().replaceFormulae(content, col_range=col_range, row_range=self, sheet=sheet)
        
    def includes(self, cell):
        row_index = cell.index[1]
        return andExec(row_index >= self.start, row_index <= self.stop)
        
    def generate(self, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        return sheet[:, self]

class EmptyRange(Range):
    def __init__(self, start=-1, stop=-2, step=1, cond_opposite_dir=False):
        if stop < start:
            start, stop = stop, start
        self.slicer = slice(start, stop+1, step)
        self.ranger = range(start, stop+1, step)
        self.start, self.stop, self.step = start, stop, step

class EmptyColRange(ColRange, EmptyRange):
    def ________():
        pass
    
class EmptyRowRange(RowRange, EmptyRange):
    def ________():
        pass
    
def RowRangeNaturalNum(start, stop, book=None, sheet=None):
    return RowRange(start-1, stop-1, book=book, sheet=sheet)
    
def ColRangeNaturalNum(start, stop, book=None, sheet=None):
    return ColRange(start-1, stop-1, book=book, sheet=sheet)
    
class CellRange(Range):
    def __init__(self, col_range, row_range, book=None, sheet=None, originCell=None, single_cell_clicked=False):
        if isinstance(sheet, Sheet):
            book = sheet.book
        elif isinstance(book, Book):
            sheet = book.sheet
        
        self.book = book
        self.sheet = sheet
        self.col_range = col_range if not isinstance(col_range, EmptyRange) else EmptyColRange()
        self.row_range = row_range if not isinstance(row_range, EmptyRange) else EmptyRowRange()
        self.single_cell_clicked = single_cell_clicked
        self.experimental = book.experimental if isinstance(book, Book) else False
        
        self.col_range.book = book
        self.col_range.sheet = sheet
        self.row_range.book = book
        self.row_range.sheet = sheet
        
        self.originCell = originCell if isinstance(originCell, Cell) else self.getStartingCell() if sheet else None
        
        self.setColLength(col_range)
        self.setRowLength(row_range)
        
    @property
    def col_range(self):
        return self._col_range
    @col_range.setter
    def col_range(self, new_col_range):
        self._col_range = new_col_range
        self.setColLength(new_col_range)
        # self.autoSelect()
        
    @property
    def row_range(self):
        return self._row_range
    @row_range.setter
    def row_range(self, new_row_range):
        self._row_range = new_row_range
        self.setRowLength(new_row_range)
        # self.autoSelect()
    
    def setColLength(self, col_range, event=None):
        self.col_length = self.nCols = col_range.stop - col_range.start + 1
    
    def setRowLength(self, row_range, event=None):
        self.row_length = self.nRows = row_range.stop - row_range.start + 1
    
    def autoSelect(self, event=None):
        self.sheet.currentCell = self.getStartingCell()
        self.sheet.lastCell    = current_range.getLastCell()
        self.sheet.removeSelectionMark()
        self.toggleCellColor(eventInherited=event, select=True)
        self.sheet.currentCell.toggleCellColor(eventInherited=event, select=False)
        
    def getStartingCell(self):
        return self.sheet[self.col_range.start, self.row_range.start]
        
    def getLastCell(self):
        return self.sheet[self.col_range.stop, self.row_range.stop]
        
    def declare_empty(self):
        return EmptyCellRange(self.col_range, self.row_range, book=self.book, sheet=self.sheet)
        
    def generateColIndexList(self):
        return self.col_range.generateColIndexList()
        
    def generateRowIndexList(self):
        return self.row_range.generateRowIndexList()
        
    def generateIndexList(self):
        return self.col_range.generateColIndexList(), self.row_range.generateRowIndexList()
        
    def getColNum(self):
        return list(self.col_range.ranger)
        
    def getRowNum(self):
        return list(self.row_range.ranger)
        
    def within(self, other):
        return andExec(self.col_range.within(other.col_range), self.row_range.within(other.row_range))
        
    def is_one_cell(self):
        return andExec(self.col_range.is_one_line(), self.row_range.is_one_line())
        
    def shift_n_units(self, col_shift=0, row_shift=0):
        col_range = self.col_range.shift_n_units(col_shift)
        row_range = self.row_range.shift_n_units(row_shift)
        return self.__class__(col_range, row_range, book=self.book, sheet=self.sheet)
        
    def extendHorizontal(self, stretch):
        self.col_range.extend_n_units(stretch)
        self.switchOffSelector()
        self.switchOnSelector()
        self.toggleCellColor(select=True, selectMultipleInOne=True)
        
    def extendVertical(self, stretch):
        self.row_range.extend_n_units(stretch)
        self.switchOffSelector()
        self.switchOnSelector()
        self.toggleCellColor(select=True, selectMultipleInOne=True)
        
    def extend_n_units(self, col_stretch=0, row_stretch=0):
        if row_stretch:
            self.extendHorizontal(row_stretch)
        if col_stretch:
            self.extendVertical(col_stretch)
        
    def __getitem__(self, index, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        return sheet[index]
        
    def __setitem__(self, index, value, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        sheet[index] = value
        
    def __repr__(self):
        sheet_display = self.sheet if not self.experimental else 'app[0,0]'
        
        if isinstance(self.sheet, Sheet):
            generated_content = self.generatePlus()
            if generated_content:
                return f'CellRange( {self.col_range} , {self.row_range} , sheet={sheet_display} )'
            else:
                return self.declare_empty().__repr__()
                # return f'EmptyCellRange( {self.col_range} , {self.row_range} , sheet={sheet_display} )'
        else:
            return f'CellRange( {self.col_range} , {self.row_range} )'

    def __str__(self):
        return self.__repr__()
        
    def getSize(self):
        return self.col_range.getSize() * self.row_range.getSize()
        
    def __iter__(self):
        for cell in self.sheet.cells:
            yield cell
        
    def __lt__(self, other):
        return self.getSize() < other.getSize()
        
    def __gt__(self, other):
        return self.getSize() > other.getSize()
        
    def __le__(self, other):
        return self.getSize() <= other.getSize()
        
    def __ge__(self, other):
        return self.getSize() >= other.getSize()
        
    def __eq__(self, other):
        if type(other) == type(self):
            return andExec(self.col_range == other.col_range, self.row_range == other.row_range)
        else:
            return False
        
    def __ne__(self, other):
        if type(other) == type(self):
            return orExec(self.col_range != other.col_range, self.row_range != other.row_range)
        else:
            return False
        
    def decRowRange(self):
        self.row_range -= 1
        
    def decColRange(self):
        self.col_range -= 1
    
    def flatten(self, to_end=False, vertical=False, book=None, sheet=None):
        col_range, row_range = self.col_range, self.row_range
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        if not isinstance(book, Book):
            book = self.book
        if vertical:
            if to_end:
                new_col_range = ColRange(col_range.stop)
                new_range = row_range.generateRange(col_range=new_col_range, book=book, sheet=sheet)
            else:
                new_col_range = ColRange(col_range.start-1)
                new_range = row_range.generateRange(col_range=new_col_range, book=book, sheet=sheet)
        else:
            if to_end:
                new_row_range = RowRange(row_range.stop)
                new_range = col_range.generateRange(row_range=new_row_range, book=book, sheet=sheet)
            else:
                new_row_range = RowRange(row_range.start-1)
                new_range = col_range.generateRange(row_range=new_row_range, book=book, sheet=sheet)
        # else:
            # new_range = CellRange(start, stop, book, sheet)
        return new_range
        
    def includes(self, cell):
        return andExec(self.col_range.includes(cell), self.row_range.includes(cell))
        
    def is_flat(self):
        return orExec(self.col_range.start == self.col_range.stop, self.row_range.start == self.row_range.stop)
        
    def is_empty(self, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        return not self.generate(sheet=sheet)
        
    def reverse(self, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        self.col_range.reverse()
        self.row_range.reverse()

    def __add__(self, other):
        if type(other) == type(self):
            return self.generate() + other.generate()
        elif isinstance(other, numbers.Number):
            start_sum, stop_sum = self.col_range+other, self.row_range+other
        return self.__class__(start_sum, stop_sum, book=self.book, sheet=self.sheet)

    def __sub__(self, other):
        # self  = set B
        # other = set A
        
        A, B = other, self
        
        if isinstance(other, numbers.Number):
            diff = self.__class__(self.col_range-other, self.row_range-other, book=self.book, sheet=self.sheet)
        
        elif andExec(A.col_range < B.col_range, A.row_range > B.row_range):
            diff_col = A.col_range-B.col_range
            smaller_row = B.row_range
            diff = self.__class__(diff_col, smaller_row, book=self.book, sheet=self.sheet)
            
        elif andExec(A.col_range > B.col_range, A.row_range < B.row_range):
            smaller_col = B.col_range
            diff_row = A.row_range-B.row_range
            diff = self.__class__(smaller_col, diff_row, book=self.book, sheet=self.sheet)
        
        elif other.within(self):
            # self  = set U
            # other = set A
            UColStart = self.col_range.start
            UColStop = self.col_range.stop
            URowStart = self.row_range.start
            URowStop = self.row_range.stop
            
            AColStart = other.col_range.start
            AColStop = other.col_range.stop
            ARowStart = other.row_range.start
            ARowStop = other.row_range.stop
            
            col_range_west = ColRange(UColStart, AColStart-1)
            col_range_east = ColRange(AColStop+1, UColStop)
            col_range_big_north = col_range_big_south = ColRange(UColStart, UColStop)
            
            row_range_west = row_range_east = RowRange(ARowStart, ARowStop)
            row_range_big_north = RowRange(URowStart, ARowStart-1)
            row_range_big_south = RowRange(ARowStop+1, URowStop)
            
            col_range_N_A = EmptyRange()
            row_range_N_A = EmptyRange()
            # area_N_A = self.__class__(col_range_N_A, row_range_N_A, book=self.book, sheet=self.sheet)
            area_N_A = EmptyCellRange(col_range_N_A, row_range_N_A, book=self.book, sheet=self.sheet)
            
            west = self.__class__(col_range_west, row_range_west, book=self.book, sheet=self.sheet) if UColStart != AColStart else area_N_A
            east = self.__class__(col_range_east, row_range_east, book=self.book, sheet=self.sheet) if UColStop != AColStop else area_N_A
            big_north = self.__class__(col_range_big_north, row_range_big_north, book=self.book, sheet=self.sheet) if URowStart != ARowStart else area_N_A
            big_south = self.__class__(col_range_big_south, row_range_big_south, book=self.book, sheet=self.sheet) if URowStop != ARowStop else area_N_A
            
            diff = CellSet(self.book, self.sheet, [west, east, big_north, big_south])
        
        elif orExec(self <= other, A.col_range.same_size_with(B.col_range), A.row_range.same_size_with(B.row_range)):
            diff = EmptyCellRange(EmptyRange(), EmptyRange(), book=self.book, sheet=self.sheet)
        
        return diff
        
    def focusOnStartingCell(self, event=None):
        startCell = self.getStartingCell()
        startCell.focusOnCell(event=event)
        
    def toggleStartingCell(self, event=None, *args, **kwargs):
        startCell = self.getStartingCell()
        startCell.toggleCell(event=event, *args, **kwargs)
        
    def focusOnRange(self, event=None):
        self.focusOnStartingCell(event=event)
        self.toggleCellColor(eventInherited=event, select=True)
        
    def toggleCellColor(self, eventInherited=None, select=True, previous_range=None, selector_trigger=True, debug=0, *args, **kwargs):
        self.mapPlus( lambda cell: (
                                cell.toggleCellColor(eventInherited=eventInherited, select=select, *args, **kwargs),
                                # print(cell) if not select else None
                                ))
        # if selector_trigger:
        if andExec(selector_trigger, not isinstance(self, EmptyRange)):
            if select:   self.switchOnSelector (*args, **kwargs)
            else:        self.switchOffSelector(*args, **kwargs)
            self.originCell.toggleCellColor(eventInherited=eventInherited, select=False)
        # if andExec(selector_trigger, isinstance(self, EmptyRange)):
            # # print(4781, self)
            # if previous_range:
                # previous_range.switchOffSelector()
            # if select:
                # for selectedRange in self.sheet.selectedCellsSet:
                    # # selectedRange.originCell.toggleCellColor(eventInherited=eventInherited, select=False)
                    # selectedRange.switchOnSelector()
            
    def checkOuterBorder(self):
        displayTypes = []
        checkOuterBorders = [lambda cell: displayTypes.append(cell.rowBorder.winfo_manager()) if cell else None,
                             lambda cell: displayTypes.append(cell.colBorder.winfo_manager()) if cell else None]
        
        for to_end, vertical in zip(to_end_list, vertical_list):
            flattened_range = self.flatten(to_end=to_end, vertical=vertical)
            flattened_range.mapPlus(checkOuterBorders[vertical])
        return displayTypes
            
    def notAllOuterBorder(self):
        return '' in self.checkOuterBorder()
            
    def checkAllBorders(self):
        displayTypes = []
        for cell_range in self.generateRangesPlus():
            displayTypes += cell_range.checkOuterBorder()
        return displayTypes
            
    def notAllBorders(self):
        return '' in self.checkAllBorders()
        
    def superFlatten(self):
        flattened_ranges = []
        for to_end, vertical in zip(to_end_list, vertical_list):
            flattened_range = self.flatten(to_end=to_end, vertical=vertical)
            flattened_range.mapPlus(flattened_ranges.append)
        return flattened_ranges
        
    def switchOnOuterBorder(self, *args, **kwargs):
        switchFuncs = [lambda cell: cell.rowBorder.install() if cell else None,
                     lambda cell: cell.colBorder.install() if cell else None]
        
        for to_end, vertical in zip(to_end_list, vertical_list):
            flattened_range = self.flatten(to_end=to_end, vertical=vertical)
            flattened_range.mapPlus(switchFuncs[vertical])
    
    def switchOffOuterBorder(self, *args, **kwargs):
        switchFuncs = [lambda cell: cell.rowBorder.takeoff() if cell else None,
                      lambda cell: cell.colBorder.takeoff() if cell else None]
        
        for to_end, vertical in zip(to_end_list, vertical_list):
            flattened_range = self.flatten(to_end=to_end, vertical=vertical)
            flattened_range.mapPlus(switchFuncs[vertical])
        
    def switchOnAllBorders(self, *args, **kwargs):
        self.mapPlus(lambda cell: cell.switchOnOuterBorder(*args, **kwargs) if cell else None)
        
    def switchOffAllBorders(self, *args, **kwargs):
        self.mapPlus(lambda cell: cell.switchOffOuterBorder(*args, **kwargs) if cell else None)
        
    def toggleAllBorders(self, *args, **kwargs):
        if self.notAllBorders():
            self.switchOnAllBorders(self, *args, **kwargs)
        else:
            self.switchOffAllBorders(self, *args, **kwargs)
        
    def toggleOuterBorder(self, *args, **kwargs):
        if self.notOuterBorder():
            self.switchOnOuterBorder(self, *args, **kwargs)
        else:
            self.switchOffOuterBorder(self, *args, **kwargs)
        
    def switchOnSelector(self, color_trigger=False, pink_purple=False, *args, **kwargs):
        switchFuncs = [lambda cell: cell.rowSelector.install(pink_purple) if cell else None,
                     lambda cell: cell.colSelector.install(pink_purple) if cell else None]
        for to_end, vertical in zip(to_end_list, vertical_list):
            flattened_range = self.flatten(to_end=to_end, vertical=vertical)
            flattened_range.mapPlus(switchFuncs[vertical])
            # flattened_range.mapPlus(lambda cell: print(f'to_end, vertical = {to_end, vertical}', cell) )
            # print(f'to_end, vertical = {to_end, vertical}', flattened_range)
        if andExec(self.col_range.stop == self.sheet.nCols, self.row_range.start == 0):
            switchFuncs[0](self.sheet.placeholderColumn.colTitleCell)
        if color_trigger:
            self.toggleCellColor(select=True)
    
    def switchOffSelector(self, color_trigger=False, to_end_list=to_end_list, vertical_list=vertical_list, debug=False, *args, **kwargs):
        # print(4864, self)
        switchFuncs = [lambda cell: cell.rowSelector.takeoff() if cell else None,
                      lambda cell: cell.colSelector.takeoff() if cell else None]
        if debug:
            print(4884, 'to_end_list, vertical_list =', to_end_list, vertical_list)
        for to_end, vertical in zip(to_end_list, vertical_list):
            flattened_range = self.flatten(to_end=to_end, vertical=vertical)
            flattened_range.mapPlus(switchFuncs[vertical])
        if andExec(self.col_range.stop == self.sheet.nCols, self.row_range.start == 0):
            switchFuncs[0](self.sheet.placeholderColumn.colTitleCell)
        if color_trigger:
            self.toggleCellColor(select=False)
        
    def replaceFormulae(self, content, col_range=None, row_range=None, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        self.generateRange(col_range=col_range, row_range=row_range, sheet=sheet).replaceFormulae(content)
        
    def generateContents(self, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
            
        selectedCellsContents1 = [[]]
        
        selectors = sheet[self.col_range.slicer, self.row_range.slicer]
        set1, set2 = selectors.concat_at_front([None]), selectors
        
        for cell_before, cell_after in zip(set1, set2):
            # print(cell_before, cell_after)
            if orExec(not cell_before, cell_before.index[0] == cell_after.index[0]):
                selectedCellsContents1[-1] += [cell_after.getContent()]
            else:
                selectedCellsContents1 += [[cell_after.getContent()]]
        
        # print('selectedCellsContents1 =', selectedCellsContents1)
                
        selectedCellsContents2 = ['\n'.join(cont) for cont in selectedCellsContents1]
        selectedCellsContents3 = '\t'.join(selectedCellsContents2)
        return selectedCellsContents3
    
    def replaceFormulae(self, content):
        self.map( lambda cell: cell.replaceFormula(content) )
    
    def isHighEnough(self):
        return self.row_range.isHighEnough() and self.col_range.isHighEnough()
    
    def map(self, function, debug=None):
        if debug: print(f'col_range, row_range = {self.col_range}, {self.row_range}')
        for colIndex in self.col_range:
            for rowIndex in self.row_range:
                if debug: print(f'colIndex, rowIndex = {colIndex}, {rowIndex}')
                try:
                    if andExec(colIndex >= 0, rowIndex >= 0):
                        function( self[colIndex, rowIndex] )
                except IndexError:
                    pass
                except Exception as e:
                    raise
                    # print(format_exc(), file=sys.stderr)
                    # print(type(e).__name__, e, f'at colIndex, rowIndex = {colIndex}, {rowIndex}')
        
    def mapPlus(self, function, book=None, sheet=None, ranges=False, debug=None, wanted_type=Cell):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        if not isinstance(book, Book):
            book = self.book
        
        if self.isHighEnough():
            sheet[self].map(function, debug=debug)
        
        # self.map(function, debug=debug)
        
        # emptyCellSet = CellSet(self.book, self.sheet)
        # array1 = sheet.indexColumn[self.row_range] if self.col_range.start == -1 else emptyCellSet
        # array2 = sheet.colTitleCells[self.col_range] if self.row_range.start == -1 else emptyCellSet
        # array3 = sheet.placeholderColumn[self.row_range] if self.col_range.stop == self.nCols else emptyCellSet
        # array4 = [col.placeholderCell for col in sheet[self.col_range] ] if self.row_range.stop == self.nRows else emptyCellSet
        
        # if andExec(self.col_range.start < 0, self.row_range.start < 0):
            # additionalArray = sheet.indexColumn.cells[self.row_range] + sheet.colTitleCells[self.col_range]
        # elif self.col_range.start < 0:
            # additionalArray = sheet.indexColumn.cells[self.row_range]
        # elif self.row_range.start < 0:
            # additionalArray = sheet.colTitleCells[self.col_range]
        # else:
            # return
        
        # additionalArray = array1 + array2 + array3 + array4
        # additionalArray.map(function, debug=debug)
        
    def generate(self, book=None, sheet=None, ranges=False):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        if not isinstance(book, Book):
            book = self.book
        col_range = self.col_range if self.col_range.start >= 0 else ColRange(0, self.col_range.stop, book=book, sheet=sheet)
        row_range = self.row_range if self.row_range.start >= 0 else RowRange(0, self.row_range.stop, book=book, sheet=sheet)
        generateArray = sheet[col_range, row_range]
        if ranges:
            generateArray = CellSet(book, sheet, map(lambda cell: cell.generateRange(), generateArray) )
        return generateArray
        
    def generatePlus(self, book=None, sheet=None, ranges=False):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        if not isinstance(book, Book):
            book = self.book
        inputCells = self.generate()
        if andExec(self.col_range.start < 0, self.row_range.start < 0):
            generateArray = sheet.indexColumn.cells[self.row_range] + sheet.colTitleCells[self.col_range] + inputCells
        elif self.col_range.start < 0:
            generateArray = sheet.indexColumn.cells[self.row_range] + inputCells
        elif self.row_range.start < 0:
            generateArray = sheet.colTitleCells[self.col_range] + inputCells
        else:
            generateArray = inputCells
        if ranges:
            generateArray = CellSet(book, sheet, map(lambda cell: cell.generateRange(), generateArray) )
        return CellSet(book, sheet, sorted(generateArray) )
        
    def generateRangesPlus(self, book=None, sheet=None):
        return self.generatePlus(book=book, sheet=sheet, ranges=True)
        
        
class EmptyCellRange(CellRange, EmptyRange):
    def __init__(self, col_range=None, row_range=None, book=None, sheet=None):
        if isinstance(sheet, Sheet):
            book = sheet.book
        elif isinstance(book, Book):
            sheet = book.sheet
        
        self.book = book
        self.sheet = sheet
        self.col_range = col_range if isinstance(col_range, ColRange) else ColRange(-2, -1)
        self.row_range = row_range if isinstance(row_range, RowRange) else RowRange(-2, -1)
        self.experimental = book.experimental
    
    def __repr__(self):
        sheet_display = self.sheet if not self.experimental else 'app[0,0]'
        if andExec(self.col_range == ColRange(-2, -1), self.row_range == RowRange(-2, -1)):
            self.range_input = ''
        elif self.col_range == ColRange(-2, -1):
            self.range_input = f'{self.row_range}'
        elif self.row_range == RowRange(-2, -1):
            self.range_input = f'{self.col_range}'
        else:
            self.range_input = f'{self.col_range} , {self.row_range}'
            
        if isinstance(self.sheet, Sheet):
            if self.range_input: self.range_input += f' , '
            return f'EmptyCellRange( {self.range_input}sheet={sheet_display} )'
        else:
            return f'EmptyCellRange( {self.range_input} )'
        
    def __str__(self):
        return self.__repr__()
            
    def declare_empty(self, cell_range):
        return self

    def focusOnStartingCell(self, event=None, toggle_cell=True):
        pass
        
    def toggleStartingCell(self, event=None, *args, **kwargs):
        pass
        
    def focusOnRange(self, event=None, toggle_cell=True):
        pass
        
    def toggleCellColor(self, eventInherited=None, *args, **kwargs):
        pass
        
    def replaceFormulae(self, content, col_range=None, row_range=None, sheet=None):
        pass
        
    def generateContents(self, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        return CellSet(self.book, sheet, [])
    
    def replaceFormulae(self, content):
        pass
        
    def map(self, function, debug=None):
        pass
        
    def generate(self, sheet=None):
        if not isinstance(sheet, Sheet):
            sheet = self.sheet
        return CellSet(self.book, sheet, [])

class CellSet:
    def __init__(self, book=None, sheet=None, cellSet=None, label_status=False, name=None):
        if isinstance(sheet, Sheet):
            book = sheet.book
        elif isinstance(book, Book):
            sheet = book.sheet
        
        self.book = book
        self.sheet = sheet
        self.cellSet = tuple(cellSet) if cellSet else ()
        self.experimental = book.experimental
        
        self.currentCellReaderLabel_status = label_status
        
        self.selectedColsIndex = []
        self.selectedRowsIndex = []
        
        self.name = name
        
    def __repr__(self):
        sheet_display = self.sheet if not self.experimental else 'app[0,0]'
        return f'CellSet(cellSet={list(self.cellSet)}, sheet={sheet_display})'
        
    def __getitem__(self, index):
        # try:
        if isinstance(index, int): return self.cellSet[index]
        elif isinstance(index, RowRange) and set(map(type,self.cellSet)) == {Column}:
            selectedCellsSet = self.__class__(self.book, self.sheet)
            for col in self.cellSet:
                selectedCellsSet += col[index]
            return self.__class__(self.book, self.sheet, selectedCellsSet )
        elif isinstance(index, Range): return self.__class__(self.book, self.sheet, self.cellSet[index.slicer] )
        elif isinstance(index, slice): return self.__class__(self.book, self.sheet, self.cellSet[index] )
        # except IndexError:
            # return
        
    def modify_currentCellSelReader(self, init=False):
        if hasattr(self, 'generateIndexList'):
            self.selectedColsIndex, self.selectedRowsIndex = self.generateIndexList()
        if self.currentCellReaderLabel_status:
            self.sheet.modify_currentCellSelReader(self.selectedColsIndex, self.selectedRowsIndex, init=init)
        
    def __setitem__(self, index, value):
        if len(self) < 0:
            cellSet = (value)
            return
        if self.cellSet[index] == value:
            return
        cellSet = ()
        if index < 0:
            index += len(self)
        for i in range(len(self.cellSet)):
            if i == index:
                cellSet += (value,)
                if self.name == 'selectedCellsSet':
                    value.toggleCellColor(select=True)
                    self.cellSet[i].toggleCellColor(select=False)
            else:
                cellSet += (self.cellSet[i],)
        self.cellSet = cellSet
        if self.name == 'selectedCellsSet':
            self.modify_currentCellSelReader()
        
    def __str__(self):
        return self.__repr__()
        
    # def __bool__(self):
        # return bool(self.cellSet)
        
    def __len__(self):
        return len(self.cellSet)
        
    def __add__(self, other):
        return self.concat(other)
        
    def __iter__(self):
        for cell_or_subset in self.cellSet:
            yield cell_or_subset
        
    # def includes(self, cell):
        # for cell_or_subset in self.cellSet:
            # 
        
    def copy(self):
        return self.__class__(self.book, self.sheet, self.cellSet)
        
    def is_flat(self):
        """ Checks whether a cell set is either a flat row or a flat column """
        flat_num, empty_num = 0, 0
        for cell_or_subset in self.cellSet:
            # If one is flat and all others are empty then return True
            if cell_or_subset.is_flat(): flat_num += 1
            elif cell_or_subset.is_empty(): empty_num += 1
            elif type(cell_or_subset) != CellRange: return False
        return andExec(flat_num == 1, empty_num == len(self.cellSet)-1)
        
    def getColNum(self):
        colNums = set()
        for cell_or_subset in self.cellSet:
            if isinstance(cell_or_subset, LineRange):
                increment = cell_or_subset.getLineNum()
            elif isinstance(cell_or_subset, CellRange):
                increment = cell_or_subset.getColNum()
            elif isinstance(cell_or_subset, Cell):
                increment = [cell_or_subset.index[0]]
            colNums.update(set(increment))
        colNums = sorted(colNums)
        return colNums
        
    def getRowNum(self):
        rowNums = set()
        for cell_or_subset in self.cellSet:
            if isinstance(cell_or_subset, LineRange):
                increment = cell_or_subset.getLineNum()
            elif isinstance(cell_or_subset, CellRange):
                increment = cell_or_subset.getRowNum()
            elif isinstance(cell_or_subset, Cell):
                increment = [cell_or_subset.index[1]]
            rowNums.update(set(increment))
        rowNums = sorted(rowNums)
        return rowNums
        
    def getColSize(self):
        nColsSize = 0
        for cell_or_subset in self.cellSet:
            if isinstance(cell_or_subset, LineRange):
                increment = cell_or_subset.getSize()
            elif isinstance(cell_or_subset, CellRange):
                increment = cell_or_subset.col_range.getSize()
            elif isinstance(cell_or_subset, Cell):
                increment = 1
            nColsSize += increment
        return nColsSize
        
    def getRowSize(self):
        nColsSize = 0
        for cell_or_subset in self.cellSet:
            if isinstance(cell_or_subset, LineRange):
                increment = cell_or_subset.getSize()
            elif isinstance(cell_or_subset, CellRange):
                increment = cell_or_subset.row_range.getSize()
            elif isinstance(cell_or_subset, Cell):
                increment = 1
            nColsSize += increment
        return nColsSize
        
    def generateColIndexList(self):
        colIndex = []
        for cell_or_subset in self.cellSet:
            colIndex_instance = cell_or_subset.generateColIndexList()
            colIndex += sorted(colIndex_instance, reverse=True) + ['']
        return colIndex
        
    def generateRowIndexList(self):
        rowIndex = []
        for cell_or_subset in self.cellSet:
            rowIndex_instance = cell_or_subset.generateRowIndexList()
            rowIndex += sorted(rowIndex_instance) + ['']
        return rowIndex
        
    def generateIndexList(self):
        colIndex, rowIndex = set(), set()
        # print('self =', type(self) )
        # print('self.cellSet =', self.cellSet)
        for cell_or_subset in self.cellSet:
            if hasattr(cell_or_subset, 'generateIndexList'):
                colIndex_instance, rowIndex_instance = cell_or_subset.generateIndexList()
                colIndex.update(colIndex_instance)
                rowIndex.update(rowIndex_instance)
        return colIndex, rowIndex
        
    def sort(self, /, *, key=None, reverse=False):
        cellSet = list(self.cellSet)
        cellSet.sort(key=key, reverse=reverse)
        self.cellSet = tuple(cellSet)
        
    def start(self, cell_or_subset, init=False):
        if isinstance(cell_or_subset, CellLabel):
            cell_or_subset = cell_or_subset.generateRange()
        self.cellSet = ( cell_or_subset , )
        if self.name == 'selectedCellsSet':
            self.modify_currentCellSelReader(init=init)
            self.sheet.focus_set()
        
    def restart(self, cell_or_subset, init=False):
        self.start(cell_or_subset, init=init)
        
    def add(self, cell_or_subset, init=False, cell_raw=False):
        if andExec(type(cell_or_subset) == CellLabel, not cell_raw):
            cell_or_subset = cell_or_subset.generateRange()
        if cell_or_subset not in self.cellSet:
            self.cellSet += ( (cell_or_subset,) )
        if self.name == 'selectedCellsSet':
            cell_or_subset.toggleCellColor(select=True)
            self.modify_currentCellSelReader(init=init)
        
    def append(self, cell_or_subset, init=False):
        self.add(cell_or_subset, init=init)
        
    def exclusify(self, exclusiveness):
        cellSet = ()
        for cell_or_subset in self.cellSet:
            if exclusiveness(cell_or_subset):
                cellSet += (cell_or_subset,)
        return self.__class__(self.book, self.sheet, self.cellSet)
        
    def remove(self, removed_cell_or_subset, init=False, affect_display=False):
        cellSet = ()
        if isinstance(removed_cell_or_subset, LineRange):
            entire_list = self[removed_cell_or_subset]
            for cell_or_subset in self.cellSet:
                if cell_or_subset in entire_list:
                    if isinstance(cell_or_subset, Cell):
                        cell_or_subset.pack_forget()
                        cell_or_subset.cellFrame.grid_forget()
                    elif isinstance(cell_or_subset, Column):
                        for rowTopSeparator in cell_or_subset.rowTopSeparators:
                            rowTopSeparator.takeoff()
                else:
                    cellSet += (cell_or_subset,)
        elif hasattr(removed_cell_or_subset, '__iter__'):
            entire_list = removed_cell_or_subset
            for cell_or_subset in self.cellSet:
                if cell_or_subset not in entire_list:
                    cellSet += (cell_or_subset,)
        else:
            for cell_or_subset in self.cellSet:
                cell_or_subset_deleted = cell_or_subset.deleted if isinstance(cell_or_subset, Cell) else False
                if andExec(cell_or_subset != removed_cell_or_subset, not cell_or_subset_deleted):
                    cellSet += (cell_or_subset,)
        self.cellSet = cellSet
        if self.name == 'selectedCellsSet':
            removed_cell_or_subset.toggleCellColor(select=False)
            self.modify_currentCellSelReader(init=init)
        
    def __delitem__(self, removed_cell_or_subset):
        self.remove(self, removed_cell_or_subset)
        
    def insert_at_index(self, index: int, value: object) -> None:
        cellSet = ()
        for i in range(len(self.cellSet)):
            if i == index: cellSet += (value, self.cellSet[i])
            else: cellSet += (self.cellSet[i],)
        self.cellSet = cellSet
        if self.name == 'selectedCellsSet':
            self.modify_currentCellSelReader(init=init)
            value.toggleCellColor(select=True)
        
    def remove_at_index(self, index: int) -> None:
        cellSet = ()
        for i in range(len(self.cellSet)):
            cell_or_subset = self.cellSet[i]
            if i == index:
                if isinstance(cell_or_subset, Cell):
                    cell_or_subset.pack_forget()
                    cell_or_subset.cellFrame.grid_forget()
            else:
                cellSet += (cell_or_subset,)
        self.cellSet = cellSet
        if self.name == 'selectedCellsSet':
            self.modify_currentCellSelReader(init=init)
            self.cellSet[index].toggleCellColor(select=False)
        
    def insert_at_front(self, value: object) -> None:
        self.insert_at_index(0, value)
        
    def concat(self, cellSet2):
        if type(cellSet2) == type(self):      concat_set = self.cellSet + cellSet2.cellSet
        elif isinstance(cellSet2, tuple):     concat_set = self.cellSet + cellSet2
        elif hasattr(cellSet2, '__iter__'):   concat_set = self.cellSet + tuple(cellSet2)
        return self.__class__(self.book, self.sheet, concat_set)
        
    def concat_at_front(self, cellSet2):
        if type(cellSet2) == type(self):      concat_set = cellSet2.cellSet + self.cellSet
        elif isinstance(cellSet2, tuple):     concat_set = cellSet2         + self.cellSet
        elif hasattr(cellSet2, '__iter__'):   concat_set = tuple(cellSet2)  + self.cellSet
        return self.__class__(self.book, self.sheet, concat_set)
        
    def shift_n_units(self, col_shift=0, row_shift=0):
        cellSet = ()
        for cell_or_subset in self.cellSet:
            cellSet += (cell_or_subset.shift_n_units(col_shift=col_shift, row_shift=row_shift) , )
        return self.__class__(self.book, self.sheet, cellSet)
        
    def toggleCellColor(self, eventInherited=None, *args, **kwargs):
        for cell_or_subset in self.cellSet:
            cellType = type(cell_or_subset)
            function = lambda cell: cell.toggleCellColor(eventInherited=eventInherited, *args, **kwargs)
            if cellType == CellRange:
                cell_or_subset.map(function)
                if cell_or_subset.row_range.start == -1:
                    for column, colTitleCell in zip(self.sheet[cell_or_subset.col_range], self.sheet.colTitleCells[cell_or_subset.col_range]):
                        column.rowSelector.manage(select=kwargs['select'])
                        colTitleCell.toggleCellColor(eventInherited=eventInherited, select=kwargs['select'])
            elif cellType == CellLabel:
                function(cell_or_subset)
        
    def destroy(self):
        for cell_or_subset in self.cellSet:
            cellType = type(cell_or_subset)
            function = lambda cell: cell.destroy()
            if cellType == CellRange:        cell_or_subset.map(function)
            elif cellType == CellLabel:      function(cell_or_subset)
        
    def cellCoord(self):
        for cell_or_subset in self.cellSet:
            cellType = type(cell_or_subset)
            if orExec(cellType == CellRange, hasattr(cellType, '__iter__')):
                cell_or_subset.map(function)
            else:
                function(cell_or_subset)
        
    def map(self, function, cell_absorbant=True, debug=None):
        for cell_or_subset in self.cellSet:
            cellType = type(cell_or_subset)
            if andExec(orExec(cellType == CellRange, hasattr(cellType, 'map')), cell_absorbant):
                cell_or_subset.map(function)
            elif cellType is None:
                continue
            else:
                function(cell_or_subset)
        
    def mapPlus(self, function, cell_absorbant=True, debug=None):
        self.map(function, cell_absorbant=cell_absorbant, debug=debug)
        # for cell_or_subset in self.cellSet:
            # cellType = type(cell_or_subset)
            # if andExec(orExec(cellType == CellRange, hasattr(cellType, '__iter__')), cell_absorbant):
                # cell_or_subset.mapPlus(function)
            # elif cellType is None:
                # continue
            # else:
                # function(cell_or_subset)
        
    def mapReturn(self, function, cell_absorbant=True, debug=None):
        cellSet = CellSet(self.book, self.sheet)
        for cell_or_subset in self.cellSet:
            cellType = type(cell_or_subset)
            if andExec(orExec(cellType == CellRange, hasattr(cellType, '__iter__')), cell_absorbant):
                cellSet.add( cell_or_subset.map(function) )
            elif cellType is None:
                continue
            else:
                cellSet.add( function(cell_or_subset) )
        return cellSet
        
    def generate(self):
        cells = self.__class__(self.book, self.sheet)
        for cell_or_subset in self.cellSet:
            cellType = type(cell_or_subset)
            if cellType == CellRange:   cells += cell_or_subset.generate()
            elif cellType == CellLabel:      cells += [cell_or_subset]
        return cells
        
    def __call__(self):
        return self.generate()
        
class HaveHeadersSelector(tk.Toplevel):
    def __init__(self, master, file_addr, sheets=None, *args, **kwargs):
        super().__init__(master, bg='#FFFFFF', *args, **kwargs)
        
        self.title("Sheet with/without Headers Selector")
        self.resizable(0,0)
        self.response = True
        
        width, height = 500, 300
        # Get the screen width and height
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Calculate the position of the window to center it on the screen
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

        # Set the geometry of the window to center it on the screen
        self.geometry(f"{width}x{height}+{x}+{y}")
        self.master = master
        
        self.file_addr = file_addr
        file_addr1, self.extension = os.path.splitext(file_addr)
        self.file_dir, self.filename = os.path.split(file_addr1)
        self.extension = self.extension.lower()
        
        self.sheets = sheets
        self.selected_sheets = []

        # self.file_label = tk.Label(self, text="Select Excel File:")
        # self.file_label.pack(pady=10)

        # self.file_button = tk.Button(self, text="Select File", command=self.select_file)
        # self.file_button.pack()

        self.sheet_label = tk.Label(self, text='Select which Sheets whose 1st row you consider to be the Column Title (Header) Row. (If there are no Sheets having headers, then simply click "Select Sheets" button.)', font='Arial 10 bold', bg='#FFFFFF', wraplength=300)

        self.canvas = tk.Canvas(self, borderwidth=0, bg='#FFFFFF')
        self.frame = tk.Frame(self.canvas, bg='#FFFFFF')
        self.vsb = tk.Scrollbar(self.canvas, orient="vertical", command=self.canvas.yview, bg='#FFFFFF')
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.frame.bind("<Configure>", self.onFrameConfigure)
        self.canvas.bind("<MouseWheel>", self.onMouseWheel)

        self.select_button = tk.Button(self, text="Select Sheets", command=self.select_sheets, bg='#FFFFFF')

        # self.sheet_label.grid(row=0, column=0, columnspan=2)
        # self.canvas.grid(row=1, column=0)
        # self.select_button.grid(row=1, column=1)
        
        self.sheet_label.pack(pady=10)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.select_button.pack()
        
        self.vsb.pack(side="right", fill="y")
        self.canvas.create_window((4,4), window=self.frame, anchor="n")
        
        self.protocol("WM_DELETE_WINDOW", self.close)
        
        self.load_sheets()
        
        # if getattr(sys, 'frozen', False):
        # self.mainloop()

    def close(self):
        self.destroy()
        self.response = False
        
    def onFrameConfigure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def onMouseWheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/240)), "units")

    def select_file(self):
        file_addr = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_addr:
            self.load_sheets(file_addr)

    def load_sheets(self):
        for widget in self.frame.winfo_children():
            widget.destroy()
        
        if self.extension.startswith('.xls'):
            # print(self.file_addr)
            wb = openpyxl.load_workbook(self.file_addr)
            self.sheets = wb.sheetnames
        elif self.extension == '.qutable':
            self.sheets = self.sheets
        elif self.extension == '.csv':
            self.sheets = [self.filename]

        self.checkboxes = {}
        for sheet in self.sheets:
            self.checkboxes[sheet] = False
            checkbox = tk.Checkbutton(self.frame, text=sheet)
            checkbox.bind('<Button-1>', lambda event, sheet=sheet: self.toggle_checkbox(sheet))
            checkbox.pack(anchor="w")

    def toggle_checkbox(self, sheet):
        self.checkboxes[sheet] = not self.checkboxes[sheet]

    def select_sheets(self_object):
        global app, self
        self_object.selected_sheets = [sheet for sheet, var in self_object.checkboxes.items() if var]
        self_object.destroy()
        self.response = True
        if 'app' not in globals():
            app = QuTable()
        self = app[0,0]

class FormulaEntry(tk.Entry):
    pass
    
class ColTitleEntry(tk.Entry):
    def __init__(parent, formulaEntryBox, *args, **kwargs):
        tk.Entry.__init__(parent, *args, **kwargs)
        self.formulaEntryBox = formulaEntryBox
    
class FormulaWizard(tk.Toplevel):
    row_count = 1
    
    def __init__(self, book, sheet, *args, **kwargs):
        tk.Toplevel.__init__(self, book.window, *args, **kwargs)
        self['bg'] = '#FFFFFF'
        # self.protocol("WM_DELETE_WINDOW", self.withdraw)
        # self.geometry('+225+415')
        
        self.title("Sheet with/without Headers Selector")
        self.resizable(0,0)
        self.response = True
        
        self._width = 1100
        self._height = 250
        self.center_window(self, width=self._width, height=self._height)
        
        # Create a frame to hold the grid system
        self.frame = tk.Frame(self, bg='#FFFFFF')
        self.frame.pack(expand=True)
        
        self.book = book
        self.sheet = sheet
        self.widgets = []
        
        self.choices = sheet.colTitles
        self.title('Column Formula Wizard')

        self.columnFormulaHelpBtn = tk.Button(self.frame, text="Click this if you need any help", command=self.columnFormulaHelpAction, bg='#FFFFFF')
        self.columnFormulaHelpBtn.grid(row=0, column=0, columnspan=3)

        self.colDropDownTitle = tk.Label(self.frame, text="\n\tNew Column Name(s)", bg='#FFFFFF', anchor='center', justify='center')
        self.formulaEntryTitle = tk.Label(self.frame, text="\n\t\t\tFormula Entry(-ies)", bg='#FFFFFF', anchor='center', justify='center')
        self.colDropDownTitle.grid(row=1, column=0)
        self.formulaEntryTitle.grid(row=1, column=1)
        
        self.rowFrame = tk.Frame(self.frame, bg='#FFFFFF')
        self.rowFrame.grid(row=2, column=0, columnspan=3)
        
        self.bottom_dist1 = tk.Frame(self.frame, height=20, bg='#FFFFFF')
        self.add_row_button = tk.Button(self.frame, text="Add Column", command=self.add_row, bg='#FFFFFF')
        self.bottom_dist2 = tk.Frame(self.frame, height=20, bg='#FFFFFF')
        self.submit_button = tk.Button(self.frame, text="Submit Formula", command=self.calculate, bg='#FFFFFF')

        self.add_row()
        
        for btn in [self.add_row_button, self.submit_button]:
            btn.bind('<Enter>', self.hoverButton)
            btn.bind('<Leave>', self.leaveButton)

        if getattr(sys, 'frozen', False):
            self.mainloop()
        
    @property
    def width(self):
        return self._width
    @width.setter
    def width(self, new_width):
        self._width = new_width
        self.center_window(self, width=self._width, height=self._height)
        
    @property
    def height(self):
        return self._height
    @height.setter
    def height(self, new_height):
        self._height = new_height
        self.center_window(self, width=self._width, height=self._height)
        
    def columnFormulaHelpAction(self):
        self.helpWindow = tk.Toplevel(self, bg='#FFFFFF')
        self.center_window(self.helpWindow, width=700, height=250)
        self.helpWindow.title('Sheet Headers Selector Help Window')
        
        # Create a frame to hold the pack system
        self.helpWindowFrame = tk.Frame(self.helpWindow, bg='#FFFFFF')
        self.helpWindowFrame.pack(expand=True)
        
        self.columnFormulaHelp = tk.Label(self.helpWindowFrame, text='Select the column whose values are to be calculated, by simply typing in the "Formula Entry(-ies)" on the right side. Type the nth column as the letter C followed by a number (column sequence).\nFor example, 1st column as C1, 2nd column as C2, and so on.\nFinally, type New Column Name(s) on the left for the name of new column(s) to be created.\n\nAn example would be "NewColFormula" on the left and "C1*C2" on the right. It means that "NewColFormula" would be created containing the products of the values of C1 and C2.\n', bg='#FFFFFF', wraplength=665)
        self.columnFormulaHelpLink = tk.Button(self.helpWindowFrame, text='Click here for more information.', bg='#FFFFFF', wraplength=300)
        
        self.columnFormulaHelp.pack()
        self.columnFormulaHelpLink.pack()
        
        if getattr(sys, 'frozen', False):
            self.helpWindow.mainloop()
        
    def center_window(self, window, width, height):
        # Get the screen width and height
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()

        # Calculate the position of the window to center it on the screen
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

        # Set the geometry of the window to center it on the screen
        window.geometry(f"{width}x{height}+{x}+{y}")
        
    def hoverButton(self, event):
        event.widget.configure(background='#d3d3d3')
        
    def leaveButton(self, event):
        event.widget.configure(background='#FFFFFF')
        
    def calculate(self):
        for colTitleEntry, formulaEntry, deleteBtn in self.widgets:
            colTitle, formula = colTitleEntry.get(), formulaEntry.get()
            resultColFormulaVars = createVarList( formula )
            resultColData = eval(formula, {var:self.book.sheet[var] for var in resultColFormulaVars})
            self.sheet.insertColRightMost()
            self.sheet[self.sheet.nCols-1].colTitle = colTitle
            self.sheet[self.sheet.nCols-1].replaceFormulae(resultColData)
        
    def add_row(self, init=False):
        # if self.height+175 > self.winfo_screenheight():
            # messagebox.showwarning('Warning', 'If you add more column formula entries, it will surpass the height of your screen. Therefore, this operation is deemed impossible to execute.', parent=self.book.window)
            
        if self.row_count > 5:
            messagebox.showwarning('Warning', 'If you add more column formula entries, it might cause unresponsiveness due to heavy works of calculations.', parent=self.book.window)
            
        else:
            self.row_count += 1
            
            entry_col = ColTitleEntry(self.rowFrame, width=35, justify='center', bg='#FFFFFF', font='Arial 14')
            entry_col.grid(row=self.row_count, column=0)
            entry_formula = FormulaEntry(self.rowFrame, width=35, justify='center', bg='#FFFFFF', font='Arial 14')
            entry_formula.grid(row=self.row_count, column=1)
            delete_button = tk.Button(self.rowFrame, text="Delete", command=lambda row=self.row_count: self.delete_row(row), bg='#FFFFFF')
            delete_button.grid(row=self.row_count, column=2)
            
            self.widgets.append( (entry_col, entry_formula, delete_button) )
            
            self.bottom_dist1.grid(row=self.row_count+2, column=0, columnspan=3)
            self.add_row_button.grid(row=self.row_count+3, column=0, columnspan=3)
            self.bottom_dist2.grid(row=self.row_count+4, column=0, columnspan=3)
            self.submit_button.grid(row=self.row_count+5, column=0, columnspan=3)
        
            self.height += 35

    def delete_row(self, row):
        for widget in self.frame.grid_slaves():
            if int(widget.grid_info()["row"]) == row:
                widget.grid_forget()
        del self.widgets[row]
        self.row_count -= 1
        
        self.height -= 35

class GraphPlotter:
    def __init__(self, book, sheet):
        self.book = book
        self.sheet = sheet
        
        self.master = master = tk.Toplevel(book.window, bg='#FFFFFF')
        master.title("Graph Plotter")
        master.geometry("1000x420")
        master.protocol("WM_DELETE_WINDOW", self.close)
        
        # Centering the window on the screen
        window_width = master.winfo_reqwidth()
        window_height = master.winfo_reqheight()
        position_right = int(master.winfo_screenwidth()/2 - window_width*2)
        position_down = int(master.winfo_screenheight()/2 - window_height)
        master.geometry(f"+{position_right}+{position_down}")
        
        # Create a frame to hold the grid system
        self.frame = tk.Frame(self.master, bg='#FFFFFF')
        self.frame.pack(expand=True, padx=20, pady=20)
        
        # Configure spacer1
        self.spacer1 = tk.Frame(self.frame, height=10, bg='#FFFFFF')
        self.spacer1.grid(column=0, row=0, sticky='nsew')
        
        # Configure spacer2
        self.spacer2 = tk.Frame(self.frame, height=50, bg='#FFFFFF')
        self.spacer2.grid(column=0, row=5, sticky='nsew', rowspan=100)
        
        sepBGWidth = 2
        sepNoBGWidth = 50
        
        # Plot Type Selection Option Menu
        # self.plot_type_label = tk.Label(self.frame, text="Select plot type:", bg='#FFFFFF')
        # self.plot_type_label.grid(column=0, row=0, sticky='nsew')
        # self.plot_type = tk.StringVar()
        # self.plot_type.set("scatter")  # Default plot type is scatter
        # self.plot_type_options = ["scatter", "bar"]
        # self.plot_type_menu = ttk.OptionMenu(self.frame, self.plot_type, self.plot_type.get(), *self.plot_type_options)
        # self.plot_type_menu.grid(column=0, row=1, sticky='nsew')
        
        # Plot Type Selection Buttons
        # self.plot_types = ["scatter", "bar"]
        self.plot_types = ["scatter"]
        self.change_plot_type(self.plot_types[0])
        self.plot_type_options = []
        self.plot_type_label = tk.Label(self.frame, text="Select plot type\n(For now,\nit's only scatter):", bg='#FFFFFF')
        self.plot_type_label.grid(column=0, row=1, sticky='nsew')
        i = 2                                        
        for plot_type in self.plot_types:
            plot_type_option = tk.Checkbutton(self.frame, text=plot_type, bg='#FFFFFF')
            plot_type_option.grid(column=0, row=i, sticky='nsew')
            self.bind_3_btns(plot_type_option, lambda event, plot_type=plot_type: self.change_plot_type(plot_type))
            self.plot_type_options.append(plot_type_option)
            i += 1
            
        self.plot_type_options[0].select()
        self.plot_type_options[0]['state'] = 'disabled'
        
        self.separator1a = tk.Frame(self.frame, height=100, width=sepNoBGWidth, bg='#FFFFFF')
        self.separator1a.grid(column=1, row=0, sticky='ns', rowspan=100)
        self.separator1 = tk.Frame(self.frame, height=100, width=sepBGWidth, bg='#000000')
        self.separator1.grid(column=2, row=0, sticky='ns', rowspan=100)
        self.separator1b = tk.Frame(self.frame, height=100, width=sepNoBGWidth, bg='#FFFFFF')
        self.separator1b.grid(column=3, row=0, sticky='ns', rowspan=100)
        
        self.column_x_label = tk.Label(self.frame, text="Select X-axis column\n(Choose only one):", bg='#FFFFFF', font='Arial 10 bold')
        self.column_x_label.grid(column=4, row=1, sticky='nsew')
        self.column_x_vars = {}
        self.column_x_options = {}
        self.one_x_selected = False
        
        self.separator2a = tk.Frame(self.frame, height=100, width=sepNoBGWidth, bg='#FFFFFF')
        self.separator2a.grid(column=5, row=0, sticky='ns', rowspan=100)
        self.separator2 = tk.Frame(self.frame, height=100, width=sepBGWidth, bg='#000000')
        self.separator2.grid(column=6, row=0, sticky='ns', rowspan=100)
        self.separator2b = tk.Frame(self.frame, height=100, width=sepNoBGWidth, bg='#FFFFFF')
        self.separator2b.grid(column=7, row=0, sticky='ns', rowspan=100)
        
        self.column_y_label = tk.Label(self.frame, text="Select Y-axis column\n(Choose one or more):", bg='#FFFFFF', font='Arial 10 bold')
        self.column_y_label.grid(column=8, row=1, sticky='nsew')
        self.column_y_vars = {}
        self.column_y_options = {}
        
        self.separator3a = tk.Frame(self.frame, height=100, width=sepNoBGWidth, bg='#FFFFFF')
        self.separator3a.grid(column=9, row=0, sticky='ns', rowspan=100)
        self.separator3 = tk.Frame(self.frame, height=100, width=sepBGWidth, bg='#000000')
        self.separator3.grid(column=10, row=0, sticky='ns', rowspan=100)
        self.separator3b = tk.Frame(self.frame, height=100, width=sepNoBGWidth, bg='#FFFFFF')
        self.separator3b.grid(column=11, row=0, sticky='ns', rowspan=100)
        
        # Plot/Unplot Buttons
        self.plot_button = tk.Button(self.frame, text="Plot Graph", command=self.plot_graph, bg='#FFFFFF')
        self.plot_button.grid(column=12, row=1, sticky='nsew')
        
        # self.unplot_button = tk.Button(self.frame, text="Unplot", command=self.unplot, bg='#FFFFFF')
        # self.unplot_button.grid(column=12, row=2, sticky='nsew')
        
        self.plotted_variables = []
        
    def change_plot_type(self, plot_type):
        self.plot_type = plot_type
        
    def close(self):
        self.master.withdraw()
        self.column_x_vars = {}
        self.column_x_options = {}
        self.column_y_vars = {}
        self.column_y_options = {}
        
    # def columnXGraphToggle(self, dictIndex, event):
        # if andExec(self.one_x_selected, event.widget['state'] == 'normal'):
            # for checkbox in self.column_x_options.values():
                # checkbox['state'] = 'disabled'
            # event.widget['state'] = 'normal'
        # else:
            # self.column_x_vars[dictIndex] = not self.column_x_vars[dictIndex]
            # self.one_x_selected = any(self.column_x_vars.values())

    def columnXGraphToggle(self, dictIndex):
        # Update the BooleanVar associated with the clicked Checkbutton
        self.column_x_vars[dictIndex].set(not self.column_x_vars[dictIndex].get())
        
        if orExec(not self.one_x_selected, self.column_x_vars[dictIndex].get()):
            self.one_x_selected = True
            for column, checkbox in self.column_x_options.items():
                if column != dictIndex:
                    checkbox['state'] = 'disabled'
        else:
            self.one_x_selected = False
            for checkbox in self.column_x_options.values():
                checkbox['state'] = 'normal'

    def columnYGraphToggle(self, dictIndex):
        self.column_y_vars[dictIndex] = not self.column_y_vars[dictIndex]
        
    def bind_3_btns(self, entity, function):
        for button_i in [1, 2, 3]:
            entity.bind(f'<Button-{button_i}>', function)
        
    def select_columns(self):
        self.columns = self.sheet.colTitles
        spacer2_init_height = self.spacer2['height']
        
        column_x_options_total_height = 0
        for column, i in zip(self.columns, range(2, len(self.columns)+2)):
            self.column_x_vars[column] = tk.BooleanVar(value=False)
            self.column_x_options[column] = tk.Checkbutton(self.frame, text=column, bg='#FFFFFF',
                                                           variable=self.column_x_vars[column],
                                                           command=lambda col=column: self.columnXGraphToggle(col))
            self.column_x_options[column].grid(column=4, row=i, sticky='nsew')
            column_x_options_total_height += 30
            
        # column_x_options_total_height = 0
        # for column, i in zip(self.columns, range(2, len(self.columns)+2)):
            # self.column_x_vars[column] = False
            # self.column_x_options[column] = tk.Checkbutton(self.frame, text=column, bg='#FFFFFF')
            # self.bind_3_btns(self.column_x_options[column], lambda event, column=column: self.columnXGraphToggle(column, event))
            # self.column_x_options[column].grid(column=4, row=i, sticky='nsew')
            # column_x_options_total_height += 30
            
        if column_x_options_total_height > spacer2_init_height:
            self.spacer1['height'] += column_x_options_total_height/8
            self.spacer2['height'] += column_x_options_total_height/2
        
        column_y_options_total_height = 0
        for column, i in zip(self.columns, range(2, len(self.columns)+2)):
            self.column_y_vars[column] = False
            self.column_y_options[column] = tk.Checkbutton(self.frame, text=column, bg='#FFFFFF')
            self.bind_3_btns(self.column_y_options[column], lambda event, column=column: self.columnYGraphToggle(column))
            self.column_y_options[column].grid(column=8, row=i, sticky='nsew')
            column_y_options_total_height += 30
        if column_y_options_total_height > spacer2_init_height:
            self.spacer1['height'] += column_y_options_total_height/8
            self.spacer2['height'] += column_y_options_total_height/2
        
    def plot_graph(self):
        # Clear columns from self.plotted_variables
        self.plotted_variables.clear()
        no_x_axis = not any( [value.get() for value in self.column_x_vars.values()] )
        no_y_axis = not any( self.column_y_vars.values() )
        
        if andExec(no_x_axis, no_y_axis):
            messagebox.showerror("Error", "Please select at least one X and one Y axis column.", parent=self.book.window)
            return
        
        if no_x_axis:
            messagebox.showerror("Error", "Please select at least one X axis column.", parent=self.book.window)
            return
        
        if no_y_axis:
            messagebox.showerror("Error", "Please select at least one Y axis column.", parent=self.book.window)
            return
        
        # Append columns to self.plotted_variables
        for column, var in self.column_x_vars.items():
            if var.get():
                self.plotted_variables.append(('x', column))
        
        for column, var in self.column_y_vars.items():
            if var:
                self.plotted_variables.append(('y', column))
        
        # Check whether columns exist within self.plotted_variables
        if self.plotted_variables:
            plt.clf()  # Clear previous plot
            
            plot_type = self.plot_type
            
            # Initialize lists to store x and y values
            x_values = []
            y_values = []
            legend_labels = []  # To store legend labels
            
            for axis, column in self.plotted_variables:
                self.sheet.df[column] = self.sheet.df[column].apply(lambda x: float(x) if is_numeric(x) else x)
                if axis == 'x':
                    x_values.append(self.sheet.df[column])
                    # legend_labels.append(column)
                elif axis == 'y':
                    y_values.append(self.sheet.df[column])
                    legend_labels.append(column)
            
            try:
                if plot_type == "scatter":
                    for x, y, label in zip(x_values, y_values, legend_labels):
                        plt.scatter(x, y, label=label)
                elif plot_type == "bar":
                    for x, y, label in zip(x_values, y_values, legend_labels):
                        plt.bar(x, y, label=label)
                    
            except UnboundLocalError:
                messagebox.showerror("Error", "Please select at least one X and one Y axis column.", parent=self.book.window)
                return
            
            plt.legend()
            plt.title("Graph")
            plt.xticks(rotation=45, ha='right')  # Rotate and align x-axis tick labels
            plt.tight_layout()  # Adjust layout to prevent clipping of tick labels
            plt.show()
            
        else:
            messagebox.showerror("Error", "Please select at least one X and one Y axis column.", parent=self.book.window)
    
    def unplot(self):
        for column in self.columns:
            self.column_x_vars[column] = False
            self.column_y_vars[column] = False

    def plot_graph_old(self):
        # Clear columns from self.plotted_variables
        self.plotted_variables.clear()
        
        # Append columns to self.plotted_variables
        for column, var in self.column_x_vars.items():
            if var.get():
                self.plotted_variables.append(('x', column))
        
        for column, var in self.column_y_vars.items():
            if var:
                self.plotted_variables.append(('y', column))
        
        # Check whether columns exist within self.plotted_variables
        if self.plotted_variables:
            plt.clf()       # Clear previous plot
            
            plot_type = self.plot_type
            
            sorted_indices = self.sheet.df.index
            for axis, column in self.plotted_variables:
                if axis == 'x':
                    try:
                        self.x_values = self.sheet.df[column].astype(float)
                    except TypeError as msg:
                        if 'Cannot cast DatetimeArray to dtype float64' in msg.args:
                            self.x_values = self.sheet.df[column]
                            # self.x_values = pd.to_datetime(self.sheet.df[column])
                        else:
                            raise
                elif axis == 'y':
                    try:
                        self.y_values = self.sheet.df[column].astype(float)
                    except TypeError as msg:
                        if 'Cannot cast DatetimeArray to dtype float64' in msg.args:
                            self.y_values = self.sheet.df[column]
                            # self.y_values = pd.to_datetime(self.sheet.df[column])
                        else:
                            raise
            
            try:
                if plot_type == "scatter":
                    plt.scatter(self.x_values, self.y_values)
                elif plot_type == "bar":
                    plt.bar(self.x_values, self.y_values)
                    
            except UnboundLocalError:
                messagebox.showerror("Error", "Please select at least one X and one Y axis column.", parent=self.book.window)
                return
            
            # for axis, column in self.plotted_variables:
                # if axis == 'x':
                    # # plt.bar(self.sheet.df[column], label=column)
                    # # plt.xlabel('Index')
                    # plt.plot(self.sheet.df[column], label=column)
                    # plt.xlabel(column)
                # elif axis == 'y':
                    # # plt.bar(self.sheet.df[column], label=column)
                    # plt.plot(self.sheet.df[column], label=column)
                    # plt.ylabel(column)
                    
            plt.legend()
            plt.title("Graph")
            plt.show()
            
        else:
            messagebox.showerror("Error", "Please select at least one X and one Y axis column.", parent=self.book.window)

class GraphPlotterA:
    def __init__(self, book, sheet):
        self.book = book
        self.sheet = sheet
        
        self.master = master = tk.Tk()
        master.title("Graph Plotter")
        master.geometry("750x420")
        master.protocol("WM_DELETE_WINDOW", self.close)
        
        # Centering the window on the screen
        window_width = master.winfo_reqwidth()
        window_height = master.winfo_reqheight()
        position_right = int(master.winfo_screenwidth()/2 - window_width*2)
        position_down = int(master.winfo_screenheight()/2 - window_height)
        master.geometry(f"+{position_right}+{position_down}")
        
        # Create a frame to hold the grid system
        self.frame = tk.Frame(self.master)
        self.frame.pack(expand=True, padx=20, pady=20)
        
        # Configure spacer1
        self.spacer1 = tk.Frame(self.frame, height=10)
        self.spacer1.grid(column=0, row=0, sticky='nsew')
        
        # Configure spacer2
        self.spacer2 = tk.Frame(self.frame, height=50)
        self.spacer2.grid(column=0, row=5, sticky='nsew', rowspan=100)
        
        self.plot_button = tk.Button(self.frame, text="Plot Graph", command=self.plot_graph)
        self.plot_button.grid(column=0, row=2, sticky='nsew')
        
        self.unplot_button = tk.Button(self.frame, text="Unplot", command=self.unplot)
        self.unplot_button.grid(column=0, row=3, sticky='nsew')
        
        sepBGWidth = 2
        sepNoBGWidth = 50
        
        # Separator and labels for X-axis column selection
        # ...

        # Separator and labels for Y-axis column selection
        # ...

        self.plotted_variables = []

    # Close method to close the window
    def close(self):
        self.master.withdraw()
        self.column_x_vars = {}
        self.column_x_options = {}
        self.column_y_vars = {}
        self.column_y_options = {}

    # Method to toggle X-axis columns
    def columnXGraphToggle(self, dictIndex):
        self.column_x_vars[dictIndex] = not self.column_x_vars[dictIndex]

    # Method to toggle Y-axis columns
    def columnYGraphToggle(self, dictIndex):
        self.column_y_vars[dictIndex] = not self.column_y_vars[dictIndex]

    # Method to select columns
    def select_columns(self):
        self.columns = self.sheet.colTitles
        spacer2_init_height = self.spacer2['height']
        
        column_x_options_total_height = 0
        for column, i in zip(self.columns, range(2, len(self.columns)+2)):
            self.column_x_vars[column] = False
            self.column_x_options[column] = tk.Checkbutton(self.frame, text=column)
            self.column_x_options[column].bind('<Button-1>', lambda event, column=column: self.columnXGraphToggle(column))
            self.column_x_options[column].bind('<Button-2>', lambda event, column=column: self.columnXGraphToggle(column))
            self.column_x_options[column].bind('<Button-3>', lambda event, column=column: self.columnXGraphToggle(column))
            self.column_x_options[column].grid(column=4, row=i, sticky='nsew')
            column_x_options_total_height += 30
        if column_x_options_total_height > spacer2_init_height:
            self.spacer1['height'] += column_x_options_total_height/8
            self.spacer2['height'] += column_x_options_total_height/2
        
        column_y_options_total_height = 0
        for column, i in zip(self.columns, range(2, len(self.columns)+2)):
            self.column_y_vars[column] = False
            self.column_y_options[column] = tk.Checkbutton(self.frame, text=column)
            self.column_y_options[column].bind('<Button-1>', lambda event, column=column: self.columnYGraphToggle(column))
            self.column_y_options[column].bind('<Button-2>', lambda event, column=column: self.columnYGraphToggle(column))
            self.column_y_options[column].bind('<Button-3>', lambda event, column=column: self.columnYGraphToggle(column))
            self.column_y_options[column].grid(column=8, row=i, sticky='nsew')
            column_y_options_total_height += 30
        if column_y_options_total_height > spacer2_init_height:
            self.spacer1['height'] += column_y_options_total_height/8
            self.spacer2['height'] += column_y_options_total_height/2
        
    # Method to plot scatter plot
    def plot_scatter(self):
        plt.clf()
        for column, _ in self.plotted_variables:
            plt.scatter(self.sheet.df.index, self.sheet.df[column], label=column)
        plt.legend()
        plt.title("Scatter Plot")
        plt.show()

    # Method to plot horizontal bar chart
    def plot_horizontal_bar(self):
        plt.clf()
        for column, _ in self.plotted_variables:
            plt.barh(self.sheet.df.index, self.sheet.df[column], label=column)
        plt.legend()
        plt.title("Horizontal Bar Chart")
        plt.show()

    # Method to plot histogram
    def plot_histogram(self):
        plt.clf()
        for column, _ in self.plotted_variables:
            plt.hist(self.sheet.df[column], bins=10, alpha=0.5, label=column)
        plt.legend()
        plt.title("Histogram")
        plt.show()

    # Method to plot pie chart
    def plot_pie_chart(self):
        plt.clf()
        for column, _ in self.plotted_variables:
            plt.pie(self.sheet.df[column], labels=self.sheet.df.index, autopct='%1.1f%%')
        plt.title("Pie Chart")
        plt.show()

    # Method to plot column (vertical bar) chart
    def plot_column_chart(self):
        plt.clf()
        for column, _ in self.plotted_variables:
            plt.bar(self.sheet.df.index, self.sheet.df[column], label=column)
        plt.legend()
        plt.title("Column Chart")
        plt.show()

    # Method to plot area chart
    def plot_area_chart(self):
        plt.clf()
        for column, _ in self.plotted_variables:
            plt.fill_between(self.sheet.df.index, self.sheet.df[column], label=column)
        plt.legend()
        plt.title("Area Chart")
        plt.show()

    # Method to plot stock chart
    def plot_stock_chart(self):
        plt.clf()
        # Implement stock chart plotting
        plt.title("Stock Chart")
        plt.show()

    # Method to plot box and whisker plot
    def plot_box_whisker(self):
        plt.clf()
        for column, _ in self.plotted_variables:
            plt.boxplot(self.sheet.df[column], labels=[column])
        plt.title("Box and Whisker Plot")
        plt.show()

    # Method to plot the graph based on user selection
    def plot_graph(self):
        # Clear columns from self.plotted_variables
        self.plotted_variables.clear()

        # Append columns to self.plotted_variables
        for column, var in self.column_x_vars.items():
            if var:
                self.plotted_variables.append(('x', column))

        for column, var in self.column_y_vars.items():
            if var:
                self.plotted_variables.append(('y', column))

        # Check whether columns exist within self.plotted_variables
        if self.plotted_variables:
            if len(self.plotted_variables) == 1:
                axis, column = self.plotted_variables[0]
                if axis == 'x':
                    self.plot_horizontal_bar()
                elif axis == 'y':
                    self.plot_scatter()
            elif len(self.plotted_variables) == 2:
                axis_types = [axis for axis, _ in self.plotted_variables]
                if andExec('x' in axis_types, 'y' in axis_types):
                    self.plot_scatter()
                elif 'x' in axis_types:
                    self.plot_horizontal_bar()
                elif 'y' in axis_types:
                    self.plot_column_chart()
        else:
            messagebox.showerror("Error", "Please select at least one X/Y axis column.", parent=self.book.window)

    # Method to unplot all columns
    def unplot(self):
        for column in self.columns:
            self.column_x_vars[column].set(False)
            self.column_y_vars[column].set(False)

class SheetInner(tk.Frame):
    def __init__(self, master, *args, **kwargs):
        tk.Frame.__init__(self, *args, **kwargs)
        self.master = master
        
    def __repr__(self):
        return f'SheetInner for {self.master}'
        
    # def __str__(self):
        # return self.__repr__()
        
class UndoRedoSystem:
    def __init__(self, book, sheet, undo_menu, redo_menu):
        self.undoActions = []       # List to store (index, undoActions) tuples
        self.undoSnapshots = []     # List to store (index, undoSnapshot) tuples for pandas DataFrame Cell snapshots
        self.undoColumns = []     # List to store (index, undoColumnSnapshot) tuples for pandas DataFrame Column snapshots
        self.undoQStats = []      # List to store (index, undoQStatsAttr: {column:current_cont, previous_cont for every column in sheet}) tuples
        
        self.redoActions = []       # List to store (index, redoActions) tuples
        self.redoSnapshots = []     # List to store (index, redoSnapshot) tuples for pandas DataFrame Cell snapshots
        self.redoColumns = []     # List to store (index, redoColumnSnapshot) tuples for pandas DataFrame Column snapshots
        self.redoQStats = []      # List to store (index, redoQStatsAttr: {column:current_cont, previous_cont for every column in sheet}) tuples

        self.index_counter = 0  # Counter for unique indexing of actions
        self.editTrackerPhase = 1  # Counter for unique indexing of undo/redo action location
        self.maxActions = 120  # Maximum number of actions allowed
        
        self.book = book
        self.sheet = sheet
        
        self.sheet.undo_btn['command'] = self.undoActionCollect
        self.sheet.redo_btn['command'] = self.redoActionCollect
        
        self.undo_menu = undo_menu
        self.redo_menu = redo_menu
        self.currentSnapshotDF, self.currentColsDF = self.sheet.generateSnapshotDF()
        self.currentQStats = self.sheet.getStatsReport(fancyReportType=True, statsReportVisible=False)
        
        # self.currentQStats = {column.index:(column.current_cont, column.previous_cont) for column in self.sheet}

        self.update_undo_menu()
        self.update_redo_menu()

    def actionCollect(self, action):
        currentSnapshotDF, currentColsDF = self.sheet.generateSnapshotDF()
        currentQStats = self.sheet.getStatsReport(fancyReportType=True, statsReportVisible=False)
        
        # Perform the action
        # Check if the number of actions exceeds the maximum limit
        if len(self.undoActions) >= self.maxActions:
            # Remove the oldest action from undoActions
            self.undoActions.pop(0)
            self.undoSnapshots.pop(0)
            self.undoColumns.pop(0)
            self.undoQStats.pop(0)

            # Decrement indices in all Undo Lists
            self.undoActions = [(idx - 1, act) for idx, act in self.undoActions]
            self.undoSnapshots = [(idx - 1, act) for idx, act in self.undoSnapshots]
            self.undoColumns = [(idx - 1, act) for idx, act in self.undoColumns]
            self.undoQStats = [(idx - 1, act) for idx, act in self.undoQStats]
            
            # Decrement indices in all Redo Lists
            self.redoActions = [(idx - 1, act) for idx, act in self.redoActions]
            self.redoSnapshots = [(idx - 1, act) for idx, act in self.redoSnapshots]
            self.redoColumns = [(idx - 1, act) for idx, act in self.redoColumns]
            self.redoQStats = [(idx - 1, act) for idx, act in self.redoQStats]
            
        else:
            # Increment index_counter only if adding a new action
            self.index_counter += 1
            self.editTrackerPhase += 1
            
        self.undoActions.append((self.index_counter, action))
        self.undoSnapshots.append((self.index_counter, self.currentSnapshotDF.copy() ))
        self.undoColumns.append((self.index_counter, self.currentColsDF.copy() ))
        self.undoQStats.append((self.index_counter, self.currentQStats.copy() ))
        
        self.currentSnapshotDF = currentSnapshotDF
        self.currentColsDF = currentColsDF
        self.currentQStats = currentQStats

        self.update_undo_menu()

        # Clear redo actions whenever a new action is performed
        self.redoActions = []
        self.update_redo_menu()

    def undoActionCollect(self, calledDirectly=True):
        if self.undoActions:
            # Pop the last action from undoActions
            index, undoneAction = self.undoActions.pop()
            index, undoneSnapshot = self.undoSnapshots.pop()
            index, undoneColSnapshot = self.undoColumns.pop()
            index, undoneQStats = self.undoQStats.pop()

            self.editTrackerPhase -= 1
            
            # Move the undone action and snapshot to redoActions and dataSnapshots
            self.redoActions.append((index, undoneAction))
            self.redoSnapshots.append((index, self.currentSnapshotDF.copy() ))
            self.redoColumns.append((index, self.currentColsDF.copy() ))
            self.redoQStats.append((index, self.currentQStats.copy() ))

            self.update_undo_menu()
            self.update_redo_menu()
            
            self.currentSnapshotDF = undoneSnapshot
            self.currentColsDF = undoneColSnapshot
            self.currentQStats = undoneQStats
            
            self.sheet.reconstructFromSnapshotDF(undoneSnapshot, undoneColSnapshot, self.editTrackerPhase)
            
            if calledDirectly:
                self.sheet.resetQStats(QStatsData=undoneQStats)
            
            return undoneSnapshot, undoneQStats, undoneColSnapshot

    def redoActionCollect(self, calledDirectly=True):
        if self.redoActions:
            # Pop the last action from redoActions
            index, redoneAction = self.redoActions.pop()
            index, redoneSnapshot = self.redoSnapshots.pop()
            index, redoneColSnapshot = self.redoColumns.pop()
            index, redoneQStats = self.redoQStats.pop()

            self.editTrackerPhase += 1

            # Move the redone action and snapshot to undoActions and dataSnapshots
            self.undoActions.append((index, redoneAction))
            self.undoSnapshots.append((index, self.currentSnapshotDF.copy() ))
            self.undoColumns.append((index, self.currentColsDF.copy() ))
            self.undoQStats.append((index, self.currentQStats.copy() ))

            self.update_undo_menu()
            self.update_redo_menu()
            
            self.currentSnapshotDF = redoneSnapshot
            self.currentColsDF = redoneColSnapshot
            self.currentQStats = redoneQStats
            
            self.sheet.reconstructFromSnapshotDF(redoneSnapshot, redoneColSnapshot, self.editTrackerPhase)
            
            if calledDirectly:
                self.sheet.resetQStats(QStatsData=redoneQStats)
            
            return redoneSnapshot, redoneQStats, redoneColSnapshot

    def update_undo_menu(self):
        self.undo_menu.delete(0, tk.END)  # Clear current menu items
        for index, action in reversed(self.undoActions):
            self.undo_menu.add_command(label=action, command=lambda idx=index: self.undo_up_to(idx))
        # self.sheet.undo_btn['state'], self.sheet.undo_menu_btn['state'] = ('normal', 'normal') if self.undoActions else ('disabled', 'disabled')
        self.sheet.undo_btn['state'] = 'normal' if self.undoActions else 'disabled'
        self.sheet.undo_menu_btn['state'] = 'normal' if self.undoActions else 'disabled'

    def update_redo_menu(self):
        self.redo_menu.delete(0, tk.END)  # Clear current menu items
        for index, action in reversed(self.redoActions):
            self.redo_menu.add_command(label=action, command=lambda idx=index: self.redo_up_to(idx))
        # self.sheet.redo_btn['state'], self.sheet.redo_menu_btn['state'] = ('normal', 'normal') if self.redoActions else ('disabled', 'disabled')
        self.sheet.redo_btn['state'] = 'normal' if self.redoActions else 'disabled'
        self.sheet.redo_menu_btn['state'] = 'normal' if self.redoActions else 'disabled'
                
    def undo_up_to(self, index):
        if self.undoActions:
            # Find the index of the specified action in undoActions
            try:
                idx = next(idx for idx, (i, _) in enumerate(self.undoActions) if i == index)
            except StopIteration:
                return
            
            # Call undoActionCollect() repeatedly until reaching the specified index
            while len(self.undoActions) > idx:
                undoneSnapshot, undoneQStats, undoneColSnapshot = self.undoActionCollect(calledDirectly=False)
                
            # self.editTrackerPhase = undone_index
            self.update_undo_menu()
            self.update_redo_menu()
            self.currentSnapshotDF = undoneSnapshot
            self.sheet.reconstructFromSnapshotDF(undoneSnapshot, undoneColSnapshot, self.editTrackerPhase)
            self.sheet.resetQStats(QStatsData=undoneQStats)
                
    def redo_up_to(self, index):
        if self.redoActions:
            # Find the index of the specified action in redoActions
            try:
                idx = next(idx for idx, (i, _) in enumerate(self.redoActions) if i == index)
            except StopIteration:
                return
            
            # Call redoActionCollect() repeatedly until reaching the specified index
            while len(self.redoActions) > idx:
                redoneSnapshot, redoneQStats, redoneColSnapshot = self.redoActionCollect(calledDirectly=False)
                
            # self.editTrackerPhase = redone_index
            self.update_undo_menu()
            self.update_redo_menu()
            self.currentSnapshotDF = redoneSnapshot
            self.sheet.reconstructFromSnapshotDF(redoneSnapshot, redoneColSnapshot, self.editTrackerPhase)
            self.sheet.resetQStats(QStatsData=redoneQStats)

class Sheet(tk.Text):
    init = True
    ready = False
    entry = None
    QStats = None
    QStatsData = None
    inCutCondition = False
    cellFormulaEdit = False
    columnFormulaEdit = None
    QStatsType = ['Sum', 'Count', 'Average', 'Maximum', 'Minimum', 'Sample Standard Deviation', 'Sample Variance', 'Population Standard Deviation', 'Population Variance']
    font = ('Arial', 12)
    colIndex, rowIndex = 0, 0
    
    cut_triggered = False
    currentCell = lastCell = None
    previous_range = None
    horizontalKeyboardMoveLeft = False
    horizontalKeyboardMoveRight = True
    verticalKeyboardMoveUp = False
    verticalKeyboardMoveDown = True
    modifier_is_pressed = [False]*4
    control_is_pressed = False
    alt_is_pressed = False
    shift_is_pressed = False
    win_is_pressed = False
    app_is_pressed = False
    dragFirstTouch = True
        
    def __init__(self, parent, book, imported, have_headers=True, sheetName=None, structData=None, df=None, columnSingleWidth=100, nCols=1, nRows=5, width=50, BossSheet=None, sheetOpenPyxl=None, *args, **kwargs):
        super().__init__(parent, cursor='arrow', state='disabled', wrap=tk.NONE, background='#FFFFFF', *args, **kwargs)
        self.bind('<1>', lambda event: self.focus_set() )
        
        self.sheetName = sheetName if sheetName is not None else 'NewSheet1'
        self.BossSheet = BossSheet if BossSheet is not None else self
        
        self.book = book
        self.sheet = self
        
        self.book.sheetsDict[sheetName] = self
        
        self.parent = parent
        self.window = book.window
        self.scrollBar = BothScrollBar(h_parent=self.book.sheetOuter0, v_parent=self.book.sheetOuter1, target=self)
        
        # self.structData = structData = self.book.structData[sheetName] = structData if extension == '.qutable' else {'have_headers': True, 'columnSingleWidth': None, 'nCols': None, 'nRows': None, 'width': None, 'columns': None, 'colTitleCells': None, 'cells': None}
         
        extension = self.book.extension
        
        if andExec(extension == '.qutable', imported):
            self.structData = structData = self.book.structData[sheetName]
            self.columnSingleWidth = columnSingleWidth = structData['columnSingleWidth']
            self.nCols = nCols = structData['nCols']
            self.nRows = nRows = structData['nRows']
            self.width = width = structData['width']
            self.have_headers = structData['have_headers']
            self.restructureStructData()
            
        else:
            self.structData = structData = self.book.structData[sheetName] = {'have_headers': True, 'columnSingleWidth': None, 'nCols': None, 'nRows': None, 'width': None, 'columns': None, 'colTitleCells': None, 'cells': None}
            self.columnSingleWidth = columnSingleWidth
            # print(f'nRows = {nRows}')
            # print(f'have_headers = {have_headers}')
            nRows += not have_headers
            # print(f'nRows = {nRows}')
            self.nCols = nCols
            self.nRows = nRows
            self.width = width
            self.have_headers = have_headers
        
        # if not sheetOpenPyxl:
        if extension.startswith('.xls'):
            if not self.have_headers:
                self.insert_row_and_translate_formulae(sheetOpenPyxl)
                # sheetOpenPyxl.insert_rows(1) #, amount= int(not have_headers) )
        else:
            sheetOpenPyxl = book.bookOpenPyxl.create_sheet(sheetName)
        self.sheetOpenPyxl = sheetOpenPyxl
        
        # self.editDF = []
        # self.editActions = []
        # self._editTrackerPhase = 1
        
        self.graphPlotter = GraphPlotter(self.book, self.sheet)
        self.graphPlotter.master.withdraw()
        
        self.sheetNum = len(book)
        self.prevSelectedCells = set()
        
        self.rightClickMenu = rightClickMenu = RightClickMenu(widget=self, book=book, purpose=self.book.sheetOuter1, parent=self)
        rightClickMenu.add_command(label="Cut", command=book.cut )
        rightClickMenu.add_command(label="Copy", command=book.copy )
        rightClickMenu.add_command(label="Paste", command=book.paste )
        rightClickMenu.add_command(label="Clear", command=book.clear )
        rightClickMenu.add_separator()
        rightClickMenu.add_command(label="Column Formula", command=book.columnFormulaStartAction )
        rightClickMenu.add_separator()
        rightClickMenu.add_command(label="Insert Column Left", command=book.insertColLeft )
        rightClickMenu.add_command(label="Insert Column Right", command=book.insertColRight )
        rightClickMenu.add_command(label="Insert Row Above", command=book.insertRowAbove )
        rightClickMenu.add_command(label="Insert Row Bottom", command=book.insertRowBelow )
        rightClickMenu.add_separator()
        rightClickMenu.add_command(label="Delete Column", command=book.deleteColumn )
        rightClickMenu.add_command(label="Delete Row", command=book.deleteRow )
        
        # self.book.createSelectorFormulaEntry()
        
        self.sheetInner = tk.Frame(master=self, background='#FFFFFF', width=1600)
        self.sheetInner.configure(cursor='plus')
            
        self.window_create('end', window=self.sheetInner, stretch=1)
        self.xview_tuple = self.xview()
        self.yview_tuple = self.yview()
        
        self.imported = imported
        
        self.columnsTotalWidth = width
        self.columnsTotalWidthActual = width

        debug = sheetOpenPyxl.title == 'Winter 2024' if book.experimental else 0
        
        self.sheet.edit_status = False
        
        self.defaultTitle = defaultTitle = '[Insert Column Name Here]'
        if orExec(str(df) == str(pd.DataFrame()), not isinstance(df, pd.DataFrame)):
            self.df = df = pd.DataFrame({f'C{i}':['']*self.nRows for i in range(1, self.nCols+1)}).astype(str)
            # print('executed 1')
        elif str(df).startswith('Empty DataFrame'):
            self.df = df = pd.DataFrame({colTitle:['']*self.nRows for colTitle in df.columns}).astype(str)
            # print('executed 2')
        elif andExec(df is not None, extension != '.qutable'):
            self.df = df = df.astype(str)
            # print('executed 3')
        elif df is None:
            self.df = df = pd.DataFrame({defaultTitle:['']*self.nRows}).astype(str)
            # print('executed 4')
        
        if andExec(imported, df is not None, extension != '.qutable'):
            self.nCols = nCols = self.df.shape[1]
            self.nRows = nRows = self.df.shape[0] + (not have_headers)
            # print('really not qutable at all')
        
        if extension == '.qutable':
            self.colTitles = []
            colIndex = 0
            while colIndex < self.nCols:
                # print(structData['columns'])
                try:
                    colDict = structData['columns'][colIndex]
                except KeyError:
                    colDict = structData['columns'][str(colIndex)]
                # print(colDict)
                self.colTitles.append( colDict['title'] if 'title' in colDict else defaultTitle )
                colIndex += 1
            df_colTitles = self.colTitles
        else:
            structData['columns'] = {}
            structData['colTitleCells'] = {}
            structData['cells'] = {}
            
            # structData['columns'] = [{} for i1 in range(nCols)]
            # structData['colTitleCells'] = [{} for i1 in range(nCols)]
            # structData['cells'] = [{} for i1 in range(nCols) for i2 in range(nRows)]
            
            if andExec(have_headers, imported):
                self.colTitles = df_colTitles = list(self.df)
            else:
                self.colTitles = [f'C{i}' for i in range(self.nCols)]
                df_colTitles = list(self.df)
        
        removeBinds(self, 'Text')
        self.bind('<MouseWheel>', self.mouseWheel )
        
        structData_original = str(structData)
        print(6710)
        # print(6694, structData_original)
        self.enhanceSheetInner(nCols, nRows, debug=debug)
        
        self.structData = structData = eval(structData_original)
        print(6715)
        # print(6696, structData_original)
        self.fill_in_data(nCols, nRows, df_colTitles, df, structData, debug=debug)
        self.undo_redo_widgets()
        
        self.initialSnapshotDF, self.initialColumnDF = self.generateSnapshotDF()
        
        # self.editDF = [self.initialSnapshotDF]
        # self.editActions = ['']
        
        self.edit_status = True
        self.init = False
        self.ready = True
        
    def enhanceSheetInner(self, nCols, nRows, debug=0):
        self.nCols = nCols
        self.nRows = nRows
        
        book = self.book
        extension = book.extension
        
        self.columns = CellSet(self.book, self.sheet)
        self.cells = CellSet(self.book, self.sheet)
        self.colTitleCells = CellSet(self.book, self.sheet)
        self.cellMap = CellSet(self.book, self.sheet)
        self.shadowBorders = CellSet(self.book, self.sheet)
        
        self.colSelectors = CellSet(self.book, self.sheet)
        self.rowSelectors = CellSet(self.book, self.sheet)
        self.colShadowBorders = CellSet(self.book, self.sheet)
        self.rowShadowBorders = CellSet(self.book, self.sheet)
        self.colBorders = CellSet(self.book, self.sheet)
        self.rowBorders = CellSet(self.book, self.sheet)
        
        self.selectedCellsSet = CellSet(self.book, self.sheet, label_status=True, name='selectedCellsSet')
        self.selectedColsIndex = []
        self.selectedRowsIndex = []
        
        self.rowDimList = []
        self.colDimList = []
        
        self.rowDimRawList = rowDimRawList = list(self.sheetOpenPyxl.row_dimensions.values()) if extension.startswith('.xls') else []
        self.colDimRawList = colDimRawList = list(self.sheetOpenPyxl.column_dimensions.values()) if extension.startswith('.xls') else []
        
        self.indexColumn = Column(parent=self.sheetInner, book=book, colIndex=-1, width=self.width, rowDimRawList=rowDimRawList, nRows=nRows, font=self.font, cellType=IndexCell, sheet=self, QStatsType=None)
        
        if not self.imported:
            self.colTitles = [self.defaultTitle]*nCols
        
        for colIndex, colTitle in zip(range(nCols), self.colTitles):
            # Conversion factor using the pixel-to-point conversion rate
            pixels_per_point = 10.5
            try:
                column_width_points = colDimRawList[colIndex].width
                column_width_pixels = column_width_points * pixels_per_point
            except (IndexError, UnboundLocalError):
                column_width_pixels = self.columnSingleWidth
            
            if self.sheetName == 'Spring 2024':
                print(f'colIndex = {colIndex}')
            
            column = Column(self.sheetInner, colIndex=colIndex, book=book, width=column_width_pixels, rowDimRawList=rowDimRawList, colTitle=colTitle, nRows=nRows, font=self.font, sheet=self)
            
            self.columnsTotalWidth += column_width_pixels
            if not self.colDimList:
                self.colDimList.append(column_width_pixels)
            
            self.columns += [column]
            self.colTitleCells.add(column.colTitleCell, init=True)
            
            self.cells += column.cells
            self.cellMap.add(column.cells)
            
        self.placeholderColumn = placeholderColumn = Column(self.sheetInner, colIndex=colIndex+1, book=book, width=500, rowDimRawList=rowDimRawList, nRows=nRows, font=self.font, sheet=self, colTitle='\t Double Click to Add New Column', is_placeholder=True)
        placeholderColumn.colTitleCell.fg = '#BBBBBB'
        placeholderColumn.colTitleCell.font = 'Arial 12 italic'
        placeholderColumn.colTitleCell.anchorVar = 'w'
        self.columnsTotalWidthActual += placeholderColumn.width
        
        placeholderColumn.colTitleCell.unbind("<Double-1>")
        placeholderColumn.colTitleCell.bind("<Double-1>", placeholderColumn.sheet.insertColRightMost)
        
        self.addRowGuide = tk.Label(self.sheetInner, text='Type something in the bottommost row to automatically insert new row.\n\n\n', fg='#BBBBBB', bg='#FFFFFF', font='Arial 12 italic')
        colSize, rowIndex = self.sheetInner.grid_size()
        self.addRowGuide.grid(row=rowIndex, column=0, columnspan=colSize)
        self.addRowGuide.bind('<1>', lambda event: self.focus_set() )
        
        currentCellColIndex, currentCellRowIndex = 0, 0
        self.columns[currentCellColIndex].cells[currentCellRowIndex].focusOnCell()
        # self.SelectionMarkByCoord(0, 0)
        
        if self.currentCell.generateRange() not in self.selectedCellsSet.cellSet:
            self.selectedCellsSet.start( self.currentCell )
        
    def undo_redo_widgets(self):
        book = self.book
        superMenu = book.SuperMenuWidget
        self.undo_btn = superMenu.undo_btns[self.sheetName] = SuperButton(parent=superMenu.SuperMenuFrame, text='Undo', image_path='', image=superMenu.undo_btn_img, book=book, font='Arial 14')
        self.undo_btn.grid(row=superMenu.undo_btn_row, column=superMenu.undo_btn_column)
        
        self.undo_menu_btn = superMenu.undo_menu_btns[self.sheetName] = SuperMenuButton(parent=superMenu.SuperMenuFrame, icon='▼', text='Undo List', book=book, state='disabled')
        self.undo_menu_btn.grid(row=superMenu.undo_menu_btn_row, column=superMenu.undo_menu_btn_column)
        
        self.redo_btn = superMenu.redo_btns[self.sheetName] = SuperButton(parent=superMenu.SuperMenuFrame, text='Redo', image_path='', image=superMenu.redo_btn_img, book=book, font='Arial 14')
        self.redo_btn.grid(row=superMenu.redo_btn_row, column=superMenu.redo_btn_column)
        
        self.redo_menu_btn = superMenu.redo_menu_btns[self.sheetName] = SuperMenuButton(parent=superMenu.SuperMenuFrame, icon='▼', text='Redo List', book=book, state='disabled')
        self.redo_menu_btn.grid(row=superMenu.redo_menu_btn_row, column=superMenu.redo_menu_btn_column)
        
        self.undo_redo_system = UndoRedoSystem(self.book, self.sheet, self.undo_menu_btn.menu, self.redo_menu_btn.menu)
        
        self.actionCollect = self.undo_redo_system.actionCollect
        
    def fill_in_data(self, nCols, nRows, df_colTitles, df, structData, debug=0):
        ''' Filling in data (if any) '''
        book = self.book
        extension = book.extension
        
        # structData = self.structData
        have_headers = self.have_headers
        
        # print(6814, f'structData = {structData}')
        
        if extension != '.qutable':
            for cell in self.cells:
                # t = threading.Thread(target=CellLabel.autoInit, args=(cell,))
                # t.start()
                cell.autoInit()
            for column in self.columns:
                # t = threading.Thread(target=ColumnTitle.autoInit, args=(column.colTitleCell,))
                # t.start()
                column.colTitleCell.autoInit()
        
        for colIndex, column, colTitle in zip(range(nCols), self.columns, df_colTitles):
            # Changing for header for Table (if any)
            # print(6832)
            if have_headers:
                # print(6834)
                cell = column.colTitleCell if have_headers else self[colIndex, 0]
                # print(6836)
                if extension == '.qutable':
                    data = structData['colTitleCells'][str(colIndex)] if have_headers else structData['cells'][cell.cellFlatIndex]
                    cell.autoFormat(data)
                    cellResult = self.autoFormat(data, cell, int(have_headers), debug=debug)
                else:
                    # print(6842)
                    cellOpenPyxlIndex = have_headers, colIndex+have_headers
                    cellResult = self.autoFormat(cellOpenPyxlIndex, cell, int(have_headers), debug=debug)
            
            # print(6846)
            row_increment = 2 if have_headers else 1
            columnCells = column.cells
            columnValues = df.iloc[:, colIndex]
            if not have_headers:
                columnValues = [colTitle] + list(columnValues)
            for cell, cellValue in zip(columnCells, columnValues):
                rowIndex = cell.index[1]
                if extension.startswith('.xls'):
                    cellOpenPyxlIndex = rowIndex+row_increment, colIndex+1
                    cellValueHyphenCheck = cellValue.strip() == '-' if isinstance(cellValue, str) else False
                    self.autoFormat(cellOpenPyxlIndex, cell, row_increment, debug=debug)
                    # t = threading.Thread(target=self.autoFormat, args=(cellOpenPyxlIndex, cell, row_increment, debug) )
                    # t.start()
                elif extension == '.csv':
                    cell.formula = cellValue
                elif extension == '.qutable':
                    data = structData['cells'][cell.cellFlatIndex if isinstance(structData['cells'], list) else str(cell.index)]
                    # print(6842, str(cell.index), data)
                    data_formula = data['formula']
                    cell.formula = data_formula if data_formula else data['content']
                    cell.autoFormat(data)
            if extension == '.qutable':
                data = structData['columns'][str(colIndex)]
                column.autoFormat(data)
                
            self.columnsTotalWidthActual += column.width
        
        if extension == '.qutable':
            self.df = pd.DataFrame({colTitle: column.cellValues for colTitle, column in zip(self.colTitles, self.columns)})
            if isinstance(structData['cells'], list):
                structData['cells'] = structData['cells'][:nCols*nRows]
            if isinstance(structData['colTitleCells'], list):
                structData['colTitleCells'] = structData['colTitleCells'][:nCols]
            if isinstance(structData['columns'], list):
                structData['columns'] = structData['columns'][:nCols]
        
        self.df = convert_dates_to_datetime(self.df)
        
    def getStatsReport(self, fancyReportType=False, statsReportVisible=True):
        StatsDict = {'': self.QStatsType}
        
        for column in self.columns:
            StatsDictCol = column.getStatsReport(statsReportVisible=statsReportVisible)
            StatsDict.update(StatsDictCol)
            
        if fancyReportType:
            StatsDict = pd.DataFrame(StatsDict)
            
        self.QStatsData = StatsDict
        
        return StatsDict
        
    def createQStats(self, QStatsData=None):
        self.QStats = QStats = ttk.Treeview(self.book.QStatsSheet, columns=list(range(self.nCols+1)), show='headings')
        HScrollBar(QStats)
        QStats.column(0, anchor='center')
        QStats.heading(0, text='')
        for column, colIndex in zip( self , range(1, self.nCols+1) ):
            QStats.column(colIndex, anchor='center')
            QStats.heading(colIndex, text=column)
        self.book.QStatses.append(QStats)
        QStats.place(relwidth=1.0, relheight=1.0)
        
        if type(QStatsData) == type(None):
            QStatsData = self.getStatsReport(fancyReportType=True)
        
        for StatsIndex in range(len(self.QStatsType)):
            stats = list(QStatsData.loc[StatsIndex])
            QStats.insert('', tk.END, values=stats)
        
    def resetQStats(self, QStatsData=None):
        quantum_mode = self.book.quantum_mode
        if quantum_mode:
            self.book.quantum_mode = False
        try:
            self.QStats.destroy()
            self.book.QStatses.remove( self.QStats )
            self.createQStats( QStatsData )
        except Exception as e:
            with open('stderr.log', 'w') as fileIO:
                fileIO.write( format_exc() )
        if quantum_mode:
            self.book.quantum_mode = True
    
    def addStructData(self):
        for cell in self.cells:
            # if cell.structData not in self.structData['cells'].values():
            self.structData['cells'][str(cell.index)] = cell.structData
        
    def restructureStructData(self):
        if isinstance(self.structData['cells'], list):
            self.structData_cells = self.structData['cells']
            self.structData['cells'] = {}
            for colIndex in range(self.nCols):
                for rowIndex in range(self.nRows):
                    self.structData['cells'][str((colIndex, rowIndex))] = self.structData_cells[colIndex*self.nRows+rowIndex]
        if isinstance(self.structData['colTitleCells'], list):
            self.structData_colTitleCells = self.structData['colTitleCells']
            self.structData['colTitleCells'] = {}
            for colIndex in range(self.nCols):
                self.structData['colTitleCells'][str(colIndex)] = self.structData_colTitleCells[colIndex]
        if isinstance(self.structData['columns'], list):
            self.structData_columns = self.structData['columns']
            self.structData['columns'] = {}
            for colIndex in range(self.nCols):
                self.structData['columns'][str(colIndex)] = self.structData_columns[colIndex]
        
    # def restructureStructData(self):
        # if isinstance(self.structData['cells'], list):
            # for cell in self.cells:
                # self.structData['cells'][str(cell.index)] = cell.structData
        # if isinstance(self.structData['colTitleCells'], list):
            # for colTitleCell in self.colTitleCells:
                # self.structData['colTitleCells'][str(colTitleCell.index)] = colTitleCell.structData
        # if isinstance(self.structData['columns'], list):
            # for column in self.columns:
                # self.structData['columns'][column.index] = column.structData
        
    # def deleteStructData(self):
        # for cell in self.cells:
            # del self.structData['cells'][str(cell.index)]
        
    def reformSheetInner(self, nCols, nRows):
        nColsDiff = nCols - self.nCols
        nRowsDiff = nRows - self.nRows
        rightMostCol, bottomMostRow = None, None
        if nColsDiff > 0:
            rightMostCol = self.insertColRightMost(nColsSelected=nColsDiff)
        elif nColsDiff == 0:
            pass
        else:
            self.deleteColRightMost(nColsSelected=-nColsDiff)
        if nRowsDiff > 0:
            bottomMostRow = self.insertRowBottomMost(nRowsSelected=nRowsDiff)
        elif nRowsDiff == 0:
            pass
        else:
            self.deleteRowBottomMost(nRowsSelected=-nRowsDiff)
        return rightMostCol, bottomMostRow
            
    def reExecFormula(self):
        for cell in self.cells:
            cell.parse_formula_determination()
        
    def generateCellMaps(self):
        return self.cellMap
        
    def generateSnapshotDF(self):
        rowIndexerforCells = ['bg', 'fg', 'anchor', 'justify', 'link', 'top_border', 'bottom_border', 'left_border', 'right_border', 'height', 'content', 'formula', 'font']
        colTitleSnapshotDF = {colTitleCell.index:colTitleCell.structData.copy().values() for colTitleCell in self.colTitleCells}
        cellSnapshotDF = {cell.index:cell.structData.copy().values() for cell in self.cells}
        combinedSnapshotDFRaw = {**colTitleSnapshotDF, **cellSnapshotDF}
        combinedSnapshotDF = pd.DataFrame(combinedSnapshotDFRaw, index=rowIndexerforCells)
        
        rowIndexerforCols = ['sumStatus', 'sumNum', 'previous_cont', 'current_cont', 'is_automated', 'width', 'height', 'title']
        colSnapshotDFRaw = {col.index:col.structData.copy().values() for col in self}
        colSnapshotDF = pd.DataFrame(colSnapshotDFRaw, index=rowIndexerforCols)
        
        combinedSnapshotDF = pd.DataFrame(self.structData['cells'])
        colSnapshotDF = pd.DataFrame(self.structData['columns'])
        
        return combinedSnapshotDF, colSnapshotDF
        
    def generateSnapshotDF(self):
        colIndexerforColTitleCells = [colTitleCell.index for colTitleCell in self.colTitleCells]
        colTitleSnapshotDF = pd.DataFrame(self.structData['colTitleCells'])
        colTitleSnapshotDF = colTitleSnapshotDF.transpose()
        colTitleSnapshotDF.index = colIndexerforColTitleCells
        colTitleSnapshotDF = colTitleSnapshotDF.transpose()
        
        colIndexerforCells = [cell.index for cell in self.cells]
        cellSnapshotDF = pd.DataFrame(self.structData['cells'])
        cellSnapshotDF = cellSnapshotDF.transpose()
        cellSnapshotDF.index = colIndexerforCells
        cellSnapshotDF = cellSnapshotDF.transpose()
        
        combinedSnapshotDF = pd.concat([colTitleSnapshotDF, cellSnapshotDF], axis=1)
        colSnapshotDF = pd.DataFrame(self.structData['columns'])
        return combinedSnapshotDF, colSnapshotDF
        
    def reconstructFromSnapshotDF(self, snapshotDF, columnDF, editTrackerPhase):
        ''' Undo/Redo Backbone '''
        if editTrackerPhase <= 1:
            snapshotDF = self.initialSnapshotDF
            columnDF = self.initialColumnDF
        # print(6957, snapshotDF)
        nCols, nRows = max(list(snapshotDF))
        nCols, nRows = nCols+1, nRows+1
        self.reformSheetInner(nCols, nRows)
        for cellIndex in snapshotDF:
            cell = self[cellIndex]
            cell.autoFormat( dict(snapshotDF[cellIndex]) )
        for colIndex in columnDF:
            col = self[int(colIndex)]
            col.autoFormat( dict(columnDF[colIndex]) )
        self.book.formulaEntry.delete(0, 'end')
        self.book.formulaEntry.insert(0, self.currentCell.formula)
        
    # def collectSnapshotDF(self, action):
        # # snapshotDF = self.generateSnapshotDF()
        # # initCond = not self.editDF[self.editTrackerPhase-1].equals(snapshotDF) if self.editDF else 1
        # # if andExec(initCond, action):
        # if action:
            # self.undo_redo_system.actionCollect(action)
    
    @property
    def editTrackerPhase(self):
        return self.undo_redo_system.editTrackerPhase
        # return self._editTrackerPhase
    
    @editTrackerPhase.setter
    def editTrackerPhase(self, new_editTrackerPhase):
        self.undo_redo_system.editTrackerPhase = new_editTrackerPhase
        # self._editTrackerPhase = new_editTrackerPhase
    
    @property
    def undoActions(self):
        return self.undo_redo_system.undoActions
    
    @property
    def redoActions(self):
        return self.undo_redo_system.redoActions
    
    @property
    def undoSnapshotDF(self):
        return self.undo_redo_system.undoSnapshots
    
    @undoSnapshotDF.setter
    def undoSnapshotDF(self, new_undoSnapshots):
        self.undo_undo_system.undoSnapshots = undoSnapshots
    
    @property
    def redoSnapshotDF(self):
        return self.undo_redo_system.redoSnapshots
    
    @redoSnapshotDF.setter
    def redoSnapshotDF(self, new_redoSnapshots):
        self.undo_redo_system.redoSnapshots = redoSnapshots
    
    def autoFormat(self, input_data, cell, row_increment, debug=0):
        if not self.imported:
            return ''
            
        if not self.book.extension.startswith('.xls'):
            return ''
        
        try:
            colIndex, rowIndex = cell.index
            rowIndex += row_increment
            sheetOpenPyxl = self.sheetOpenPyxl
            cellOpenPyxlIndex = input_data
            column, row = cell.counting_index
            row += 1 #int(self.have_headers)
            try:
                cellOpenPyxl = sheetOpenPyxl.cell(row=row, column=column)
                # cellOpenPyxl = sheetOpenPyxl._cells[cellOpenPyxlIndex]
            except KeyError:
                return ''
            
            # print(7104, sheetOpenPyxl)
            # print(cellOpenPyxl, cellOpenPyxl.value)
            # print('cell-coord', cellOpenPyxl.coordinate, 'to', cell)
            
            # print(7108, cell.formula.startswith('=') )
            # if cell.formula.startswith('='):
            
            cell.formula = cellOpenPyxl.value
            
            # print(7113, cell.formula )
            fill_color = cellOpenPyxl.fill.fgColor
            font = cellOpenPyxl.font
            alignmentSet = cellOpenPyxl.alignment
            horizontal = alignmentSet.horizontal
            vertical = alignmentSet.vertical
            hyperlinkSet = cellOpenPyxl.hyperlink
            border = cellOpenPyxl.border
            
            # if andExec(not cell.is_placeholder, not cell.column.is_placeholder):
                # cell.autoInit()
                
            if fill_color:
                fill_color_value = fill_color.value
                # if debug: print(f'{cell} fill_color_value = {fill_color_value}')
                if isinstance(fill_color_value, str):
                    if fill_color_value[:2] == '00':
                        cell.bg = '#FFFFFF'
                    else:
                        cell.bg = '#' + fill_color.value[2:]
            if font.color:
                font_color_value = font.color.value
                # if debug: print(f'{cell} font_color_value = {font_color_value}')
                if hyperlinkSet:
                    if hyperlinkSet.display:
                        cell.fg = '#0000FF'
                elif isinstance(font_color_value, str):
                    # print(7150, (not cell.is_placeholder, not cell.column.is_placeholder) )
                    cell.fg = '#' + font.color.value[2:]
                    
            if hyperlinkSet:
                if hyperlinkSet.display:
                    cell.link = hyperlinkSet.display
                    
            fontDisplay = ()
            if font.name: fontDisplay += (f'{font.name}',)
            if font.size:
                fontDisplay += (f'{int(font.size)}',)
                cell_height = get_text_height(f'{font.name} {int(font.size)}')*1.35
            if font.bold: fontDisplay += ('bold',) #; print(cell, fontDisplay)
            if font.italic: fontDisplay += ('italic',)
            if font.underline: fontDisplay += ('underline',)
            # print(7150, cell, fontDisplay)
            # print()
            cell.font = fontDisplay
            cell.anchorVar, cell.justify = alignmentRule[horizontal, vertical][cell.is_numeric()]
            
            # cell.column.fonts += [cell.font]
            
            # if self.sheetName == 'Spring 2024':
                # print(f"{cell}\t{cellOpenPyxl}\t{cellOpenPyxlIndex}\t{horizontal, vertical}\t{cell.anchorVar}")
            
            # if debug:
                # print(cell, str(border).replace("type='indexed', ", "type='indexed',\n\n"), sep='\n', end='\n\n')
            
            if border.right.style:     cell.switchBorderRight()
            if border.left.style:      cell.switchBorderLeft()
            if border.bottom.style:    cell.switchBorderBottom()
            if border.top.style:       cell.switchBorderTop()
            
            # print([rowIndex+row_increment, colIndex+1], cellOpenPyxl.value)
            
            # cell.autoInit()
            
            return cellOpenPyxl.value
            
        # except (KeyError, IndexError, AttributeError) as e:
        except Exception as e:
            print("%s\t%s\t%s\n%s\n\n" % (7208, [rowIndex+row_increment, colIndex+1], e, format_exc()) , file=sys.stderr)
            # print(7183, [rowIndex+row_increment, colIndex+1], e)
            # print(7184, f"{cell}\t{cellOpenPyxl}\t{cellOpenPyxlIndex}\t{horizontal, vertical}\t{cell.anchorVar}")
            # print(7185, format_exc(), file=sys.stderr)
            raise
            return ''
        
    # def autoFormatReverse(self, sheetOpenPyxl, cellOpenPyxlIndex, cell, row_increment, debug=0):
        # # font = cellOpenPyxl.font
        # border = cellOpenPyxl.border
        # hyperlinkSet = cellOpenPyxl.hyperlink
        # cellOpenPyxl.fill.fgColor = cell.bg
        # alignmentSet.horizontal, alignmentSet.vertical = alignmentReverse[anchor]
        # # font.color.value = cell.fg
        # hyperlinkSet.display = cell.link
        
        # # Retrieve font attributes from Each and Every Cell
        # tkinter_font = cell.cget("font")
        # tkinter_font_attributes = tkFont.Font(font=tkinter_font)
        # font_name = tkinter_font_attributes.actual()["family"]
        # font_size = tkinter_font_attributes.actual()["size"]
        # font_bold = tkinter_font_attributes.actual()["weight"] == "bold"
        # font_color = cell.fg
        # cellOpenPyxl.font = openpyxl.styles.Font(name=font_name, size=font_size, bold=font_bold, color=font_color)
        
        # if cell.right_border:     border.right.style = 'thin'
        # if cell.left_border:      border.left.style = 'thin'
        # if cell.bottom_border:    border.bottom.style = 'thin'
        # if cell.top_border:       border.top.style = 'thin'
        
    def insert_row_and_translate_formulae(self, ws):
        # Insert a row at the top
        ws.insert_rows(1, amount=1)
        
        # Translate all formulae one row below
        max_row = ws.max_row
        max_col = ws.max_column
    
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if cell.data_type == 'f':  # If the cell contains a formula
                    formula = cell.value
                    translated_formula = translate_formula(formula, 1)
                    cell.value = translated_formula
        
        # for row in range(2, max_row + 1):
            # for col in range(1, max_col + 1):
                # cell = ws.cell(row, col)
                # if cell.data_type == 'f':  # If the cell contains a formula
                    # formula = cell.value
                    # translated_formula = Translator(formula, origin="A{}:{}".format(2, max_row),
                                                    # target="A{}:{}".format(3, max_row + 1)).translate_formula()
                    # cell.value = translated_formula
        
    @property
    def nCols(self):
        return self._nCols
    @nCols.setter
    def nCols(self, new_nCols):
        structData = self.structData
        self._nCols = structData['nCols'] = new_nCols
        
    @property
    def nRows(self):
        return self._nRows
    @nRows.setter
    def nRows(self, new_nRows):
        structData = self.structData
        self._nRows = structData['nRows'] = new_nRows
        
    @property
    def width(self):
        return self._width
    @width.setter
    def width(self, new_width):
        self._width = self.structData['width'] = new_width
        
    @property
    def height(self):
        return self._height
    @height.setter
    def height(self, new_height):
        self._height = self.structData['height'] = new_height
        
    @property
    def columnSingleWidth(self):
        return self._columnSingleWidth
    @columnSingleWidth.setter
    def columnSingleWidth(self, new_columnSingleWidth):
        self._columnSingleWidth = self.structData['columnSingleWidth'] = new_columnSingleWidth
        
    @property
    def colTitles(self):
        return self._colTitles
    @colTitles.setter
    def colTitles(self, new_colTitles):
        self._colTitles = new_colTitles
        # if self.extension != '.qutable':
            # self.structDatacolumns'] = [None]
        # for colIndex, new_colTitle in zip(range(self.nCols), new_colTitles):
            # colDict = self.structData['columns'][colIndex]
            # if hasattr(colDict, '__iter__'):
                # colDict['title'] = new_colTitle
        
    @property
    def selectedCellsSet(self):
        return self._selectedCellsSet
    @selectedCellsSet.setter
    def selectedCellsSet(self, new_selectedCellsSet):
        self._selectedCellsSet = new_selectedCellsSet
        self.selectedColsIndex = []
        self.selectedRowsIndex = []
        for cell_range in new_selectedCellsSet:
            cell_range.toggleCellColor(select=True)
            self.selectedColsIndex += list(cell_range.col_range)
            self.selectedRowsIndex += list(cell_range.row_range)
        
    def clickAction(self, event, link, debug=0):
        x, y = event.x_root, event.y_root
        widgetStart, widgetEnd = event.widget, event.widget.winfo_containing(x, y)
        
        if widgetStart == widgetEnd:
            if is_browser_only_link(link):
                confirm = messagebox.askyesno('Security Verification', 'QuTable is unable to verify whether or not the link is secure.\n\nIf you want to open the link, open it at your own risk.\n\nAre you sure you want to open the link?', parent=self.book.window)
                if not confirm: return
            if link:
                os.startfile(link)
        
    def mouseWheel(self, event=None):
        # event.state == 8 ==> Vertical Up-Down Scroll
            # + == Up
            # - == Down
        # event.state == 9 ==> Horizontal Right-Left Scroll
            # + == Right
            # - == Left
        # if event.state == 9: print('mouseWheel xview =', self.xview() )
        # elif event.state == 8: print('mouseWheel yview =', self.yview() )
        # print('mouseWheel event.num =', event.num)
        # self.mouseWheelEvent = event
        
        # print('mouseWheel event.state =', event.state)
        # print('mouseWheel event.delta =', event.delta)
        # print()
        
        if event.state == 8:
            # yview_action = self.xview_tuple == self.xview()
            # if yview_action:
            current_yview = self.yview()
            delta = -event.delta/6000
            self.yview('moveto', current_yview[0]+delta)
            self.yview_tuple = self.yview()
                # return
            
        # if event.state == 9 or not yview_action:
        if event.state == 9:
            # print('mouseWheel sheet.current_yview =', self.yview() )
            # print('mouseWheel sheet.yview_tuple =', self.yview_tuple)
            # print('mouseWheel self.yview_tuple != self.yview() =', self.yview_tuple != self.yview() )
            
            current_xview = self.xview()
            delta = event.delta/6000
            self.xview('moveto', current_xview[0]+delta)
            self.xview_tuple = self.xview()
            
            # print()
            # print()
            
        if self.entry:
            self.currentCell.edit_stop(event=event, focusOut=True)
        
    def cells_to_CellRange(self, cell1, cell2, cond_col_opposite_dir=False, cond_row_opposite_dir=False):
        return CellRange( ColRange(cell1.index[0], cell2.index[0], cond_opposite_dir=cond_col_opposite_dir) ,
                          RowRange(cell1.index[1], cell2.index[1], cond_opposite_dir=cond_row_opposite_dir) ,
                          sheet=self , book=self.book )
        
    def cellIndex_to_CellRange(self, cellIndexCode, cond_col_opposite_dir=False, cond_row_opposite_dir=False):
        cellStartIndexCode, cellStopIndexCode = cellIndexCode.split(':')
        cellStartIndexTuple = self.handleIndex(cellStartIndexCode, True)
        cellStopIndexTuple = self.handleIndex(cellStopIndexCode, True)
        return CellRange( ColRange(cellStartIndexTuple[0], cellStopIndexTuple[0], cond_opposite_dir=cond_col_opposite_dir) ,
                          RowRange(cellStartIndexTuple[1], cellStopIndexTuple[1], cond_opposite_dir=cond_row_opposite_dir) ,
                          sheet=self , book=self.book )
    
    # QuTable code handling
    def handleIndex(self, colCodeWIndex, return_tuple=False):
        if colCodeWIndex.startswith('C'):
            if 'R' in colCodeWIndex:
                indices1 = re.split('[CR]', colCodeWIndex)[1:]
                indices2 = tuple( int(index)-1 for index in indices1 )
            else:
                indices1 = colCodeWIndex.split('C')[1]
                indices2 = int(indices1)-1
            if return_tuple:
                return indices2
            return self[indices2]
        
    def __getitem__(self, index):
        if type(index) == int:
            if index == -1:
                return self.indexColumn
            elif index == self.nCols:
                return self.placeholderColumn
            else:
                return self.columns[index]
            
        elif type(index) == ColRange:
            slice_range = index.slicer
            indexColumnSet = CellSet(self.book, self.sheet, [self.indexColumn] if index.start == -1 else [] )
            index_start = index.start
            if andExec(index.start < 0, index.stop > index.start):
                index.start = 0
            ordinaryCellSet = CellSet(self.book, self.sheet, self.columns[index] )
            placeholderColumnSet = CellSet(self.book, self.sheet, [self.placeholderColumn] if index.stop == self.nCols else [] )
            index.start = index_start
            selectedColsSet = indexColumnSet + ordinaryCellSet + placeholderColumnSet
            return CellSet(self.book, self.sheet, selectedColsSet)
            
        elif type(index) == range:
            slice_range = slice(index.start, index.stop, index.step)
            return CellSet(self.book, self.sheet, self.columns[slice_range])
            
        elif type(index) == CellRange:
            # cells = []
            # for col in self[index.col_range.slicer]:
                # cells += col.cells[index.row_range.slicer]
            cells = self[index.col_range][index.row_range]
            return CellSet(self.book, self.sheet, cells)
            
        elif type(index) == str:
            if index.startswith('C'):
                if ':' in index:
                    indicesRaw = index.split(',')
                    indices = CellSet(self.book, self.sheet)
                    for indexRaw in indicesRaw:
                        selector = self.cellIndex_to_CellRange(indexRaw)
                        indices.add( self[selector] )
                    return indices
                return self.handleIndex(index)
            return super().__getitem__(index)
            
        elif hasattr(index, '__iter__'):
            if len(index) == 1:
                return self.columns[index[0]]
            elif len(index) == 2:
                if andExec(type(index[0]) == slice, type(index[1]) == slice):
                    cells = []
                    for col in self[index[0]]:
                        cells += col.cells[index[1]]
                    return CellSet(self.book, self.sheet, cells)
                elif orExec(isinstance(index[0], ColRange), isinstance(index[1], RowRange)):
                    col_range, row_range = index
                    cells = self[col_range][row_range]
                    # for col in self[index[0].slicer]:
                        # cells += col.cells[index[1].slicer]
                    return CellSet(self.book, self.sheet, cells)
                elif isinstance(index[0], ColRange):
                    return CellSet(self.book, self.sheet, [col.cells[index[1]] for col in self[index[0].slicer]])
                elif isinstance(index[1], RowRange):
                    return CellSet(self.book, self.sheet, [col.cells[index[1].slicer] for col in self[index[0]]])
                else:
                    if andExec(index[0] == -1, index[1] == -1):
                        return self.indexColumn.colTitleCell
                    elif andExec(index[0] == self.nCols, index[1] == self.nRows):
                        return self.placeholderColumn.placeholderCell
                    elif index[0] == -1:
                        # if index[0] == self.nCols:
                            # return self.indexColumn.placeholderCell
                        if index[1] == self.nRows:
                            return self.indexColumn.placeholderCell
                        else:
                            return self.indexColumn[index[1]]
                    elif index[1] == -1:
                        if index[0] == self.nCols:
                            return self.placeholderColumn.colTitleCell
                        else:
                            return self.colTitleCells[index[0]]
                    elif index[0] == self.nCols:
                        if index[1] == self.nRows:
                            return self.placeholderColumn.placeholderCell
                        else:
                            return self.placeholderColumn.cells[index[1]]
                    elif index[1] == self.nRows:
                        return self.columns[index[0]].placeholderCell
                    else:
                        # print(6933, index)
                        return self.cells[index[0]*self.nRows+index[1]]
                        # return self.columns[index[0]].cells[index[1]]
            
        elif type(index) == slice:
            return CellSet(self.book, self.sheet, self.columns[index])
        
        elif orExec(isinstance(index, ColRange), isinstance(index[1], RowRange)):
            return CellSet(self.book, self.sheet, self.columns[index.slicer])
        
    def __setitem__(self, index, value):
        if type(index) == int:
            self.columns[index] = value
            
        elif type(index) == range:
            slice_range = slice(index.start, index.stop, index.step)
            self.columns[slice_range] = value
            
        elif type(index) == str:
            super().__setitem__(index, value)
            
        elif hasattr(index, '__iter__'):
            if len(index) == 1:
                self.columns[index[0]] = value
            elif len(index) == 2:
                if andExec(type(index[0]) == slice, type(index[1]) == slice):
                    for col in self[index[0]]:
                        col.cells[index[1]] = value
                elif type(index[0]) == slice:
                    for col in self[index[0]]:
                        col.cells[index[1]] = value
                elif andExec(isinstance(index[0], ColRange), isinstance(index[1], RowRange)):
                    for col in self[index[0].slicer]:
                        col.cells[index[1].slicer] = value
                elif isinstance(index[0], ColRange):
                    for col in self[index[0].slicer]:
                        col.cells[index[1]] = value
                elif isinstance(index[1], RowRange):
                    for col in self[index[0]]:
                        col.cells[index[1].slicer] = value
                else:
                    self.columns[index[0]].cells[index[1]] = value
                    self.cells[index[0]*self.nRows+index[1]] = value
            
        elif type(index) == slice:
            self.columns[index] = value
        
        elif orExec(isinstance(index, ColRange), isinstance(index[1], RowRange)):
            self.columns[index.slicer] = value
        
    def __delitem__(self, index):
        if type(index) == int:
            del self.columns[index]
            
        elif type(index) == range:
            slice_range = slice(index.start, index.stop, index.step)
            del self.columns[slice_range]
            
        # elif type(index) == str:
            # super().__setitem__(index, value)
            
        elif hasattr(index, '__iter__'):
            if len(index) == 1:
                del self.columns[index[0]]
            elif len(index) == 2:
                if andExec(type(index[0]) == slice, type(index[1]) == slice):
                    for col in self[index[0]]:
                        del col.cells[index[1]]
                elif type(index[0]) == slice:
                    for col in self[index[0]]:
                        del col.cells[index[1]]
                elif andExec(isinstance(index[0], ColRange), isinstance(index[1], RowRange)):
                    for col in self[index[0].slicer]:
                        del col.cells[index[1].slicer]
                elif isinstance(index[0], ColRange):
                    for col in self[index[0].slicer]:
                        del col.cells[index[1]]
                elif isinstance(index[1], RowRange):
                    for col in self[index[0]]:
                        del col.cells[index[1].slicer]
                else:
                    del self.columns[index[0]].cells[index[1]]
                    del self.cells[index[0]*self.nRows+index[1]]
            
        elif type(index) == slice:
            del self.columns[index]
        
        elif orExec(isinstance(index, ColRange), isinstance(index[1], RowRange)):
            del self.columns[index.slicer]
        
    def convertStructDataToXLSCell(self, cell):
        ws = self.sheetOpenPyxl
        
        # self.structData = {'bg':self.bg, 'fg':self.fg, 'anchor':self.anchorVar, 'justify':self.justify, 'link':self.link, 'top_border':int(self.top_border), 'bottom_border':int(self.bottom_border), 'left_border':int(self.left_border), 'right_border':int(self.right_border), 'height':self.row_height, 'content':self.content, 'formula':self.formula, 'font':self.font}
        
        structDataAttrs = ['top_border', 'right_border', 'link', 'left_border', 'justify', 'height', 'formula', 'font', 'fg', 'content', 'bottom_border', 'bg', 'anchor']
        
        structData = cell.structData
        column, row = cell.counting_index
        row += 1
        
        for structDataAttr in structDataAttrs:
            if structDataAttr not in cell.structData:
                cell.structData[structDataAttr] = ''
        
        cellOpenPyxl = ws.cell(row=row, column=column)
        
        # Set the value/content of the cell
        if structData['formula']:
            qutableFormula = structData['formula']
            cellOpenPyxl.value = convert_qutable_to_excel(qutableFormula)
        else:
            cellOpenPyxl.value = structData['content']

        # Apply background color
        if structData['bg']:
            bgMod = structData['bg'].strip('#')
            fill = PatternFill(start_color=bgMod, end_color=bgMod, fill_type='solid')
            cellOpenPyxl.fill = fill

        # Apply font color
        # :
        fgMod = structData['fg'].strip('#') if structData['fg'] else '000000'
        # cellOpenPyxl.font = Font(color=fgMod)

        # Apply font
        if structData['font']:
            # Retrieve font attributes from Each and Every Cell
            tkinter_font = cell.cget("font")
            tkinter_font_attributes = tkFont.Font(font=tkinter_font)
            font_name = tkinter_font_attributes.actual()["family"]
            font_size = tkinter_font_attributes.actual()["size"]
            font_bold = tkinter_font_attributes.actual()["weight"] == "bold"
            font_color = cell.fg
            cellOpenPyxl.font = openpyxl.styles.Font(name=font_name, size=font_size, bold=font_bold, color=fgMod)
            # cellOpenPyxl.font = structData['font']

        # Apply alignment
        if andExec(structData['anchor'], structData['justify']):
            anchor = structData['anchor']
            horizontal, vertical = alignmentReverse[anchor]
            cellOpenPyxl.alignment = Alignment(horizontal=horizontal, vertical=vertical)

        # Apply hyperlink
        if structData['link']:
            cellOpenPyxl.hyperlink = structData['link']

        # Apply borders
        if orExec(structData['top_border'], structData['bottom_border'], structData['left_border'], structData['right_border']):
            border = Border(
                top=structData['top_border'],
                bottom=structData['bottom_border'],
                left=structData['left_border'],
                right=structData['right_border']
            )
            cellOpenPyxl.border = border

        # Apply row height
        if structData['height']:
            ws.row_dimensions[row].height = structData['height']
        
    def convertStructDataToXLS(self):
        for cell in self.cells:
            self.convertStructDataToXLSCell(cell)
        
    def __repr__(self):
        return self.sheetName
        
    def __str__(self):
        return self.sheetName
        
    def __iter__(self):
        for column in self.columns:
            yield column

    def __lt__(self, other):
        return self.columns < other.columns

    def __gt__(self, other):
        return self.columns > other.columns

    def __le__(self, other):
        return self.columns <= other.columns

    def __ge__(self, other):
        return self.columns >= other.columns
        
    def __len__(self):
        return len(self.columns)

    def __add__(self, other):
        if type(other) == type(self):
            self.columns += other.columns
            self.nCols += len(other.columns)
        elif hasattr(other, '__iter__'):
            self.columns += other
            self.nCols += len(other)
        
    def parse_expr_with_sheet(self, expr_str):
        from sympy import symbols, Symbol
        
        # Parse the expression
        expr = parse_expr(expr_str)
        
        return srepr(expr)
        
        # Iterate through the symbols in the expression
        for symbol in expr.free_symbols:
            # Replace the symbol with a new symbol including the sheet name
            expr = expr.subs(symbol, Symbol(symbol.name, sheet='.'))
            expr = expr.subs(symbol, Symbol(symbol.name, sheet='.'))

        return srepr(expr).replace('True', f'app["{self.book.title}"]["{self.sheetName}"]')
            
    def makeIndexReadable(self, List, indicator):
        newList = [str(index+1) for index in List if isinstance(index, numbers.Number)]
        newList.sort()
        return List
        return indicator + f', {indicator}'.join( newList )

    def modify_currentCellSelReader(self, selectedColsIndex, selectedRowsIndex, init=False):
        if not init:
            startCellCode = self.selectedCellsSet[-1].getStartingCell().cellIndexCode
            lastCellCode = self.selectedCellsSet[-1].getLastCell().cellIndexCode
            
            display = startCellCode if startCellCode == lastCellCode else f'{startCellCode}:{lastCellCode}'
            if self.book.currentCellSelector:
                self.book.currentCellSelector.delete(0, 'end')
                self.book.currentCellSelector.insert(0, display)
            
            # display = f'Selected Cell: {self.currentCell.cellIndexCode}\t\t\tSelected Columns: {self.makeIndexReadable(selectedColsIndex, "C")}\t\t\tSelected Rows: {self.makeIndexReadable(selectedRowsIndex, "R")}'
            # self.book.currentCellSelector['text'] = display
        
    def generateColIndex(self, readable=False):
        return [column.index+1 if readable else column.index for column in self]
        
    def remove(self, column):
        self.columns.pack_forget()
        self.columns.remove(column)
        
    def selectAllCells(self):
        self.currentCell.selectbyDragging(col_range=range(self.nCols), row_range=range(self.nRows))
        
    def clear(self):
        def clearExt(cell):
            cell.column.previous_cont = cell.content
            cell.column.current_cont = '0'
            cell.formula = ''
        self.selectedCellsSet.map(clearExt)
        self.book.formulaEntry.delete(0, 'end')
        self.resetQStats()
        
    def clearCells(self):
        def execute(selectedCell):
            selectedCell.formula = ''
        self.selectedCellsSet.map( execute )
        
    def copyStrGenerate(self, cell_range=None, colSeparator='\t', rowSeparator='\n', debug=None):
        text = ''
        if isinstance(cell_range, LineRange):
            cell_range = cell_range.generateRange()
        cell_ranges = [cell_range] if cell_range else self.selectedCellsSet
        for selectedRange in cell_ranges:
            selectedCellsFromRange = selectedRange.generatePlus()
            nCols = selectedRange.nCols
            nRows = selectedRange.nRows
            for i1 in range(nRows):
                for i2 in range(nCols):
                    text += selectedCellsFromRange[i1+i2*nRows].getContent()
                    if i2 < nCols-1:
                        text += colSeparator
                text += rowSeparator
        return text.strip()
        
    def copyHTMLGenerate(self, debug=None):
        html = ''
        for selectedRange in self.selectedCellsSet:
            html += '<table>\n'
            selectedCellsFromRange = selectedRange.generatePlus()
            nCols = selectedRange.nCols
            nRows = selectedRange.nRows
            for i1 in range(nRows):
                html += '<tr>\n'
                for i2 in range(nCols):
                    cell = selectedCellsFromRange[i2*nRows+i1]
                    cell_css = convert_to_css(cell.font)
                    subtext = cell.getContent()
                    tag = 'th' if isinstance(cell, ColumnTitle) else 'td'
                    html += f"<{tag} style='{cell_css}'>{subtext}</{tag}>\n"
                html += '</tr>\n'
            html += '</table>'
        return html
        
    def copy(self):
        klembord.set_with_rich_text(self.copyStrGenerate(), self.copyHTMLGenerate())
        # selectedCellsContents3 = self.copyStrGenerate()
        # self.window.clipboard_clear()
        # self.window.clipboard_append(selectedCellsContents3)
        
    def cut(self):
        self.cut_triggered = True
        self.copy()
        
    def paste(self, starting_cell=None, text1=None):
        if self.cut_triggered:
            self.clear()
            self.cut_triggered = False
            
        if type(text1) != str:
            text1 = self.clipboard_get()
        text2 = text1.strip().replace('\r', '\n').replace('\v', '\n').replace('\f', '\n').split('\n')
        text3 = [t2.split('\t') for t2 in text2]
        
        if not starting_cell: starting_cell = self.currentCell
        colIndexBegin, rowIndexBegin = starting_cell.index
        
        if andExec(not text1, not starting_cell):
            text_old = self.copyStrGenerate()
        
        index_error = []
        for t3c in range(colIndexBegin, colIndexBegin+len(text3[0]) ):
            for t3, t3r in zip( text3 , range(rowIndexBegin, rowIndexBegin+len(text3) ) ):
                try:
                    # print('t3c, t3r =', t3c, t3r, end='\t')
                    text3a = t3[t3c-colIndexBegin]
                    # print('text3a =', text3a)
                    cell_t3c_t3r = self[t3c, t3r]
                except IndexError as index_error_:
                    index_error += [type(index_error_)]
                    continue
                if andExec(not text1, not starting_cell):
                    cell_t3c_t3r.toggleCell(prev_color=cell_t3c_t3r.cget('bg'))
                if isinstance(cell_t3c_t3r, Cell):
                    cell_t3c_t3r.formula = text3a
            if andExec(index_error, hasattr(index_error, '__iter__')):
                if index_error[0] == IndexError:
                    break
        
        # self.changesMade += [(    lambda: self.paste(starting_cell=starting_cell, text1=text1)      ,
                                  # lambda: self.paste(starting_cell=starting_cell, text1=text_old)      )]
        
    def changeTextColor(self):
        # colorPicker = CTkColorPicker.AskColor(title='Text Color')
        # colorChosen = colorPicker.get()
        colorChosen = colorchooser.askcolor(parent=self.book.window, title='Text color', color=self.currentCell.fg)[1]
        if colorChosen:
            self.selectedCellsSet.map( lambda selectedCell: selectedCell.changeTextColor(colorChosen) )
            self.sheet.actionCollect(action='Change Text Color')
        
    def changeFillColor(self):
        # colorPicker = CTkColorPicker.AskColor(title='Fill Color')
        # colorChosen = colorPicker.get()
        colorChosen = colorchooser.askcolor(parent=self.book.window, title='Fill color', color=self.currentCell.bg)[1]
        if colorChosen:
            self.selectedCellsSet.map( lambda selectedCell: selectedCell.changeFillColor(colorChosen) )
            self.sheet.actionCollect(action='Change Fill Color')
    
    def switchAlignmentBtn(self, cell=None):
        pass
        # if not cell:
            # cell = self.selectedCellsSet[0].originCell
        # hAlign, vAlign = alignmentReverse[cell.anchorVar]
        # alignmentHBtns, alignmentVBtns = self.book.SuperMenuWidget.alignmentHBtns, self.book.SuperMenuWidget.alignmentVBtns
        # for alignmentHBtn in alignmentHBtns.values():
            # alignmentHBtn.switchOff()
        # for alignmentVBtn in alignmentVBtns.values():
            # alignmentVBtn.switchOff()
        # alignmentHBtns[hAlign].switchOn()
        # alignmentVBtns[vAlign].switchOn()
        
    def switchBorderBtn(self):
        pass
        # toggleBorderBtns = self.book.SuperMenuWidget.toggleBorderBtns
        # for toggleBorderBtn in toggleBorderBtns:
            # to_end, vertical = toggleBorderBtn.to_end, toggleBorderBtn.vertical
            # self.execBorder(func1=toggleBorderBtn.switchOff, func2=toggleBorderBtn.switchOn, to_end=to_end, vertical=vertical, switch_cond=True)
        
    def switchBtnsCombo(self):
        self.switchAlignmentBtn()
        self.switchBorderBtn()
    
    def toggleSelectBorder(self, flattened_range_cond=True, switch_cond=False, *args, **kwargs):
        if not self.selectedCellsSet: return
        cell_range = self.selectedCellsSet[-1]
        colBorderFunc = lambda cell: managers.append( cell.colSelector.winfo_manager() if cell else None )
        rowBorderFunc = lambda cell: managers.append( cell.rowSelector.winfo_manager() if cell else None )
        switchFuncs = [lambda cell: cell.rowSelector.install() if cell else None,
                 lambda cell: cell.colSelector.install() if cell else None]
        
        for to_end, vertical in zip(to_end_list, vertical_list):
            flattened_range = cell_range.flatten(to_end=to_end, vertical=vertical)
            RowColSepFunc = colBorderFunc if vertical else rowBorderFunc
            flattened_range.mapPlus(switchFuncs[vertical])
        
    def execBorder(self, func1=None, func2=None, to_end=True, vertical=True, flattened_range_cond=False, switch_cond=False, cellSet=None, slicer=LineRange(), *args, **kwargs):
        colBorderFunc = lambda cell: managers.append( cell.colBorder.winfo_manager() if cell else None )
        rowBorderFunc = lambda cell: managers.append( cell.rowBorder.winfo_manager() if cell else None )
        RowColSepFunc = colBorderFunc if vertical else rowBorderFunc
        
        if not cellSet:
            cellSet = self.selectedCellsSet
        
        for cell_range in cellSet:
            managers = []
            flattened_range = cell_range.flatten(to_end=to_end, vertical=vertical)
            flattened_range.mapPlus(RowColSepFunc)
            # flattened_range.mapPlus(lambda cell: print() )
            if flattened_range_cond:
                flattened_range.mapPlus(func1 if '' in managers else func2)
            else:
                func1(*args, **kwargs) if '' in managers else func2(*args, **kwargs)
        
    def toggleVBorder(self, cellSet=None, right=True):
        install_func = lambda cell: cell.colBorder.install()
        lift_func = lambda cell: cell.colBorder.takeoff()
        self.execBorder(install_func, lift_func, to_end=right, vertical=True, flattened_range_cond=True, cellSet=cellSet)
        self.switchBorderBtn()
        
    def toggleHBorder(self, cellSet=None, bottom=False):
        install_func = lambda cell: cell.rowBorder.install()
        lift_func = lambda cell: cell.rowBorder.takeoff()
        self.execBorder(install_func, lift_func, to_end=bottom, vertical=False, flattened_range_cond=True, cellSet=cellSet)
        self.switchBorderBtn()
        
    def toggleBorderLeft(self, cellSet=None):
        self.toggleVBorder(right=False, cellSet=cellSet)
        
    def toggleBorderRight(self, cellSet=None):
        self.toggleVBorder(right=True, cellSet=cellSet)
        
    def toggleBorderTop(self, cellSet=None):
        self.toggleHBorder(bottom=False, cellSet=cellSet)
        
    def toggleBorderBottom(self, cellSet=None):
        self.toggleHBorder(bottom=True, cellSet=cellSet)
        
    def toggleBorderAll(self):
        toggle_func = lambda cell_range: cell_range.switchOnAllBorders() if cell_range.notAllBorders() else cell_range.switchOffAllBorders()
        self.selectedCellsSet.map(toggle_func, cell_absorbant=False)
        self.switchBorderBtn()
        
    def toggleBorderNone(self):
        lift_func = lambda cell_range: cell_range.switchOffAllBorders()
        self.selectedCellsSet.map(lift_func, cell_absorbant=False)
        self.switchBorderBtn()
        
    def insertRowBottomMost(self, nRowsSelected=None):
        if not isinstance(nRowsSelected, numbers.Number):
            nRowsSelected = self.selectedCellsSet.getRowSize()
        
        self.indexColumn.insertRowBottomMost(nRowsSelected)
        for column in self.columns:
            column.insertRowBottomMost(nRowsSelected)
        self.placeholderColumn.insertRowBottomMost(nRowsSelected)
                    
        self.cells = CellSet(self.book, self.sheet)
        for col in self:
            self.cells += col.cells
            col.placeholderCell.rowIndex += nRowsSelected
        
        # self.nRows = len(self.columns[0])
        self.nRows += nRowsSelected
        
        self.indexColumn.placeholderCell.rowIndex += nRowsSelected
        self.placeholderColumn.placeholderCell.rowIndex += nRowsSelected
        self.addRowGuide.grid_configure(row=self.sheetInner.grid_size()[1])
        
        # self.addStructData()
        self.resetQStats()
        if isinstance(nRowsSelected, numbers.Number):
            self.currentCell.focusOnCell()
    
    def insertColRightMost(self, nColsSelected=None):
        if not isinstance(nColsSelected, numbers.Number):
            nColsSelected = self.selectedCellsSet.getColSize()
        
        for i in range(nColsSelected):
            newColumn = Column(self.sheetInner, colIndex=self.nCols, book=self.book, width=self.columnSingleWidth, rowDimRawList=self.rowDimRawList, nRows=self.nRows, font=self.font, fonts=self[self.sheet.nCols-1].fonts, sheet=self, inserted=True)
            self + [newColumn]
            self.cells += newColumn.cells
            self.colTitleCells.add(newColumn.colTitleCell, init=True)
            
        self.placeholderColumn.colIndex += nColsSelected
        # self.placeholderColumn.recalibrateIndex(col_translation=nColsSelected)
        
        colSize, rowIndex = self.sheetInner.grid_size()
        self.addRowGuide.grid_configure(columnspan=colSize)
        
        # self.addStructData()
        self.resetQStats()
        if isinstance(nColsSelected, numbers.Number):
            self.currentCell.focusOnCell()
        
        rightMostCol = newColumn
        return rightMostCol
        
    def insertRow(self, below, debug=False):
        selectedCellsSet1 = self.sheet.selectedCellsSet.copy()
        selectedCellsSet2 = selectedCellsSet1.copy()#.shift_n_units(col_shift=below)
        selectedCellsSet2.sort(reverse=1, key=lambda cell_range: (cell_range.col_range.start, cell_range.row_range.start) )
        # print('selectedCellsSet2 =', selectedCellsSet2)
        
        row_range_prev = None
        for current_range in selectedCellsSet2:
            current_range.focusOnStartingCell()
            row_range = current_range.row_range
            if row_range == row_range_prev:
                continue
            nRowsSelected = row_range.getSize()
            row_range_new = row_range+nRowsSelected
            self.insertRowBottomMost(nRowsSelected)
            
            row_range_to_bottom_raw = row_range(inclusive=not below, sheet=self)
            row_range_to_bottom = row_range_to_bottom_raw.generateRange()
            row_range_to_bottom_new = row_range_to_bottom_raw+nRowsSelected
            current_range.switchOffSelector(color_trigger=True)
            
            row_range_used = row_range_new if below else row_range
            start = (0, row_range_used.start+nRowsSelected)
            starting_cell = self[start]
            text = self.copyStrGenerate(cell_range=row_range_to_bottom)
            row_range_used.replaceFormulae('')
            self.paste(starting_cell, text1=text)
            
            row_range_prev = row_range
        
        current_range.focusOnStartingCell()
        
        
    def insertColumn(self, right, debug=False):
        selectedCellsSet1 = self.sheet.selectedCellsSet.copy()
        selectedCellsSet2 = selectedCellsSet1.copy()
        selectedCellsSet2.sort(reverse=1, key=lambda cell_range: (cell_range.col_range.start, cell_range.row_range.start) )
        
        col_range_prev = None
        for current_range in selectedCellsSet2:
            col_range = current_range.col_range
            if col_range == col_range_prev:
                continue
            nColsSelected = col_range.getSize()
            if not right:
                current_range.shift_n_units(col_shift=-nColsSelected)
            current_range.focusOnStartingCell()
            col_range = current_range.col_range
            nColsSelected = col_range.getSize()
            col_range_new = col_range+nColsSelected
            self.insertColRightMost(nColsSelected)
            
            col_range_to_right_raw = col_range(inclusive=not right, sheet=self)
            col_range_to_right = col_range_to_right_raw.generateRange()
            col_range_to_right_new = col_range_to_right_raw+nColsSelected
            current_range.switchOffSelector(color_trigger=True)
            
            col_range_used = col_range_new if right else col_range
            start = (col_range_used.start+nColsSelected, 0)
            starting_cell = self[start]
            text = self.copyStrGenerate(cell_range=col_range_to_right)
            col_range_used.replaceFormulae('')
            
            colTitles = []
            self.colTitleCells[col_range_to_right_raw].map(lambda cell: colTitles.append( cell.getContent() ) )
            for colTitle, colTitleCellNew in zip(colTitles, self.colTitleCells[col_range_to_right_new]):
                colTitleCellNew.formula = colTitle
            self.colTitleCells[col_range_used].map(lambda cell: cell.replaceFormula( f'NewCol{cell.index[0]-col_range.start}' ) )
            self.paste(starting_cell, text1=text)
            
            col_range_prev = col_range
        
        current_range.focusOnStartingCell()
        
        self.addRowGuide.grid_configure({'columnspan': self.sheetInner.grid_size()[0] })
        
    def insertRowAbove(self):
        self.insertRow(below=False)
        self.resetQStats()
        
    def insertRowBelow(self):
        self.insertRow(below=True)
        self.resetQStats()
        
    def insertColLeft(self):
        self.insertColumn(right=False)
        self.resetQStats()
        
    def insertColRight(self):
        self.insertColumn(right=True)
        self.resetQStats()
        
    def reselectCell(self, selectedCellsSet):
        selectedCellsSet[0].focusOnStartingCell()
        selectedCellsSet.toggleCellColor(select=True)
        self.sheet.selectedCellsSet = selectedCellsSet
        selectedColsIndex, selectedRowsIndex = selectedCellsSet.generateIndexList()
        self.modify_currentCellSelReader(selectedColsIndex, selectedRowsIndex)
        
    def deleteCell(self, cell):
        cell.cellFrame.grid_forget()
        cell.pack_forget()
        self.deleteSeparators(cell)
        del self.structData['cells'][str(cell.index)]
        
        if hasattr(cell, 'column'):
            cell.column.cells.remove(cell, affect_display=True)
        if isinstance(cell, Cell):
            cell.deleted = True
        elif isinstance(cell, IndexCell):
            self.cells.remove(cell, affect_display=True)
        
    def deleteCells(self, cells, delete_row=True):
        for cell in cells:
            self.deleteCell(cell)
            if delete_row:
                self.indexColumn.cells.remove(cell, affect_display=True)
        
    def deleteSeparators(self, cell):
        for separator in cell.cell_separators:
            try: separator.takeoff()
            except IndexError: pass
            if isinstance(separator, ColSeparator):
                if separator.sepType == 'ShadowBorder':
                    self.colShadowBorders.remove(separator, affect_display=True)
                    separator.parent.column.colShadowBorders.remove(separator, affect_display=True)
                elif separator.sepType == 'Selector':
                    self.colSelectors.remove(separator, affect_display=True)
                    separator.parent.column.colSelectors.remove(separator, affect_display=True)
                elif separator.sepType == 'Border':
                    self.colBorders.remove(separator, affect_display=True)
                    separator.parent.column.colBorders.remove(separator, affect_display=True)
                    
            if isinstance(separator, RowSeparator):
                if separator.sepType == 'ShadowBorder':
                    self.rowShadowBorders.remove(separator, affect_display=True)
                    separator.parent.column.rowShadowBorders.remove(separator, affect_display=True)
                elif separator.sepType == 'Selector':
                    self.rowSelectors.remove(separator, affect_display=True)
                    separator.parent.column.rowSelectors.remove(separator, affect_display=True)
                elif separator.sepType == 'Border':
                    self.rowBorders.remove(separator, affect_display=True)
                    separator.parent.column.rowBorders.remove(separator, affect_display=True)
        
    def deleteRowFunc(self, cellRangesSet):
        def cellRecalibrate(cell):
            cell.rowIndex -= nRowsSelected
        
        def placeholderCellRecalibrate(col):
            col.placeholderCell.rowIndex -= nRowsSelected
        
        for current_range in cellRangesSet:
            row_range = current_range.row_range
            nRowsSelected = row_range.getSize()
            row_range_to_bottom_raw = row_range(inclusive=0)
            row_range_to_bottom = row_range_to_bottom_raw.generateRange(super_inclusive=True)
            current_range.switchOffSelector(color_trigger=True)
            cellRecalibrate(self.indexColumn.placeholderCell)
            row_range_to_bottom.map(cellRecalibrate)
            inclusive_range = row_range.generateRange(super_inclusive=True)
            inclusive_range.mapPlus(self.deleteCell)
            self[inclusive_range.col_range].map(placeholderCellRecalibrate)
            self.deleteCells(self.indexColumn.cells[-nRowsSelected:])
            self.nRows -= nRowsSelected
        
        allCellSet = CellSet(self.book, self.sheet)
        row_list = cellRangesSet.generateIndexList()[1]
        
        for cell in self.cells:
            if not cell.deleted:
                allCellSet.add(cell, cell_raw=True)
            
        self.cells = allCellSet
        self[self.currentCell.index].focusOnCell()
        self.resetQStats()
        
    def deleteColumnFunc(self, cellRangesSet):
        def cellRecalibrate(cell):
            cell.colIndex -= nColsSelected
        
        def colRecalibrate(column):
            column.colIndex -= nColsSelected
        
        for current_range in cellRangesSet:
            col_range = current_range.col_range
            nColsSelected = col_range.getSize()
            col_range_to_right_raw = col_range(inclusive=0)
            col_range_to_right = col_range_to_right_raw.generateRange(super_inclusive=True)
            current_range.switchOffSelector(color_trigger=True)
            col_range_to_right.map(cellRecalibrate)
            self[col_range_to_right_raw].map(colRecalibrate)
            inclusive_range = col_range.generateRange(super_inclusive=True)
            inclusive_range.mapPlus(self.deleteCell)
            self.colTitleCells.remove(col_range)
            self.columns.remove(col_range)
            
            del self.structData['columns'][col_range.slicer]
            self.sheetOpenPyxl.delete_cols(col_range.start, nColsSelected)
            
            self.nCols -= nColsSelected
        
        allCellSet = CellSet(self.book, self.sheet)
        col_list = cellRangesSet.generateIndexList()[0]
        
        for cell in self.cells:
            if not cell.deleted:
                allCellSet.add(cell, cell_raw=True)
            
        self.cells = allCellSet
        self[self.currentCell.index].focusOnCell()
        self.resetQStats()
        
    def deleteRowBottomMost(self, nRowsSelected):
        if self.nRows <= 1:
            return
        col_range = ColRangeNaturalNum(1, 1, book=self.book, sheet=self)
        row_range = RowRangeNaturalNum(self.nRows-nRowsSelected+1, self.nRows, book=self.book, sheet=self)
        cellRangesSet = CellSet(self.book, self.sheet, [row_range.generateRange(col_range=col_range)])
        self.deleteRowFunc(cellRangesSet)
        
    def deleteColRightMost(self, nColsSelected):
        if self.nCols <= 1:
            return
        col_range = ColRangeNaturalNum(self.nCols-nColsSelected+1, self.nCols, book=self.book, sheet=self)
        row_range = RowRangeNaturalNum(1, 1, book=self.book, sheet=self)
        cellRangesSet = CellSet(self.book, self.sheet, [col_range.generateRange(row_range=row_range)])
        self.deleteColumnFunc(cellRangesSet)
        
    def deleteRow(self, nRowsSelected=None):
        if self.nRows <= 1:
            return
        if not isinstance(nRowsSelected, numbers.Number):
            nRowsSelected = self.selectedCellsSet.getRowSize()
        selectedCellsSetMod = self.selectedCellsSet.copy()
        selectedCellsSetMod.sort(reverse=0, key=lambda cell_range: (cell_range.row_range.start, cell_range.col_range.start) )
        self.deleteRowFunc(selectedCellsSetMod)
        
    def deleteColumn(self, nColsSelected=None):
        if self.nCols <= 1:
            return
        if not isinstance(nColsSelected, numbers.Number):
            nColsSelected = self.selectedCellsSet.getColSize()
        selectedCellsSetMod = self.selectedCellsSet.copy()
        selectedCellsSetMod.sort(reverse=1, key=lambda cell_range: (cell_range.col_range.start, cell_range.col_range.start) )
        self.deleteColumnFunc(selectedCellsSetMod)
        
    def generateSheetInnerGridInfo(self):
        """
        Convert a Tkinter frame containing widgets in a grid layout into a Pandas DataFrame 
        with visual representation.
        
        Args:
        - frame (tk.Frame): The Tkinter frame containing widgets in a grid layout.
        
        Vars:
        - grid_size (tuple): Tuple specifying the grid size (rows, columns).
        
        Returns:
        - pd.DataFrame: DataFrame representing the grid layout with widget names.
        """
        
        grid_size = self.sheetInner.grid_size()
        
        # Initialize a list to hold data for each cell
        cells_data = []

        # Iterate through all widgets in the frame
        for row in range(grid_size[0]):
            row_data = []
            for col in range(grid_size[1]):
                widget = self.sheetInner.grid_slaves(row=row, column=col)
                if widget:
                    cell_label = f"{widget[0]} at col {col}, row {row}"
                else:
                    cell_label = ""
                row_data.append(cell_label)
            cells_data.append(row_data)
        
        # Create DataFrame from cells_data
        df = pd.DataFrame(cells_data)
        
        # Rename columns with their indices
        df.columns = [f"{col}" for col in range(grid_size[1])]
        
        return df
        
        # """
        # Converts a tkinter grid layout within a frame into a pandas DataFrame.
        
        # Parameters:
        # - frame (tk.Frame): The tkinter frame containing the grid layout.
        
        # Returns:
        # - pd.DataFrame: DataFrame where each row corresponds to a widget in the grid.
                        # Columns: ['Row', 'Column', 'Content']
        # """
        # # Initialize an empty list to hold rows of data
        # data = []

        # # Loop through all widgets in the frame
        # for child in self.sheetInner.winfo_children():
            # subdata = []
            
            # # Check if the widget is a label or an entry (assuming for simplicity)
            # info = child.grid_info()
            
            # if child.winfo_manager() == 'grid':
                # row = info['row']
                # column = info['column']
                
                # # Append data as a tuple (column, row, content)
                # subdata.append(child)
                
                # # # Append data as a tuple (column, row, content)
                # # data.append((column, row, child))
                
            # # if isinstance(child, (tk.Label, tk.Entry)):
                # # # Get row and column information from the widget
                # # info = child.grid_info()
                # # row = info['row']
                # # column = info['column']

                # # # Get the content of the widget (assuming it has a .get() method)
                # # content = child.cget('text') if isinstance(child, tk.Label) else child.get()

                # # # Append data as a tuple (row, column, content)
                # # data.append((row, column, content))

        # # Convert the data into a pandas DataFrame
        # summary = pd.DataFrame(data, columns=['Row', 'Column', 'Content'])
        # return summary
        
    def makeChart(self):
        self.graphPlotter.master.deiconify()
        self.graphPlotter.select_columns()

    def Scroll(self, EventDelta):
        MoveBy = 0.05 if EventDelta < 0 else -0.05
        self.yview_moveto(self.yview()[0]+MoveBy)

    def removeSelectionMark(self, event=None, currentCell=None):
        # Remove Selected CellLabel from Lists of selectedCellsSet
        for cell_range in self.selectedCellsSet:
            if cell_range.is_empty():
                continue
            cell_range.toggleCellColor(eventInherited=event, select=False, selectMultipleInOne=True)
            cell_range.switchOffSelector()
            if cell_range.row_range.start == -1:
                for column, colTitleCell in zip(self[cell_range.col_range], self.colTitleCells[cell_range.col_range]):
                    column.rowSelector.takeoff()
                    colTitleCell.toggleCellColor(eventInherited=event, select=False, selectMultipleInOne=True)
        if self.entry:
            self.currentCell.edit_stop(event=event, focusOut=True)

    def MoveMarkByCoord(self, colMove, rowMove, event=None):
        try:
            colIndex1 = self.colIndex+colMove
            if andExec(colIndex1 >= 0, colIndex1 < len(self.columns)): self.colIndex = colIndex1
            
            rowIndex1 = self.rowIndex+rowMove
            if andExec(rowIndex1 >= 0, rowIndex1 < len(self.columns[0].cells)): self.rowIndex = rowIndex1
            
            self.SelectionMarkByCoord(colIndex=self.colIndex, rowIndex=self.rowIndex, event=event)
            
        except IndexError:
            pass

    def SelectionMarkByCoord(self, colIndex, rowIndex, event=None):
        global current_range
        self.removeSelectionMark()
        if self.currentCell is not None:
            self.currentCell.selected = False
        
        self.selectedCell = self[colIndex, rowIndex]
        self.selectedCell.previous_range = None
        self.selectedCell.selected = True
        # self.selectedCell.cellFrame.focus_set()
        current_range = self.selectedCell.generateRange()
        current_range.switchOnSelector()
        
        self.sheet.previous_range = None
        
        self.switchAlignmentBtn(self.selectedCell)
        
        self.currentCell = self.lastCell = self.selectedCell
        self.column = self.currentCell.column
            
        self.selectedCell.changeColorBtn()
        
        if self.book.ready:
            self.book.formulaEntry.delete(0, 'end')
            self.book.formulaEntry.insert(0, self.currentCell.formula)
        
        self.selectedCellsSet.restart( self.currentCell )
        
        self.colIndex, self.rowIndex, self.initialCell = colIndex, rowIndex, self.selectedCell
    
    def SheetsToDataFrame(self, toDataFrame=True, event=None):
        """Convert Sheets to dicts or dataframes"""
        df = {}
        
        for column in self.columns:
            df |= column.ColToDict()
        
        return pd.DataFrame(df) if toDataFrame else df

    def shiftSelectCellToggle(self, event=None):
        colSlice = slice(self.start_col, self.current_col+1) if self.current_col >= self.start_col else slice(self.current_col, self.start_col+1)
        rowSlice = slice(self.start_row, self.current_row+1) if self.current_row >= self.start_row else slice(self.current_row, self.start_row+1)
        
        # print( 'Column Slice:', colSlice )
        # print( 'Row Slice:', rowSlice )
        
        shiftSelectedCells = set( self[colSlice, rowSlice] )
        unSelectedCells = self.prevSelectedCells.difference(shiftSelectedCells)
        newlySelectedCells = shiftSelectedCells.difference(self.prevSelectedCells)
        
        for newlySelectedCell in newlySelectedCells:
            newlySelectedCell.toggleCell(event=event, select=True, selectMultipleInOne=True)
            
        for unSelectedCell in unSelectedCells:
            unSelectedCell.toggleCell(event=event, select=False, selectMultipleInOne=True)
        
        self.prevSelectedCells = shiftSelectedCells
        
    def ButtonRelease1Handle(self, event):
        for cell_range in self.selectedCellsSet:
            cell_range.toggleCellColor(eventInherited=event, select=True)
        
    def keyReleaseHandle(self, event):
        # print('KeyReleased =', event
        if event.keysym.lower().startswith('control'):   self.control_is_pressed = False
        if event.keysym.lower().startswith('alt'):       self.alt_is_pressed = False
        if event.keysym.lower().startswith('shift'):     self.shift_is_pressed = False
        if event.keysym.lower().startswith('win'):       self.win_is_pressed = False
        
    def keyShiftArrowHandle(self, event):
        # elif event.state == 1 or event.state == 262145:
        self.start_col, self.start_row = self.initialCell.index
        self.current_col, self.current_row = self.currentCell.index
        
        if event.keysym == 'Right':
            self.currentCell.selectbyDragging(widgetEnd=self.lastCell.moveRight())
        elif event.keysym == 'Left':
            self.currentCell.selectbyDragging(widgetEnd=self.lastCell.moveRight(-1))
        elif event.keysym == 'Up':
            self.currentCell.selectbyDragging(widgetEnd=self.lastCell.moveDown(-1))
        elif event.keysym == 'Down':
            self.currentCell.selectbyDragging(widgetEnd=self.lastCell.moveDown())
        elif event.keysym == 'Tab':
            self.MoveMarkByCoord(-1, 0)
        elif event.keysym == 'Return':
            self.MoveMarkByCoord(0, -1)
        elif orExec(event.keysym in string.ascii_lowercase, event.keysym in string.ascii_uppercase):
            self.currentCell.edit_start(text=event.char)
    
    def keyNoModHandle(self, event):
        # print(event.keysym)
        # if event.char:
            # self.currentCell.edit_start(text=event.char)
        if event.keysym == 'Escape':
            return
        elif event.keysym == 'Tab':
            self.MoveMarkByCoord(+1, 0)
        elif event.keysym == 'Return':
            self.MoveMarkByCoord(0, +1)
    # elif event.state == 0 or event.state == 262144:
        elif event.keysym == 'Right':
            self.MoveMarkByCoord(+1, 0)
        elif event.keysym == 'Left':
            self.MoveMarkByCoord(-1, 0)
        elif event.keysym == 'Up':
            self.MoveMarkByCoord(0, -1)
        elif event.keysym == 'Down':
            self.MoveMarkByCoord(0, +1)
        elif event.keysym == 'F2':
            self.currentCell.edit_start()
        elif andExec(event.char, not event.keysym.lower().startswith('control'), not event.keysym.lower().startswith('alt'), not event.keysym.lower().startswith('shift'), not event.keysym.lower().startswith('win'), not event.keysym.lower().startswith('app'), not event.keysym.startswith('F')):
            self.currentCell.edit_start(text=event.char)
        
    def keyHandle(self, event):
        # if cursor in columnFormulaEdit TextBox
        # if type(self.book.window.focus_get()) == tk.Text:
        # if type(self.book.window.focus_get()) == tk.Entry:
        if type(self.entry) == tk.Entry:
            return
        
        control, alt, shift, win, app = self.control_is_pressed, self.alt_is_pressed, self.shift_is_pressed, self.win_is_pressed, self.app_is_pressed
        self.modifier_is_pressed = [control, alt, shift, win, app]
        
        # print('event.num =', event.num)
        # print('event.state =', event.state)
        # print( control, alt, shift, event.keysym.lower() )
        # print()
        
        # If any of the Modifier keys are pressed:
        if orExec(event.keysym.lower().startswith('control'), event.keysym.lower().startswith('alt'), event.keysym.lower().startswith('shift')):
            if andExec(event.keysym.lower().startswith('control'), not control):
                self.control_is_pressed = True
            if andExec(event.keysym.lower().startswith('alt'), not alt):
                self.alt_is_pressed = True
            if andExec(event.keysym.lower().startswith('shift'), not shift):
                self.shift_is_pressed = True
            if andExec(event.keysym.lower().startswith('win'), not win):
                self.win_is_pressed = True
            if andExec(event.keysym.lower().startswith('app'), not app):
                self.app_is_pressed = True
            return
            
        
        # If Both Control and Alt Keys are pressed
        # elif event.state == 131084:
        if andExec(control, alt):
            if event.keysym.lower() == 's':
                self.book.saveAsBook()
            elif event.keysym.lower() == 'o':
                self.book.openBook()
            # else: return
            
        # If Both Control and Shift Keys are pressed
        # elif event.state == 13:
        elif andExec(control, shift):
            if event.keysym.lower() == 's':
                self.book.saveAsBook()
            elif event.keysym.lower() == 'o':
                self.book.openBook()
            # else: return
                
        # If Shift Key is pressed
        # if event.state == 9:
        elif shift:
            if event.keysym == 'Tab':
                self.MoveMarkByCoord(-1, 0)
            elif event.keysym == 'Return':
                self.MoveMarkByCoord(0, -1)
            else:
                self.keyShiftArrowHandle(event=event)
            
        # If Control Key is pressed
        # elif event.state == 12:
        elif control:
            if event.keysym.lower() == 'c':
                self.copy()
            elif event.keysym.lower() == 'v':
                self.book.paste()
            elif event.keysym.lower() == 's':
                self.book.saveBook()
            elif event.keysym.lower() == 'o':
                self.book.openBook()
            # else: return
        
        # If no Modifier Keys like Shift, Control, Alt, and Function are pressed
        # if event.state == 8:
        else:
            self.keyNoModHandle(event=event)
            
        # print('keyNoModHandle')
                

class SuperMenuSection(tk.Frame):
    def __init__(self, parent, book):
        super().__init__(parent, cursor='arrow', background='#FFFFFF')
        
class SuperMenu(tk.Text):
    column_grid = 0
    
    def __init__(self, parent, book):
        super().__init__(parent, state='disabled', cursor='arrow', background='#FFFFFF')
        self.SuperMenuFrame = tk.Frame(self, background='#FFFFFF')
        
        self.book = book
        self.SuperMenu1 = SuperMenuSection(self.SuperMenuFrame, book=book)
        
        self.new_book_sheet(book)
        self.open(book)
        self.undo_redo(book)
        self.save_save_as(book)
        self.copy_paste(book)
        self.text_fill_format(book)
        self.alignment_format(book)
        self.toggle_border(book)
        self.insert_delete_row_column(book)
        self.formula_part(book)
        
    def new_book_sheet(self, book):
        # Create New Book/Create New Sheet
        NewBook_btn_img_path = 'icons/NewBook.png'
        NewSheet_btn_img_path = 'icons/NewSheet.png'
        self.NewBook_btn_img = ImageTk.PhotoImage( imageDict['NewBook'].resize((60, 72)) , master=book.window )
        self.NewSheet_btn_img = ImageTk.PhotoImage( imageDict['NewSheet'].resize((60, 72)) , master=book.window )
        
        self.NewBook_btn = SuperButton(parent=self.SuperMenu1, text='New Book', image_path=NewBook_btn_img_path, image=self.NewBook_btn_img, book=book, font='Arial 12', command=book.newBook)
        self.NewSheet_btn = SuperButton(parent=self.SuperMenu1, text='New Sheet', image_path=NewSheet_btn_img_path, image=self.NewSheet_btn_img, book=book, font='Arial 12', command=book.createNewSheet)
        self.SuperMenu1.grid(row=0, column=self.column_grid); self.column_grid += 1
        self.NewBook_btn.grid(row=0, column=0)
        self.NewSheet_btn.grid(row=0, column=1)
        
        self.separator1 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator1.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
    def open(self, book):
        # Open
        open_btn_img_path = 'icons/Open.png'
        self.open_btn_img = ImageTk.PhotoImage( imageDict['Open'].resize((60, 72)) , master=book.window )
        self.open_btn = SuperButton(parent=self.SuperMenuFrame, text='Open File', image_path=open_btn_img_path, image=self.open_btn_img, book=book, font='Arial 14', command=book.openBook)
        self.open_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator2 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator2.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
    def undo_redo(self, book):
        # Undo/Redo
        undo_btn_img_path = 'icons/Undo.png'
        self.undo_btn_img = ImageTk.PhotoImage( imageDict['Undo'] , master=book.window )
        self.redo_btn_img = ImageTk.PhotoImage( imageDict['Redo'] , master=book.window )
        
        self.undo_btns = {}
        # self.undo_btn = SuperButton(parent=self.SuperMenuFrame, text='Undo', image_path=undo_btn_img_path, image=self.undo_btn_img, book=book, font='Arial 14', command=book.openBook)
        # self.undo_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        self.undo_btn_row = 0
        self.undo_btn_column = self.column_grid
        self.column_grid += 1
        
        self.undo_menu_btns = {}
        # for sheetName in book.sheetsDict:
            # self.undo_menu_btns[sheetName] = tk.Menubutton(self.SuperMenuFrame, text="▼", relief='flat', background='#FFFFFF')
            # self.undo_menu_btns[sheetName].grid(row=0, column=self.column_grid)
        # self.column_grid += 1
        self.undo_menu_btn_row = 0
        self.undo_menu_btn_column = self.column_grid
        self.column_grid += 1
        
        self.redo_btns = {}
        self.redo_btn_row = 0
        self.redo_btn_column = self.column_grid
        self.column_grid += 1
        
        self.redo_menu_btns = {}
        self.redo_menu_btn_row = 0
        self.redo_menu_btn_column = self.column_grid
        self.column_grid += 1
        
        self.separator2a = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator2a.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        
    def save_save_as(self, book):
        # Save/Save As
        self.SuperMenu1a = SuperMenuSection(self.SuperMenuFrame, book=book)
        
        save_btn_img_path = 'icons/Save.png'
        save_as_btn_img_path = 'icons/SaveAs.png'
        self.save_btn_img = ImageTk.PhotoImage( imageDict['Save'] , master=book.window )
        self.save_as_btn_img = ImageTk.PhotoImage( imageDict['SaveAs'] , master=book.window )
        
        self.save_btn = SuperButton(parent=self.SuperMenu1a, text='Save', image_path=save_btn_img_path, image=self.save_btn_img, book=book, font='Arial 12', command=book.saveBook)
        self.save_as_btn = SuperButton(parent=self.SuperMenu1a, text='Save As', image_path=save_as_btn_img_path, image=self.save_as_btn_img, book=book, font='Arial 12', command=book.saveAsBook)
        self.SuperMenu1a.grid(row=0, column=self.column_grid); self.column_grid += 1
        self.save_btn.pack(fill='both', expand=True)
        self.save_as_btn.pack(fill='both', expand=True)
        
        self.separator3 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator3.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
    def copy_paste(self, book):
        # Copy/Paste
        copy_btn_img_path = 'icons/Copy.png'
        paste_btn_img_path = 'icons/Paste.png'
        self.copy_btn_img = ImageTk.PhotoImage( imageDict['Copy'] , master=book.window )
        self.paste_btn_img = ImageTk.PhotoImage( imageDict['Paste'] , master=book.window )
        
        self.copy_btn = SuperButton(parent=self.SuperMenuFrame, text='Copy', image_path=copy_btn_img_path, image=self.copy_btn_img, book=book, font='Arial 12', command=book.copy, compound='bottom')
        self.paste_btn = SuperButton(parent=self.SuperMenuFrame, text='Paste', image_path=paste_btn_img_path, image=self.paste_btn_img, book=book, font='Arial 12', command=book.paste, compound='bottom')
        self.copy_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        self.paste_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator4 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator4.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
    def text_fill_format(self, book):
        # Text/Fill Formatting
        self.SuperMenu2 = SuperMenuSection(self.SuperMenuFrame, book=book)
        
        self.bold_btn = SuperButton(parent=self.SuperMenu2, text='B', book=book, font='Arial 12 bold', command=book.bold)
        self.italic_btn = SuperButton(parent=self.SuperMenu2, text='I', book=book, font='{Times New Roman} 12 italic', command=book.italic)
        self.underline_btn = SuperButton(parent=self.SuperMenu2, text='U', book=book, font='Arial 12 underline', command=book.underline)
        
        TextColor_btn_img_path = 'icons/TextColor.png'
        FillColor_btn_img_path = 'icons/FillColor.png'
        
        self.TextColor_btn_imgPIL = imageDict['TextColor'].convert('RGBA')
        self.FillColor_btn_imgPIL = imageDict['FillColor'].convert('RGBA')
        
        self.TextColor_btn_img = ImageTk.PhotoImage( self.TextColor_btn_imgPIL , master=book.window )
        self.FillColor_btn_img = ImageTk.PhotoImage( self.FillColor_btn_imgPIL , master=book.window )
        
        self.TextColor_btn = SuperButton(parent=self.SuperMenu2, text='Text/Font Color', image_path=TextColor_btn_img_path, image=self.TextColor_btn_img, book=book, font='Arial 12', command=book.changeTextColor)
        self.FillColor_btn = SuperButton(parent=self.SuperMenu2, text='Fill Color', image_path=FillColor_btn_img_path, image=self.FillColor_btn_img, book=book, font='Arial 12', command=book.changeFillColor)
        
        column_inner, row_inner = 0, 0
        self.bold_btn.grid(row=row_inner, column=column_inner); column_inner += 1
        self.italic_btn.grid(row=row_inner, column=column_inner); column_inner += 1
        self.underline_btn.grid(row=row_inner, column=column_inner); column_inner += 1
        
        column_inner, row_inner = 0, row_inner+1
        self.FillColor_btn.grid(row=row_inner, column=column_inner); column_inner += 1
        self.TextColor_btn.grid(row=row_inner, column=column_inner); column_inner += 1
        
        self.SuperMenu2.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator5 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator5.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
    def alignment_format(self, book):
        self.SuperMenu3 = SuperMenuSection(self.SuperMenuFrame, book=book)
        
        self.alignmentHBtns = {}
        self.alignmentVBtns = {}
        self.alignmentLabel = SuperLabel(parent=self.SuperMenu3, text='Alignment', font='Arial 10', book=book)
        self.alignmentLabel.grid(row=0, column=0, columnspan=3)
        
        toggleAlignTop_btn_img_path = 'icons/AlignTop.png'
        self.toggleAlignTop_btn_img = ImageTk.PhotoImage( imageDict['AlignTop'] , master=book.window )
        self.toggleAlignTop_btn = self.alignmentVBtns['top'] = SuperButton(parent=self.SuperMenu3, text='Align Top', image_path=toggleAlignTop_btn_img_path, image=self.toggleAlignTop_btn_img, book=book, font='Arial 12', command=book.toggleAlignTop, to_end=False, vertical=True)
        self.toggleAlignTop_btn.grid(row=1, column=0)
        
        toggleAlignMiddleV_btn_img_path = 'icons/AlignCenterV.png'
        self.toggleAlignMiddleV_btn_img = ImageTk.PhotoImage( imageDict['AlignCenterV'] , master=book.window )
        self.toggleAlignMiddleV_btn = self.alignmentVBtns['center'] = SuperButton(parent=self.SuperMenu3, text='Align Middle', image_path=toggleAlignMiddleV_btn_img_path, image=self.toggleAlignMiddleV_btn_img, book=book, font='Arial 12', command=book.toggleAlignMiddleV, to_end=False, vertical=True)
        self.toggleAlignMiddleV_btn.grid(row=1, column=1)
        
        toggleAlignBottom_btn_img_path = 'icons/AlignBottom.png'
        self.toggleAlignBottom_btn_img = ImageTk.PhotoImage( imageDict['AlignBottom'] , master=book.window )
        self.toggleAlignBottom_btn = self.alignmentVBtns['bottom'] = SuperButton(parent=self.SuperMenu3, text='Align Bottom', image_path=toggleAlignBottom_btn_img_path, image=self.toggleAlignBottom_btn_img, book=book, font='Arial 12', command=book.toggleAlignBottom, to_end=False, vertical=True)
        self.toggleAlignBottom_btn.grid(row=1, column=2)
        
        toggleAlignLeft_btn_img_path = 'icons/AlignLeft.png'
        self.toggleAlignLeft_btn_img = ImageTk.PhotoImage( imageDict['AlignLeft'] , master=book.window )
        self.toggleAlignLeft_btn = self.alignmentHBtns['left'] = SuperButton(parent=self.SuperMenu3, text='Align Left', image_path=toggleAlignLeft_btn_img_path, image=self.toggleAlignLeft_btn_img, book=book, font='Arial 12', command=book.toggleAlignLeft, to_end=False, vertical=True)
        self.toggleAlignLeft_btn.grid(row=2, column=0)
        
        toggleAlignCenterH_btn_img_path = 'icons/AlignCenterH.png'
        self.toggleAlignCenterH_btn_img = ImageTk.PhotoImage( imageDict['AlignCenterH'] , master=book.window )
        self.toggleAlignCenterH_btn = self.alignmentHBtns['center'] = SuperButton(parent=self.SuperMenu3, text='Align Center', image_path=toggleAlignCenterH_btn_img_path, image=self.toggleAlignCenterH_btn_img, book=book, font='Arial 12', command=book.toggleAlignCenterH, to_end=False, vertical=True)
        self.toggleAlignCenterH_btn.grid(row=2, column=1)
        
        toggleAlignRight_btn_img_path = 'icons/AlignRight.png'
        self.toggleAlignRight_btn_img = ImageTk.PhotoImage( imageDict['AlignRight'] , master=book.window )
        self.toggleAlignRight_btn = self.alignmentHBtns['right'] = SuperButton(parent=self.SuperMenu3, text='Align Right', image_path=toggleAlignRight_btn_img_path, image=self.toggleAlignRight_btn_img, book=book, font='Arial 12', command=book.toggleAlignRight, to_end=False, vertical=True)
        self.toggleAlignRight_btn.grid(row=2, column=2)
        
        self.SuperMenu3.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator6 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator6.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
    def toggle_border(self, book):
        # Toggle Border
        self.SuperMenu4 = SuperMenuSection(self.SuperMenuFrame, book=book)
        
        self.toggleBorderLabel = SuperLabel(parent=self.SuperMenu4, text='Toggle Border', font='Arial 10', book=book)
        self.toggleBorderLabel.grid(row=0, column=0, columnspan=4)
        
        toggleBorderLeft_btn_img_path = 'icons/BorderLeft.png'
        self.toggleBorderLeft_btn_img = ImageTk.PhotoImage( imageDict['BorderLeft'] , master=book.window )
        self.toggleBorderLeft_btn = SuperButton(parent=self.SuperMenu4, text='Left Border', image_path=toggleBorderLeft_btn_img_path, image=self.toggleBorderLeft_btn_img, book=book, font='Arial 12', command=book.toggleBorderLeft, to_end=False, vertical=True)
        self.toggleBorderLeft_btn.grid(row=1, column=0)
        
        toggleBorderRight_btn_img_path = 'icons/BorderRight.png'
        self.toggleBorderRight_btn_img = ImageTk.PhotoImage( imageDict['BorderRight'] , master=book.window )
        self.toggleBorderRight_btn = SuperButton(parent=self.SuperMenu4, text='Right Border', image_path=toggleBorderRight_btn_img_path, image=self.toggleBorderRight_btn_img, book=book, font='Arial 12', command=book.toggleBorderRight, to_end=True, vertical=True)
        self.toggleBorderRight_btn.grid(row=1, column=1)
        
        toggleBorderTop_btn_img_path = 'icons/BorderTop.png'
        self.toggleBorderTop_btn_img = ImageTk.PhotoImage( imageDict['BorderTop'] , master=book.window )
        self.toggleBorderTop_btn = SuperButton(parent=self.SuperMenu4, text='Top Border', image_path=toggleBorderTop_btn_img_path, image=self.toggleBorderTop_btn_img, book=book, font='Arial 12', command=book.toggleBorderTop, to_end=False, vertical=False)
        self.toggleBorderTop_btn.grid(row=1, column=2)
        
        toggleBorderBottom_btn_img_path = 'icons/BorderBottom.png'
        self.toggleBorderBottom_btn_img = ImageTk.PhotoImage( imageDict['BorderBottom'] , master=book.window )
        self.toggleBorderBottom_btn = SuperButton(parent=self.SuperMenu4, text='Bottom Border', image_path=toggleBorderBottom_btn_img_path, image=self.toggleBorderBottom_btn_img, book=book, font='Arial 12', command=book.toggleBorderBottom, to_end=True, vertical=False)
        self.toggleBorderBottom_btn.grid(row=1, column=3)
        
        toggleBorderAll_btn_img_path = 'icons/BorderAll.png'
        self.toggleBorderAll_btn_img = ImageTk.PhotoImage( imageDict['BorderAll'] , master=book.window )
        self.toggleBorderAll_btn = SuperButton(parent=self.SuperMenu4, text='All Borders', image_path=toggleBorderAll_btn_img_path, image=self.toggleBorderAll_btn_img, book=book, font='Arial 12', command=book.toggleBorderAll, to_end=True, vertical=False)
        self.toggleBorderAll_btn.grid(row=2, column=0)
        
        toggleBorderOuter_btn_img_path = 'icons/BorderOuter.png'
        self.toggleBorderOuter_btn_img = ImageTk.PhotoImage( imageDict['BorderOuter'] , master=book.window )
        self.toggleBorderOuter_btn = SuperButton(parent=self.SuperMenu4, text='Outer Borders', image_path=toggleBorderOuter_btn_img_path, image=self.toggleBorderOuter_btn_img, book=book, font='Arial 12', command=book.toggleBorderOuter, to_end=True, vertical=False)
        self.toggleBorderOuter_btn.grid(row=2, column=1)
        
        toggleBorderInner_btn_img_path = 'icons/BorderInner.png'
        self.toggleBorderInner_btn_img = ImageTk.PhotoImage( imageDict['BorderInner'] , master=book.window )
        self.toggleBorderInner_btn = SuperButton(parent=self.SuperMenu4, text='Inner Borders', image_path=toggleBorderInner_btn_img_path, image=self.toggleBorderInner_btn_img, book=book, font='Arial 12', command=book.toggleBorderInner, to_end=True, vertical=False)
        self.toggleBorderInner_btn.grid(row=2, column=2)
        
        toggleBorderNone_btn_img_path = 'icons/BorderNone.png'
        self.toggleBorderNone_btn_img = ImageTk.PhotoImage( imageDict['BorderNone'] , master=book.window )
        self.toggleBorderNone_btn = SuperButton(parent=self.SuperMenu4, text='No Borders', image_path=toggleBorderNone_btn_img_path, image=self.toggleBorderNone_btn_img, book=book, font='Arial 12', command=book.toggleBorderNone, to_end=True, vertical=False)
        self.toggleBorderNone_btn.grid(row=2, column=3)
        
        self.SuperMenu4.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.toggleBorderBtns = [self.toggleBorderLeft_btn, self.toggleBorderRight_btn, self.toggleBorderTop_btn, self.toggleBorderBottom_btn]
        
    def insert_delete_row_column(self, book):
        # Insert Row/Column
        self.SuperMenu5 = SuperMenuSection(self.SuperMenuFrame, book=book)
        
        insertColLeft_btn_img_path = 'icons/InsertColLeft.png'
        insertColRight_btn_img_path = 'icons/InsertColRight.png'
        insertRowAbove_btn_img_path = 'icons/InsertRowAbove.png'
        insertRowBelow_btn_img_path = 'icons/InsertRowBelow.png'
        
        self.insertColLeft_btn_img = ImageTk.PhotoImage( imageDict['InsertColLeft'] , master=book.window )
        self.insertColRight_btn_img = ImageTk.PhotoImage( imageDict['InsertColRight'] , master=book.window )
        self.insertRowAbove_btn_img = ImageTk.PhotoImage( imageDict['InsertRowAbove'] , master=book.window )
        self.insertRowBelow_btn_img = ImageTk.PhotoImage( imageDict['InsertRowBelow'] , master=book.window )
        
        self.insertColRowLabel = SuperLabel(parent=self.SuperMenu5, text='Insert & Delete Row(s) & Column(s)', font='Arial 10', book=book)
        
        self.insertColLeft_btn = SuperButton(parent=self.SuperMenu5, text='Insert Column(s) Left', image_path=insertColLeft_btn_img_path, image=self.insertColLeft_btn_img, book=book, font='Arial 12', command=book.insertColLeft)
        self.insertColRight_btn = SuperButton(parent=self.SuperMenu5, text='Insert Column(s) Right', image_path=insertColRight_btn_img_path, image=self.insertColRight_btn_img, book=book, font='Arial 12', command=book.insertColRight)
        self.insertRowAbove_btn = SuperButton(parent=self.SuperMenu5, text='Insert Row(s) Above', image_path=insertRowAbove_btn_img_path, image=self.insertRowAbove_btn_img, book=book, font='Arial 12', command=book.insertRowAbove)
        self.insertRowBelow_btn = SuperButton(parent=self.SuperMenu5, text='Insert Row(s) Below', image_path=insertRowBelow_btn_img_path, image=self.insertRowBelow_btn_img, book=book, font='Arial 12', command=book.insertRowBelow)
        self.SuperMenu5.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.insertColRowLabel.grid(row=0, column=0, columnspan=4)
        self.insertColLeft_btn.grid(row=1, column=0)
        self.insertColRight_btn.grid(row=1, column=1)
        self.insertRowAbove_btn.grid(row=2, column=0)
        self.insertRowBelow_btn.grid(row=2, column=1)
        
        self.separator7 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator7.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        # Delete Row/Column
        deleteRow_btn_img_path = 'icons/DeleteRow.png'
        deleteColumn_btn_img_path = 'icons/DeleteCol.png'
        
        self.deleteRow_btn_img = ImageTk.PhotoImage( imageDict['DeleteRow'] , master=book.window )
        self.deleteColumn_btn_img = ImageTk.PhotoImage( imageDict['DeleteCol'] , master=book.window )
        
        self.deleteRow_btn = SuperButton(parent=self.SuperMenu5, text='Delete Row(s)', image_path=deleteRow_btn_img_path, image=self.deleteRow_btn_img, book=book, font='Arial 12', command=book.deleteRow)
        self.deleteColumn_btn = SuperButton(parent=self.SuperMenu5, text='Delete Column(s)', image_path=deleteColumn_btn_img_path, image=self.deleteColumn_btn_img, book=book, font='Arial 12', command=book.deleteColumn)
        
        self.SuperMenu5.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.deleteRow_btn.grid(row=1, column=2)
        self.deleteColumn_btn.grid(row=2, column=2)
        
        self.separator8 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator8.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
    def formula_part(self, book):
        # formula_part
        self.ColFormula_btn = SuperButton(parent=self.SuperMenuFrame, text='Column\nFormula', book=book, font='Arial 14', command=book.columnFormulaStartAction )
        self.ColFormula_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator9 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator9.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        self.Chart_btn = SuperButton(parent=self.SuperMenuFrame, text='Make\nChart', book=book, font='Arial 12', command=book.makeChart )
        self.Chart_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator10 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator10.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        self.HeaderSelectToggle_btn = SuperButton(parent=self.SuperMenuFrame, text='Header\nSelection:\nColumns', book=book, font='Arial 12', command=book.headerSelectToggle )
        self.HeaderSelectToggle_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator11 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator11.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        self.QuantumToggle_btn = SuperButton(parent=self.SuperMenuFrame, text='Quantum\nMode:\nOff', book=book, font='Arial 12', command=book.quantumToggle )
        self.QuantumToggle_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator12a = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator12a.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        self.separator12b = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator12b.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        self.separator12c = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator12c.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        self.CheckUpdates_btn = SuperButton(parent=self.SuperMenuFrame, text='Get New\nUpdates', book=book, font='Arial 12', command=book.getNewUpdates )
        self.CheckUpdates_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator13 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator13.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        self.QuAppsWeb_btn = SuperButton(parent=self.SuperMenuFrame, text='Quantum\nApps\nWebsite', book=book, font='Arial 12', command=book.quAppsWebsite )
        self.QuAppsWeb_btn.grid(row=0, column=self.column_grid); self.column_grid += 1
        
        self.separator14 = ttk.Separator(self.SuperMenuFrame, orient='vertical')
        self.separator14.grid(row=0, column=self.column_grid, sticky="ns"); self.column_grid += 1
        
        self.pack(fill='both', expand=True)
        self.window_create('end', window=self.SuperMenuFrame, stretch=1)
        HScrollBar(self)

class Book:
    init = True
    ready = False
    sheet = None
    colTitles_select_mode = False
    quantum_mode = False
    quantum_mode_display = 'Off'
    renameSheetStatus = False
    # window = tk.Tk()
    extension = ''
    sheet_relwidth = 0.989
    filetypes = (
    ("QuTable files", ("*.qutable") ),
    ("Microsoft Excel File Format", ("*.xlsx", ".xlsm", ".xltx", ".xltm") ),
    ("CSV (Comma-Separated Values)", ("*.csv") ),
    ("All Files", "*.*")
    )
    
    def __init__(self, app, newID=1, experimental=False, activate=False, file_addr=None, imported=False):
        self.book = self
        self.app = app
        self.app.booksList.append(self)
        
        self.sheetsDict = {}
        self.structData = {}
        
        self.newID = newID
        
        self.file_addr = file_addr
        self.file_dir = os.path.split(file_addr)[0] if file_addr else os.path.join(os.path.expanduser('~'), 'Documents')
        
        self.experimental = experimental
        self.bookOpenPyxl = openpyxl.Workbook()
        self.bookOpenPyxl.remove(self.bookOpenPyxl.active)
        
        # print('activate, experimental')
        # print(activate, experimental)
        # print()
        
        # if not (activate or experimental):
            # return
        
        
        window = self.window = tk.Tk()
        window.geometry('1500x900+150+30')
        
        self.icon = ImageTk.PhotoImage(image=imageDict['QuTableIcon'] , master=window )
        window.iconphoto(True, self.icon)
        
        window.pack_propagate(0)
        self.maximize()
        
        window.bind('<Control-X>', self.cut )
        window.bind('<Control-C>', self.copy )
        window.bind('<Control-V>', self.paste )
        window.bind('<Delete>', self.clear )

        window.bind('<Control-O>', lambda event: self.openBook() )
        window.bind('<Control-S>', lambda event: self.saveBook() )
        window.protocol("WM_DELETE_WINDOW", self.close_window_clicked)
        
        window.configure(bg='#FFFFFF')
        window.option_add('*background', '#FFFFFF')
        
        
        self.bookPanes = bookPanes = tk.PanedWindow(self.window, orient='vertical', sashwidth=10, background='#FFFFFF')
        bookPanes.pack(fill='both')
        
        
        self.TopMenuFrame = tk.Frame(bookPanes)
        self.SuperMenuWidget = SuperMenu(parent=self.TopMenuFrame, book=self)
        for widget in vars(self.SuperMenuWidget).values():
            if is_tk_widget(widget):
                widget.bind('<1>', lambda event: self.sheet.focus_set() )
        bookPanes.add(self.TopMenuFrame, height=125)
        
        self.sheetToggleText = tk.Text(bookPanes, state='disabled', cursor='arrow', background='#FFFFFF')
        self.sheetToggleFrame = tk.Frame(self.sheetToggleText, background='#FFFFFF')
        self.sheetToggleText.pack(fill='both', expand=True)
        self.sheetToggleText.window_create('end', window=self.sheetToggleFrame, stretch=1)
        HScrollBar(self.sheetToggleText)
        bookPanes.add(self.sheetToggleText, height=81)
        
        self.currentCellSel_Formula = tk.Frame(bookPanes, highlightthickness=1, highlightbackground='#000000', background='#d3d3d3')
        self.currentCellSel_Formula.grid_rowconfigure(0, weight=1)
        
        self.book.currentCellSelector = None
        
        # self.book.currentCellSelector = tk.Label(self.currentCellSel_Formula, anchor='w', background='#FFFFFF')
        # self.book.currentCellSelector.pack(fill='both', expand=True)
        
        bookPanes.add(self.currentCellSel_Formula, height=25)
        
        
        self.sheetOuter0 = tk.Frame(bookPanes, background='#FFFFFF')
        self.sheetOuter1 = tk.Frame(self.sheetOuter0, background='#FFFFFF')
        self.sheetOuter1.pack(fill='both', expand=True)
        bookPanes.add(self.sheetOuter0, height=500)
        
        self.QStatsSheet = tk.Frame(bookPanes, highlightthickness=1, highlightbackground='#000000', background='#FFFFFF')
        bookPanes.add(self.QStatsSheet, height=250)
        
        self.sheets = []
        self.QStatses = []
        self.sheetToggleButtons = []
        
        self.sheetNewButton = tk.Button(self.sheetToggleFrame, text='+', command=self.createNewSheet, font=('Arial', 16))
        self.sheetNewButton.bind('<1>', lambda event: self.sheet.focus_set() )
        
        self.columnFormula()
        self.formulaEntryReturnStatus = False
        
        if experimental:
            self.openBook(file_addr=file_addr_experimental)
            
        else:
            if file_addr:
                filename = os.path.split(file_addr)[1]
                window.title(f'{filename} - QuTable')
                self.title = filename
            else:
                self.title = f'Untitled{self.newID}'
                window.title(f'{self.title} - QuTable')
                if not imported:
                    self.createNewSheet()
                    self.lastCmd()
                    self.sheetNewButton.pack(side='left')
                    self.init = False
                    self.ready = True
                    window.after(100, lambda window=self.window: self.experimental_warning(window) )
                    if getattr(sys, 'frozen', False):
                        window.mainloop()
            
            
    def lastCmd(self):
        self.app.booksDict[self.title] = self
        for sheet in self:
            sheet.place(relwidth=self.sheet_relwidth, relheight=1.0)
        self.window.bind('<Configure>', self.on_window_configure )
            
        
    # @property
    # def extension(self):
        # return self._extension
    # @extension.setter
    # def extension(self, new_extension):
        # self._extension = new_extension
        
    def on_window_configure(self, event):
        for sheet in self:
            if self.window.state() == 'normal':
                sheet.place(relwidth=1.0-21/(event.width), relheight=1.0)
            elif self.window.state() == 'zoomed':
                sheet.place(relwidth=self.sheet_relwidth, relheight=1.0)
        
    def createSelectorFormulaEntry(self):
        column_grid = 0
        
        # Separator Frame Column
        self.formulaEntrySep1 = tk.Frame(self.currentCellSel_Formula, width=5, borderwidth=0, background='#d3d3d3')
        self.formulaEntrySep1.grid(row=0, column=column_grid, sticky="nsew"); column_grid += 1
        
        self.currentCellSelector = SuperEntry(self, self.currentCellSel_Formula, focusOutFuncExt=self.currentCellSelectorFuncOut, tooltipText='Current Cell\nSelection', background='#FFFFFF', font='Arial 12', borderwidth=0, width=12)
        self.currentCellSelector.grid(row=0, column=column_grid, sticky="nsew")
        # self.currentCellSelector.bind("<FocusIn>", self.currentCellSelectorFuncIn )
        # self.currentCellSelector.bind("<FocusOut>", lambda event: self.sheet.focus_set() )
        # self.currentCellSelector.bind("<FocusOut>", lambda event: self.currentCellSelector.configure(takefocus=0) )
        self.currentCellSel_Formula.grid_columnconfigure(column_grid, weight=1); column_grid += 1
        
        # Separator Frame Column
        self.formulaEntrySep2 = tk.Frame(self.currentCellSel_Formula, width=20, borderwidth=0, background='#d3d3d3')
        self.formulaEntrySep2.grid(row=0, column=column_grid, sticky="nsew"); column_grid += 1
        
        self.formulaEntry = SuperEntry(self, self.currentCellSel_Formula, focusOutFuncExt=self.formulaEntryFuncOut, tooltipText='Formula Entry', background='#FFFFFF', font='Arial 12', borderwidth=0)
        # self.formulaEntry.bind("<Return>", self.formulaEntryFuncOut )
        self.formulaEntry.grid(row=0, column=column_grid, sticky="nsew")
        # self.formulaEntry.bind("<FocusOut>", lambda event: self.sheet.focus_set() )
        # self.formulaEntry.bind("<FocusOut>", lambda event: self.formulaEntry.configure(takefocus=0) )
        self.currentCellSel_Formula.grid_columnconfigure(column_grid, weight=99); column_grid += 1
        
        # Separator Frame Column
        self.formulaEntrySep3 = tk.Frame(self.currentCellSel_Formula, width=20, borderwidth=0, background='#d3d3d3')
        self.formulaEntrySep3.grid(row=0, column=column_grid, sticky="nsew"); column_grid += 1
        
        self.formulaEntry.delete(0, 'end')
        self.formulaEntry.insert(0, self.sheet.currentCell.formula)
        
        selectedColsIndex, selectedRowsIndex = self.sheet.selectedCellsSet.generateIndexList()
        self.sheet.modify_currentCellSelReader(selectedColsIndex, selectedRowsIndex)
        
    def currentCellSelectorFuncOut(self, event=None):
        index = self.currentCellSelector.get()
        self.sheet.focus_set()
        if ':' in index:
            indicesRaw = index.split(',')
            for indexRaw in indicesRaw:
                cellStartIndexCode, cellStopIndexCode = indexRaw.split(':')
                self.sheet[cellStartIndexCode].selectbyDragging(widgetEnd=self.sheet[cellStopIndexCode])
            return
        elif ',' in index:
            indicesRaw = index.split(',')
            for indexRaw in indicesRaw:
                self.sheet[indexRaw].toggleCell(select=True)
            return
        self.sheet[index].focusOnCell()
        # self.sheet[index].toggleCellColor(select=True)
        
    def formulaEntryFuncOut(self, event=None):
        formula = self.formulaEntry.get()
        self.formulaEntry.delete(0, 'end')
        formulaEntryNOTReturnStatus = event.keysym != 'Return'
        print(8694, formula, formulaEntryNOTReturnStatus)
        self.sheet.currentCell.formula = formula, formulaEntryNOTReturnStatus
        self.sheet.focus_set()
        
    # def createQStats(self, sheet):
        # QStats = ttk.Treeview(self.QStatsSheet, columns=list(range(sheet.nCols+1)), show='headings')
        # sheet.QStats = QStats
        # HScrollBar(QStats)
        # QStats.column(0, anchor='center')
        # QStats.heading(0, text='')
        # for column, colIndex in zip( sheet , range(1, sheet.nCols+1) ):
            # QStats.column(colIndex, anchor='center')
            # QStats.heading(colIndex, text=column)
        # self.QStatses += [QStats]
        # QStats.place(relwidth=1.0, relheight=1.0)
        
        # QStatsData = sheet.getStatsReport(fancyReportType=True)
        
        # for StatsIndex in range(len(sheet.QStatsType)):
            # stats = list(QStatsData.loc[StatsIndex])
            # QStats.insert('', tk.END, values=stats)
        
    def createNewSheet(self, newSheetName=None, nCols=1, nRows=5):
        self.sheetNames = [sheet.sheetName for sheet in self.sheets]
        
        i = 0
        if not newSheetName:
            while 1:
                newSheetName=f'NewSheet{i}'
                if newSheetName not in self.sheetNames:
                    sheet = Sheet(parent=self.sheetOuter1, book=self, imported=False, sheetName=newSheetName, nCols=nCols, nRows=nRows, QStatsType=None)
                    sheet.createQStats()
                    break
                else:
                    i += 1
        else:
            sheet = Sheet(parent=self.sheetOuter1, book=self, imported=False, sheetName=newSheetName, nCols=nCols, nRows=nRows, QStatsType=None)
            sheet.createQStats()
        
        self.sheets += [sheet]
        self.sheetNewButton.pack_forget()
        
        sheetToggleButton = SheetToggleButton(sheet=sheet, book=self, parent=self.sheetToggleFrame, text=newSheetName, sheetBtnNum=len(self)-1, font=('Arial', 14) )
        self.sheetToggleButtons += [sheetToggleButton]
        sheet.place(relwidth=self.sheet_relwidth, relheight=1.0)
        sheetToggleButton.toggleTheSheet()
        
        self.sheetNewButton.pack(side='left')
        
    def menu_config(self, menuButtonFrame):
        pass
        
    def maximize(self):
        self.window.state('zoomed')
        
    def minimize(self):
        self.window.withdraw()
            
    # def get_text_height(self, text, font):
        # return get_text_height(self.window, text, font)
        
    def close(self_book, close_usual=True):
        global self
        self_book.minimize()
        for sheet in self_book:
            sheet.graphPlotter.master.withdraw()
        self_book.window.destroy()
        self_book.app.booksList.remove(self_book)
        del self_book.app.booksDict[self_book.title]
        if andExec(not self_book.app.booksList, close_usual):
            self_book.close_all()
            # print(9111, close_usual, file=open('close_usual.log', 'w'))
        elif self_book.app.booksList:
            self = self_book.app[0,0]
        
    def close_window_clicked(self):
        # If Book is empty and not saved
        if andExec(set([cell.cget('text') for sheet in self.sheets for cell in sheet.cells]) == {''}, not self.file_addr):
            self.close()
            return
        save_yes_no = messagebox.askyesnocancel('Save Document Confirmation', 'Would you like to save this Document?', parent=self.window)
        if save_yes_no:
            self.saveBook(notify=False)
            self.close()
        elif save_yes_no is None:
            pass
        elif not save_yes_no:
            self.close()
            
    def close_all(self):
        local_app_data_dir = os.getenv('LOCALAPPDATA')
        qu_table_folder = os.path.join(local_app_data_dir, "QuTable")
        qu_table_exit_file = os.path.join(qu_table_folder, "QuTable-Exit.exe")
        self.qu_table_folder = qu_table_folder
        self.qu_table_exit_file = qu_table_exit_file
        os.makedirs(qu_table_folder, exist_ok=True)
        
        qu_table_exit_file_cont = open(qu_table_exit_file, 'rb').read() if os.path.isfile(qu_table_exit_file) else None
        
        self.exit_file_cont = QuTableExitContent
        
        # Rename the Folder if that same path is a Folder instead of a File
        if os.path.exists(qu_table_exit_file):
            if not os.path.isfile(qu_table_exit_file):
                file_addr, extension = os.path.splitext(qu_table_exit_file)
                extension = extension.lower()
                i = 1
                while 1:
                    newFileAddr = qu_table_exit_file.replace(file_addr, f'{file_addr}-{i}')
                    if not os.path.exists(newFileAddr):
                        break
                    i += 1
                os.rename(qu_table_exit_file, newFileAddr)
        
        # Check whether the File exists and, if exists, whether the File contain the proper QuTableExitContent binary codes
        if not andExec(os.path.isfile(qu_table_exit_file), qu_table_exit_file_cont == self.exit_file_cont):
            os.makedirs(qu_table_folder, exist_ok=True)
            exitIO = open(qu_table_exit_file, 'wb')
            exitIO.write(self.exit_file_cont)
            exitIO.close()
            
        os.startfile(qu_table_exit_file)
        
        # sys.exit()
        # exit()
        # subprocess.run(['taskkill', '/F', '/IM', 'QuTable.exe'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE, creationflags=subprocess.CREATE_NO_WINDOW)
        
    def columnFormula(self):
        self.columnFormulaWidget = tk.Frame(self.TopMenuFrame)
        self.columnFormulaHelp = tk.Label(self.columnFormulaWidget, text='Select the column whose\nvalues are to be calculated\nor simply type the nth column\nas the letter C followed by\na number (column sequence).\nFor example, 1st column as C1,\n2nd column as C2, and so on:')
        self.columnFormulaSep1 = tk.Label(self.columnFormulaWidget, text=' '*5)
        self.columnFormulaInput = tk.Text(self.columnFormulaWidget, height=5)
        self.columnFormulaSep2 = tk.Label(self.columnFormulaWidget, text=' '*5)
        self.columnFormulaEditEnd = tk.Frame(self.columnFormulaWidget)
        self.columnFormulaSep3 = tk.Label(self.columnFormulaWidget, text=' '*5)
        
        self.columnFormulaSep1.grid(row=0, column=0)
        self.columnFormulaHelp.grid(row=0, column=1)
        self.columnFormulaSep2.grid(row=0, column=2)
        self.columnFormulaInput.grid(row=0, column=3)
        self.columnFormulaSep3.grid(row=0, column=4)
        self.columnFormulaEditEnd.grid(row=0, column=5)
        
        self.columnFormulaOK = tk.Button(self.columnFormulaEditEnd, text='OK', command=lambda: self.columnFormulaEndAction('OK') )
        self.columnFormulaCancel = tk.Button(self.columnFormulaEditEnd, text='Cancel', command=lambda: self.columnFormulaEndAction('Cancel') )
        self.columnFormulaOK.grid(row=0, column=0)
        self.columnFormulaCancel.grid(row=0, column=1)
        
        self.columnFormulaSep3a = tk.Label(self.columnFormulaEditEnd, text=' '*5)
        self.columnFormulaToggleWindow = tk.Button(self.columnFormulaEditEnd, text='Popup in\nNew Window', command=self.columnFormulaPopup )
        self.columnFormulaSep3a.grid(row=1, column=0, columnspan=2)
        self.columnFormulaToggleWindow.grid(row=2, column=0, columnspan=2)
        
    def columnFormulaPopupOld(self):
        self.columnFormulaEndAction(is_popup=True)
        self.columnFormulaWidget = tk.Toplevel(self.window, bg='#FFFFFF')
        self.columnFormulaWidget.title(f'Edit Column C{self.sheet.columnFormulaEdit} Formula')
        self.columnFormulaWidget.resizable(False, False)
        self.columnFormulaWidget.geometry('+425+415')
        # self.columnFormulaWidget.attributes('-toolwindow', True)
        self.columnFormulaSep1 = tk.Label(self.columnFormulaWidget, text=' '*5, bg='#FFFFFF')
        self.columnFormulaHelp = tk.Label(self.columnFormulaWidget, text='Select the column whose values are to be calculated or simply type the nth column as the letter C followed by a number (column sequence).\nFor example, 1st column as C1, 2nd column as C2, and so on:', bg='#FFFFFF')
        self.columnFormulaInput = tk.Text(self.columnFormulaWidget, height=5)
        self.columnFormulaInput.bind('<Return>', lambda event: self.columnFormulaEndAction('OK') )
        self.columnFormulaEditEnd = tk.Frame(self.columnFormulaWidget, bg='#FFFFFF')
        self.columnFormulaSep2 = tk.Label(self.columnFormulaWidget, text=' '*5, bg='#FFFFFF')
        self.columnFormulaSep3 = tk.Label(self.columnFormulaWidget, text=' '*5, bg='#FFFFFF')
        
        self.columnFormulaSep1.grid(row=0, column=0)
        self.columnFormulaHelp.grid(row=0, column=1)
        self.columnFormulaInput.grid(row=1, column=1)
        self.columnFormulaEditEnd.grid(row=2, column=1)
        self.columnFormulaSep2.grid(row=0, column=2)
        self.columnFormulaSep3.grid(row=3, column=0)
        
        self.columnFormulaOK = tk.Button(self.columnFormulaEditEnd, text='OK', command=lambda: self.columnFormulaEndAction(status='OK'), bg='#FFFFFF' )
        self.columnFormulaCancel = tk.Button(self.columnFormulaEditEnd, text='Cancel', command=lambda: self.columnFormulaEndAction(status='Cancel'), bg='#FFFFFF' )
        self.columnFormulaOK.grid(row=0, column=0)
        self.columnFormulaCancel.grid(row=0, column=1)
        
        self.columnFormulaSep3a = tk.Label(self.columnFormulaEditEnd, text=' '*5, bg='#FFFFFF')
        self.columnFormulaToggleWindow = tk.Button(self.columnFormulaEditEnd, text='Put on Top Menu Bar', command=self.columnFormulaPutOnTop, bg='#FFFFFF' )
        self.columnFormulaSep3a.grid(row=1, column=0, columnspan=2)
        self.columnFormulaToggleWindow.grid(row=2, column=0, columnspan=2)
        
    def columnFormulaPopup(self):
        self.columnFormulaEndAction(is_popup=True)
        self.columnFormulaWidget = FormulaWizard(self, self.sheet, bg='#FFFFFF')
        
    def columnFormulaPutOnTop(self):
        self.bookPanes.paneconfigure(tagOrId=self.TopMenuFrame, height=150)
        if isinstance(self.columnFormulaWidget, tk.Toplevel):
            self.columnFormulaWidget.destroy()
        self.columnFormula()
        self.SuperMenuWidget.pack_forget()
        self.columnFormulaWidget.pack(fill='both', expand=True)
        
    def columnFormulaStartAction(self):
        messagebox.showinfo('Column Formula under Development', 'The "Column Formula" feature is still under developments. In the future, you will be able to simply put column index and then create a new column containing all the consistent formula.\nFor example, 1st column as C1, 2nd column as C2, and so on.\nFinally, type New Column Name(s) on the left for the name of new column(s) to be created.\n\nAn example would be "NewColFormula" on the left and "C1*C2" on the right. It means that "NewColFormula" would be created containing the products of the values of C1 and C2.', parent=self.window)
        return
        
        self.sheet.columnFormulaEdit = self.sheet.currentCell.index[0]
        self.columnFormulaPopup()
        
    def columnFormulaEndAction(self, status=None, is_popup=False):
        # If OK button is pressed (otherwise: Cancel button is pressed or Toggle Window)
        if status == 'OK':
            colEditIndex = self.sheet.columnFormulaEdit
            expr = self.columnFormulaInput.get('0.0', 'end')
            # if expr != '\n':
            if self.columnFormulaInput.compare('end-1c', '!=', '1.0'):
                resultColFormula = srepr( custom_parse_expr( expr ) )
                resultColData = eval( resultColFormula )
                self.sheet[colEditIndex].replaceFormulae(resultColData)
            
        # Close Separate Window Remove or columnFormulaWidget from Top Menu
        if isinstance(self.columnFormulaWidget, tk.Toplevel):
            self.columnFormulaWidget.destroy()
        else:
            self.columnFormulaWidget.pack_forget()
            
            # Set back to height of 105 px
            self.bookPanes.paneconfigure(tagOrId=self.TopMenuFrame, height=125)
            
            # Pack back Top Menu Widgets
            self.SuperMenuWidget.pack(fill='both', expand=True)
            
        # If this is called to end columnFormula edit, instead of Toggle Window
        if not is_popup:
            self.sheet.columnFormulaEdit = None
        
    def __repr__(self):
        return self.title
            
    def __str__(self):
        return self.__repr__()
        
    def __getitem__(self, index):
        if orExec(type(index) == int, type(index) == str):
            return self.sheets[index]
            
        elif type(index) == range:
            slice_range = slice(index.start, index.stop, index.step)
            return CellSet(self.book, self.sheet, self.sheets[slice_range])
            
        elif hasattr(index, '__iter__'):
            if len(index) == 1:
                return self.sheets[index[0]]
            elif len(index) == 2:
                return self.sheets[index[0]][index[1]]
            elif len(index) == 3:
                return self.sheets[index[0]][index[1]].cells[index[2]]
            
        elif type(index) == slice:
            return self.sheets[index]
        
        elif orExec(isinstance(index, ColRange), isinstance(index[1], RowRange)):
            return CellSet(self.book, self.sheet, self.sheets[index.slicer])
        
    def __iter__(self):
        for sheet in self.sheets:
            yield sheet

    def __lt__(self, other):
        return self.sheets < other.sheets

    def __gt__(self, other):
        return self.sheets > other.sheets

    def __le__(self, other):
        return self.sheets <= other.sheets

    def __ge__(self, other):
        return self.sheets >= other.sheets
        
    def __len__(self):
        return len(self.sheets)
        
    def changeTextColor(self):
        self.sheet.changeTextColor()
        
    def changeFillColor(self):
        self.sheet.changeFillColor()
        
    def cut(self, event=None):
        self.sheet.cut()
        
    def copy(self, event=None):
        self.sheet.copy()
        
    def clear(self, event=None):
        self.sheet.clear()
        self.sheet.actionCollect(action='Clear')
        
    def paste(self, event=None):
        self.sheet.paste()
        self.sheet.actionCollect(action='Paste')
        
    def bold(self):
        self.mappingFunc(CellLabel.bold)
        self.sheet.actionCollect(action='Bold')
        
    def italic(self):
        self.mappingFunc(CellLabel.italic)
        self.sheet.actionCollect(action='Italic')
        
    def underline(self):
        self.mappingFunc(CellLabel.underline)
        self.sheet.actionCollect(action='Underline')
        
    def toggleAlignTop(self):
        self.mappingFunc(CellLabel.toggleAlignTop)
        self.sheet.actionCollect(action='Toggle Align Top')
        
    def toggleAlignMiddleV(self):
        self.mappingFunc(CellLabel.toggleAlignMiddleV)
        self.sheet.actionCollect(action='Toggle Align Middle')
        
    def toggleAlignBottom(self):
        self.mappingFunc(CellLabel.toggleAlignBottom)
        self.sheet.actionCollect(action='Toggle Align Bottom')
        
    def toggleAlignLeft(self):
        self.mappingFunc(CellLabel.toggleAlignLeft)
        self.sheet.actionCollect(action='Toggle Align Left')
        
    def toggleAlignCenterH(self):
        self.mappingFunc(CellLabel.toggleAlignCenterH)
        self.sheet.actionCollect(action='Toggle Align Center')
        
    def toggleAlignRight(self):
        self.mappingFunc(CellLabel.toggleAlignRight)
        self.sheet.actionCollect(action='Toggle Align Right')
        
    def toggleBorderLeft(self):
        self.sheet.toggleBorderLeft()
        self.sheet.actionCollect(action='Toggle Border Left')
        
    def toggleBorderRight(self):
        self.sheet.toggleBorderRight()
        self.sheet.actionCollect(action='Toggle Border Right')
        
    def toggleBorderTop(self):
        self.sheet.toggleBorderTop()
        self.sheet.actionCollect(action='Toggle Border Top')
        
    def toggleBorderBottom(self):
        self.sheet.toggleBorderBottom()
        self.sheet.actionCollect(action='Toggle Border Bottom')
        
    def toggleBorderOuter(self):
        self.toggleBorderLeft()
        self.toggleBorderRight()
        self.toggleBorderTop()
        self.toggleBorderBottom()
        self.sheet.actionCollect(action='Toggle Border Outer')
        
    def toggleBorderAll(self):
        self.sheet.toggleBorderAll()
        self.sheet.actionCollect(action='Toggle Border All')
        
    def toggleBorderInner(self):
        self.toggleBorderAll()
        self.toggleBorderOuter()
        self.sheet.actionCollect(action='Toggle Border Inner')
        
    def toggleBorderNone(self):
        self.sheet.toggleBorderNone()
        self.sheet.actionCollect(action='Toggle Border None')
        
    def insertRowAbove(self):
        self.sheet.insertRowAbove()
        self.sheet.actionCollect(action='Insert Row Above')
        
    def insertRowBelow(self):
        self.sheet.insertRowBelow()
        self.sheet.actionCollect(action='Insert Row Bottom')
        
    def insertColRight(self):
        self.sheet.insertColRight()
        self.sheet.actionCollect(action='Insert Column Right')
        
    def insertColLeft(self):
        self.sheet.insertColLeft()
        self.sheet.actionCollect(action='Insert Column Left')
        
    def deleteColumn(self):
        self.sheet.deleteColumn()
        self.sheet.actionCollect(action='Delete Column')
        
    def deleteRow(self):
        self.sheet.deleteRow()
        self.sheet.actionCollect(action='Delete Row')
        
    def makeChart(self):
        self.sheet.makeChart()
        
    def headerSelectToggle(self):
        if self.colTitles_select_mode:
            self.SuperMenuWidget.HeaderSelectToggle_btn['text'] = 'Header\nSelection:\nColumns'
            self.colTitles_select_mode = False
        else:
            self.SuperMenuWidget.HeaderSelectToggle_btn['text'] = 'Header\nSelection:\nCol Titles'
            self.colTitles_select_mode = True
        
    def getNewUpdates(self):
        os.startfile('https://quantumapps100.wixsite.com/quantumapps/try-qutable')
        
    def quAppsWebsite(self):
        os.startfile('https://quantumapps100.wixsite.com/quantumapps')
        
    def quantumToggle(self):
        if self.quantum_mode:
            self.SuperMenuWidget.QuantumToggle_btn['text'] = 'Quantum\nMode:\nOff'
            self.quantum_mode = False
        else:
            self.SuperMenuWidget.QuantumToggle_btn['text'] = 'Quantum\nMode:\nOn'
            self.quantum_mode = True
        
    def newBook(self):
        newBook = Book(app=self.app, newID=self.newID+1)
        self.app.booksList.append(newBook)
        
    def fillDecision(self, newBook, imported_data, QuTable_format, extension, bookOpenPyxl, haveHeadersList):
        if imported_data is None:
            newBook.createNewSheet()
        else:
            newBook.fill_in_data(imported_data, QuTable_format=QuTable_format, extension=extension, bookOpenPyxl=bookOpenPyxl, haveHeadersList=haveHeadersList)
        
    def openBook(self, file_addr=None):
        try:
            # Labeling the Workbook as per File Name
            file_addr, imported_data, bookOpenPyxl, QuTable_format, extension, haveHeadersList = self.importBook(file_addr=file_addr, open_status=True)
            
            if not file_addr: return
            # print('openBook after importBook, return conds, ')
            if andExec(not isinstance(imported_data, dict), imported_data != None): return
            
            if extension.startswith('.xls'):
                file_addr = None
            newBook = Book(app=self.app, newID=self.newID, activate=True, file_addr=file_addr, imported=True)
            
            self.fillDecision(newBook, imported_data, QuTable_format=QuTable_format, extension=extension, bookOpenPyxl=bookOpenPyxl, haveHeadersList=haveHeadersList)
            
            if not extension.startswith('.xls'):
                _, filename = os.path.split(file_addr)
                newBook.title = filename
                newBook.window.title(f'{newBook.title} - QuTable')
                # print(f'{newBook.title} - QuTable')
            
            newBook.lastCmd()
            newBook.toggleSheet(0)
            
            newBook.init = False
            newBook.ready = True
            newBook.window.after(100, lambda window=newBook.window: self.experimental_warning(window) )
            
            if extension.startswith('.xls'):
                newBook.window.after(100, lambda: messagebox.showwarning('Open Excel File as a New File Instead', 'Open Excel File as a New File Instead\n\nThe Excel file that you requested can only be imported and converted to QuTable ".qutable" file or CSV, instead of editing Excel file directly. Please export it through the Save As Dialog Box if you really would like to save it.', parent=newBook.window) )
                
            # If Book is empty and not saved
            if andExec(set([cell.cget('text') for sheet in self.sheets for cell in sheet.cells]) == {''}, not self.file_addr):
                # Close current window
                self.close(close_usual=False)
            
            if getattr(sys, 'frozen', False):
                newBook.window.mainloop()
            
        except Exception as e:
            with open('stderr.log', 'w') as fileIO:
                fileIO.write( format_exc() )
        
    def importBook(self, file_addr=None, open_status=False):
        if self.experimental:
            file_addr = file_addr_experimental
        else:
            file_addr = filedialog.askopenfilename(parent=self.window, initialdir=self.file_dir, title='Select a File', filetypes=self.filetypes)
        
        # If Book is empty and not saved
        if andExec(open_status, set([cell.cget('text') for sheet in self.sheets for cell in sheet.cells]) == {''}, not self.file_addr, file_addr):
            # Hide (Pre-close) current window
            self.window.withdraw()
        
        if file_addr:
            if andExec(open_status, hasattr(app, '__iter__')):
                for book in app:
                    if file_addr == book.file_addr:
                        return None, None, None, None, None, None
                
            file_addr1, extension = os.path.splitext(file_addr)
            file_dir, filename = os.path.split(file_addr1)
            extension = extension.lower()
            self.bookOpenPyxl = bookOpenPyxl = None
            QuTable_format = extension == '.qutable'
            print(9530, file_addr)
            print(9531, QuTable_format)
            
            if QuTable_format:
                self.imported_data = imported_data = decompress_byte_to_json(file_addr)
            elif extension == '.csv':
                try:
                    self.imported_data = imported_data = {filename: pd.read_csv(file_addr, encoding='utf-8')}
                except UnicodeDecodeError:
                    # Try another encoding if utf-8 fails
                    try:
                        self.imported_data = imported_data = {filename: pd.read_csv(file_addr, encoding='latin-1')}
                    except UnicodeDecodeError:
                        messagebox.showerror('Error', 'Failed to decode the CSV file. Please try another encoding.', parent=self.window)
                        raise
            elif extension.startswith('.xls'):
                self.bookOpenPyxl = bookOpenPyxl = openpyxl.load_workbook(file_addr)
                self.imported_data = imported_data = pd.read_excel(file_addr, sheet_name=None)
            else:
                messagebox.showerror('Error', 'Unsupported file format.', parent=self.window)
                return TypeError('Unsupported file format.')
            
            if not QuTable_format:
                self.haveHeadersSelector = HaveHeadersSelector(self.window, file_addr, sheets=list(imported_data) )
                self.haveHeadersSelector.wait_window()
                if not self.haveHeadersSelector.response:
                    return None, None, None, None, None, None
                haveHeadersList = self.haveHeadersSelector.selected_sheets
            else:
                haveHeadersList = None
                
            if open_status:
                return file_addr, imported_data, bookOpenPyxl, QuTable_format, extension, haveHeadersList
            else:
                self.fill_in_data(imported_data, QuTable_format=QuTable_format, bookOpenPyxl=bookOpenPyxl)
                return None, None, None, None, None, None
        else:
            return None, None, None, None, None, None
        
    def fill_in_data(self, imported_data, QuTable_format, extension, haveHeadersList, bookOpenPyxl=None):
        self.extension = extension
        if extension.startswith('.xls'):
            self.bookOpenPyxl = bookOpenPyxl
        self.book.structData = imported_data
        
        # print('haveHeadersList =', haveHeadersList)
        for (sheetName, df_raw), i in zip(imported_data.items(), range(len(imported_data))):
            sheetOpenPyxl = bookOpenPyxl [ sheetName ] if bookOpenPyxl else None
            
            if QuTable_format:
                sheet = Sheet(parent=self.sheetOuter1, book=self, imported=True, sheetName=sheetName, structData=df_raw, openpyxl_sheet=None)
            else:
                df = df_raw.fillna('')
                have_headers = sheetName in haveHeadersList
                # print('have_headers =', have_headers, 'for', sheetName)
                sheet = Sheet(parent=self.sheetOuter1, book=self, have_headers=have_headers, imported=True, sheetName=sheetName, sheetOpenPyxl=sheetOpenPyxl, df=df, openpyxl_sheet=None)
                
            self.sheets += [sheet]
            
            sheetToggleButton = SheetToggleButton(sheet=sheet, book=self, parent=self.sheetToggleFrame, sheetBtnNum=i, text=sheetName, font=('Arial', 14, 'bold') )
            self.sheetToggleButtons += [sheetToggleButton]
            
        self.sheetNames = []
        for sheet in self.sheets:
            self.sheetNames += [sheet.sheetName]
            sheet.createQStats()
            
        self.sheetNewButton.pack_forget()
        self.sheetNewButton.pack(side='left')
        
        # Close workbook application
        if bookOpenPyxl:
            bookOpenPyxl.close()
        
    def saveBook(self, notify=True):
        if self.file_addr is None:
            self.saveAsBook(notify=notify)
        else:
            filename, extension = os.path.splitext(self.file_addr)
            if extension.startswith('.xls'):
                messagebox.showwarning('Cannot edit Excel files', 'QuTable cannot edit Excel files, ONLY QuTable ".qutable" files and CSV files. Please export it instead through the Save As Dialog Box.', parent=self.window)
                self.saveAsBook(notify=notify)
            else:
                self.saveAsBook(file_addr=self.file_addr, notify=notify)
        
    def saveAsBook(self, file_addr=None, notify=True):
        if file_addr is None:
            self.window.grab_set()  # Grab focus to prevent interactions with other windows
            try:
                filename, extension = os.path.splitext(self.title)
                extension = extension.lower()
                self.fileIO = filedialog.asksaveasfile(parent=self.window, initialfile=f'{filename}.qutable', initialdir=self.file_dir, filetypes=self.filetypes, defaultextension=self.filetypes)
            finally:
                self.window.grab_release()  # Release grab even if an exception occurs
            
            if not self.fileIO: return
            self.file_addr = file_addr = self.fileIO.name
            
        if self.title in self.app.booksDict:
            del self.app.booksDict[self.title]
        self.file_dir, self.title = os.path.split(self.file_addr)
        self.window.title(f'{self.title} - QuTable')
        self.app.booksDict[self.title] = self
        
        dfDict = self.BookToDataFrame()
        filename, extension = os.path.splitext(file_addr)
        self.extension = extension = extension.lower()
        
        try:
            if self.extension == '.csv':
                if messagebox.askyesno('CSV File Save Confirmation', 'A single CSV file cannot contain multiple sheets. Instead, what QuTable can do is to save multiple files with each CSV files representing each sheet with their respective names. Finally, a new folder is created with the name that you tried to save before. Would you like to continue?', parent=self.window):
                    os.makedirs(file_addr, exist_ok=True)
                    for dfTitle, df in dfDict.items():
                        df.to_csv(f'{file_addr}/{dfTitle}.csv', index=False)
            
            elif self.extension.startswith('.xls'):
                for sheet in self:
                    sheet.convertStructDataToXLS()
                self.bookOpenPyxl.save(file_addr)
            
            elif self.extension == '.qutable':
                compress_json_to_byte(self.structData, file_addr)
            
            if notify:
                messagebox.showinfo('Successfully saved', f'The file {filename} has been successfully saved! Hooray!', parent=self.window)
            
        except PermissionError:
            messagebox.showerror('File might be opened in another application', f'The file {filename} might be opened in another application.', parent=self.window)
        

    def toggleSheet(self, sheetNum):
        specifiedSheet = self.sheets[sheetNum]
        self.sheet.scrollBar.uninstall() if self.sheet else None
        self.sheet = specifiedSheet
        self.mappingFunc = self.sheet.selectedCellsSet.mapPlus
        
        specifiedQStats = self.QStatses[sheetNum]
        specifiedSheetButton = self.sheetToggleButtons[sheetNum]
        
        for SheetButton in self.sheetToggleButtons:
            SheetButton.bind_unclicked()
            
        specifiedSheetButton.unbind_clicked()
        
        # for sheet, QStats in zip(self.sheets, self.QStatses):
            # sheet.place_forget()
            # QStats.pack_forget()
        
        self.sheet.undo_btn.lift()
        self.sheet.undo_menu_btn.lift()
        self.sheet.redo_btn.lift()
        self.sheet.redo_menu_btn.lift()
        
        specifiedSheet.lift()
        specifiedQStats.lift()
        # self.SuperMenuWidget.undo_menu_btns[specifiedSheet.sheetName].lift()
        
        self.bindSheet(specifiedSheet=specifiedSheet)
        self.sheet.scrollBar.install()
        
        if self.currentCellSelector is None:
            self.createSelectorFormulaEntry()
            
        
        # self.SuperMenuWidget.undo_menu_btns[specifiedSheet].grid(row=self.SuperMenuWidget.undo_menu_btn_row, column=self.SuperMenuWidget.undo_menu_btn_column)
        
        # specifiedSheet.modify_currentCellSelReader(sheet=specifiedSheet)
        
    def duplicateSheet(self, sheetNum=0):
        specifiedSheet = self.sheets[sheetNum]
        specifiedQStats = self.QStatses[sheetNum]
        specifiedSheetButton = self.sheetToggleButtons[sheetNum]
        
        # parent = widget.nametowidget(widget.winfo_parent())
        # type(widget)(parent)
        
        newSheet = widgetDeepCopy(specifiedSheet, type(specifiedSheet)( parent=specifiedSheet.parent , book=self ) )
        newQStats = widgetDeepCopy(specifiedQStats, type(specifiedQStats)( get_parent(specifiedQStats) ) )
        newSheetButton = SheetToggleButton(sheet=newSheet, book=self, parent=self.sheetToggleFrame, text=specifiedSheetButton.cget('text'), sheetBtnNum=len(self), font=('Arial', 14, 'bold') )
        # newSheetButton = widgetDeepCopy( specifiedSheetButton, type(specifiedSheetButton)( sheet=specifiedSheet , book=self , parent=specifiedSheetButton.parent , sheetBtnNum=len(self)-1 ) )
        
        self.sheets += [newSheet]
        self.QStatses += [newQStats]
        self.sheetToggleButtons += [newSheetButton]
        
        self.sheetNewButton.pack_forget()
        self.sheetNewButton.pack(side='left')
        
        newSheetButton.unbind_clicked()
        
        specifiedSheetButton['background'] = '#FFFFFF'
        specifiedSheetButton.activate()
        specifiedSheetButton.configure(font=('Arial', 14) )
        
    def deleteSheet(self, sheetNum=0):
        specifiedSheet = self.sheets[sheetNum]
        specifiedQStats = self.QStatses[sheetNum]
        specifiedSheetButton = self.sheetToggleButtons[sheetNum]
        
        specifiedSheet.destroy()
        specifiedQStats.destroy()
        specifiedSheetButton.destroy()
        
        subsequentSheetButtons = self.sheetToggleButtons[sheetNum+1:]
            
        # for sheet, i in zip ( subsequentSheets, range(sheetNum+1,+sheetNum+1) ):
            # sheet.configure( command = lambda x=i: self.toggleSheet(x) )
        
        del self.sheetsDict[specifiedSheet.sheetName]
        self.sheets.remove( specifiedSheet )
        self.QStatses.remove( specifiedQStats )
        self.sheetToggleButtons.remove( specifiedSheetButton )
        
        for subsequentSheetButton in subsequentSheetButtons:
            subsequentSheetButton.sheetBtnNum -= 1
            subsequentSheetButton.activate()
            subsequentSheetButton.toggleTheSheet()
        
        self.sheetToggleButtons[sheetNum-1].toggleTheSheet()
        
    def renameSheet(self, sheetNum=0):
        self.renameSheetStatus = True
        specifiedSheet = self.sheets[sheetNum]
        specifiedQStats = self.QStatses[sheetNum]
        specifiedSheetButton = self.sheetToggleButtons[sheetNum]
        
        
        x = specifiedSheetButton.winfo_x()
        y = specifiedSheetButton.winfo_y()
        
        width = specifiedSheetButton.winfo_width()
        height = specifiedSheetButton.winfo_height()
        
        # Moving across cells using keyboard arrows
        self.book.window.unbind('<Key>')
        
        sheetNameEntry = tk.Entry(self.sheetToggleFrame)
        
        def renameSheetOK(event):
            del self.sheetsDict[specifiedSheet.sheetName]
            new_name = sheetNameEntry.get()
            specifiedSheetButton['text'] = specifiedSheet.sheetName = new_name
            self.sheetsDict[new_name] = specifiedSheet
            sheetNameEntry.destroy()
            self.renameSheetStatus = False
            
        def renameSheetCancel(event):
            new_name = sheetNameEntry.get()
            specifiedSheetButton['text'] = new_name
            sheetNameEntry.destroy()
            self.renameSheetStatus = False
        
        self.renameSheetOK = renameSheetOK
        self.renameSheetCancel = renameSheetCancel
        
        sheetNameEntry.bind('<Return>', renameSheetOK)
        sheetNameEntry.bind('<Escape>', renameSheetCancel)
        sheetNameEntry.place(x=x, y=y, width=width, height=height)
        
        
    def bindSheet(self, sheetNum=0, specifiedSheet=None):
        if specifiedSheet is None:
            specifiedSheet = self.sheets[sheetNum]

        # Moving across cells using keyboard arrows, shift, tab, and return
        self.window.bind('<Key>', lambda event: specifiedSheet.keyHandle(event=event) )
        self.window.bind('<KeyRelease>', lambda event: specifiedSheet.keyReleaseHandle(event=event) )
        
        specifiedSheet.cellFunctions = [

        ('<Right>', lambda event: specifiedSheet.MoveMarkByCoord(+1, 0) ),
        ('<Left>', lambda event: specifiedSheet.MoveMarkByCoord(-1, 0) ),
        ('<Up>', lambda event: specifiedSheet.MoveMarkByCoord(0, -1) ),
        ('<Down>', lambda event: specifiedSheet.MoveMarkByCoord(0, +1) ),

        ('<Tab>', lambda event: specifiedSheet.MoveMarkByCoord(+1, 0) ),
        ('<Shift-Tab>', lambda event: specifiedSheet.MoveMarkByCoord(-1, 0) ),
        ('<Shift-Return>', lambda event: specifiedSheet.MoveMarkByCoord(0, -1) ),
        ('<Return>', lambda event: specifiedSheet.MoveMarkByCoord(0, +1) ),
        
        ('<F2>', lambda event: specifiedSheet.currentCell.edit_start(event=event) ),
        
        ('<Key>', lambda event: specifiedSheet.keyHandle(event=event) ),
        
        ]
        
        self.cellFunctions = specifiedSheet.cellFunctions

    def BookToDataFrame(self, toDataFrame=True, event=None):
        """Convert Sheets to dicts or dataframes"""
        return {sheet.sheetName:sheet.SheetsToDataFrame(toDataFrame=toDataFrame) for sheet in self.sheets}
            
    def experimental_warning(self, window):
        local_app_data_dir = os.getenv('LOCALAPPDATA')
        qu_table_folder = os.path.join(local_app_data_dir, "QuTable")
        qu_table_experimental_warning_file = os.path.join(qu_table_folder, "QuTable-Experimental-Warning.txt")
        self.qu_table_folder = qu_table_folder
        self.qu_table_experimental_warning_file = qu_table_experimental_warning_file
        os.makedirs(qu_table_folder, exist_ok=True)
        
        qu_table_experimental_warning_file_cont = open(qu_table_experimental_warning_file).read() if os.path.isfile(qu_table_experimental_warning_file) else None
        
        self.experimental_warning_text = '''WARNING!

Please read the Warning carefully below:

Quantum Apps QuTable is experimental and still under developement, so be prepared that QuTable does not function as expected. As a result, this has NOT been intended for real-world uses. (If you would like to use it, it is SOLELY your responsibility for each and every step of your actions.) Furthermore, please be prepared to back up your data prior to using QuTable. (Please kindly remember that this is one of the facets that made it possible for this version of the software to be free of charge.)

Pardon for any incoveniences because my team might or might not be considered as a team. The "team" is only a "team" of one person, which is me and only me. Therefore, developments might be real slow.

Furthermore, if any parts of the QuTable is depicted slightly differently in the Documentation (https://quantumapps100.wixsite.com/quantumapps) from the real appearance of the software, it is perhaps caused by screenshot rendering, so please kindly be patient with that.

(Note that this Warning would NOT appear again as long as a file related to your affirmation to this warning is still written in a file having a location of "%LOCALAPPDATA%/QuTable/QuTable-Experimental-Warning.txt")

Would you like to proceed to use QuTable?'''
        
        self.experimental_warning_file_cont = f'When you (user or users) first opened QuTable, you have HEEDED to this following Experimental/Under Development Warning below by clicking the "OK" Button in the Warning. Here is the Experimental/Under Development Warning below:\n\n{self.experimental_warning_text}'
        
        # Rename if that same path is not a file
        if os.path.exists(qu_table_experimental_warning_file):
            if not os.path.isfile(qu_table_experimental_warning_file):
                file_addr, extension = os.path.splitext(qu_table_experimental_warning_file)
                extension = extension.lower()
                i = 1
                while 1:
                    newFileAddr = qu_table_experimental_warning_file.replace(file_addr, f'{file_addr}-{i}')
                    if not os.path.exists(newFileAddr):
                        break
                    i += 1
                os.rename(qu_table_experimental_warning_file, newFileAddr)
        
        if not andExec(os.path.isfile(qu_table_experimental_warning_file), qu_table_experimental_warning_file_cont == self.experimental_warning_file_cont):
            heed_warning = messagebox.askyesno('Experimental/Under Development Warning!', self.experimental_warning_text, parent=window)
            
            # print(9485, heed_warning)
            if heed_warning:
                os.makedirs(qu_table_folder, exist_ok=True)
                warningIO = open(qu_table_experimental_warning_file, 'w+')
                warningIO.write(self.experimental_warning_file_cont)
                warningIO.close()
            else:
                self.close_all()
        

class QuTable:
    def __init__(self):
        self.booksDict = {}
        self.booksList = []
        Book(app=self, activate=True, experimental=experimentalOverall)
        # try:
            # loadingApp.destroy()
        # except tk.TclError:
            # pass
        
    def __getitem__(self, index):
        if type(index) == int:
            return self.booksList[index]
            
        elif type(index) == range:
            slice_range = slice(index.start, index.stop, index.step)
            return self.booksList[slice_range]
            
        elif type(index) == str:
            return super().__getitem__(index)
            
        elif hasattr(index, '__iter__'):
            if len(index) == 1:
                return self.booksList[index[0]]
            elif len(index) == 2:
                return self.booksList[index[0]][index[1]]
            elif len(index) == 3:
                return self.booksList[index[0]][index[1]][index[2]]
            elif len(index) >= 4:
                return self.booksList[index[0]][index[1]][index[2]].cells[index[3]]
            
        elif type(index) == slice:
            return self.booksList[index]
        
        elif orExec(isinstance(index, ColRange), isinstance(index[1], RowRange)):
            return self.booksList[index.slicer]
        
    def __iter__(self):
        for book in self.booksList:
            yield book

    def __lt__(self, other):
        return self.booksList < other.booksList

    def __gt__(self, other):
        return self.booksList > other.booksList

    # def __eq__(self, other):
        # return self.booksList == other.booksList

    def __le__(self, other):
        return self.booksList <= other.booksList

    def __ge__(self, other):
        return self.booksList >= other.booksList
        
    def __len__(self):
        return len(self.booksList)

try:
    loadingApp = LoadingApp()
    if getattr(sys, 'frozen', False):
        loadingApp.mainloop()
except:
    print(format_exc(), file=sys.stderr)